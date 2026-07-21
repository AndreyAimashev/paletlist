#!/usr/bin/env python3
import datetime
import hashlib
import io
import json
import math
import os
import re
import secrets
import shutil
import socket
import sqlite3
import subprocess
import tempfile
import threading
import time
import unicodedata
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

try:
    from fpdf import FPDF
    from fpdf.enums import Align, WrapMode

    HAVE_FPDF = True
except ImportError:
    HAVE_FPDF = False
    FPDF = None  # type: ignore
    Align = None  # type: ignore
    WrapMode = None  # type: ignore

try:
    from barcode import Code128
    from barcode.codex import Gs1_128
    from barcode.writer import ImageWriter

    HAVE_CODE128_BARCODE = True
    HAVE_GS1_128_BARCODE = True
except ImportError:
    HAVE_CODE128_BARCODE = False
    HAVE_GS1_128_BARCODE = False
    Code128 = None  # type: ignore
    Gs1_128 = None  # type: ignore
    ImageWriter = None  # type: ignore

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False
    Workbook = None  # type: ignore
    load_workbook = None  # type: ignore
    Alignment = None  # type: ignore
    Border = None  # type: ignore
    Font = None  # type: ignore
    PatternFill = None  # type: ignore
    Side = None  # type: ignore
    get_column_letter = None  # type: ignore


BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "warehouse.db"
JSON_SEED_PATH = BASE_DIR / "nomenclature.json"
UPDATES_JSON_PATH = BASE_DIR / "updates.json"
HOST = "127.0.0.1"

try:
    from packing_sheets_generic import (
        build_generic_packing_sheets_html,
        build_generic_packing_sheets_pdf_bytes,
        generic_packing_error_message,
        sort_assemble_pallets_by_number,
    )
except ImportError:
    build_generic_packing_sheets_html = None  # type: ignore[misc, assignment]
    build_generic_packing_sheets_pdf_bytes = None  # type: ignore[misc, assignment]
    generic_packing_error_message = None  # type: ignore[misc, assignment]
    sort_assemble_pallets_by_number = None  # type: ignore[misc, assignment]

PORT = 8081
DB_LOCK = threading.Lock()
PRESENCE_LOCK = threading.Lock()
ASSEMBLE_PRESENCE_TTL_SEC = 15
_ASSEMBLE_PRESENCE: dict[int, dict[str, dict]] = {}


def _norm_api_path(path: str) -> str:
    p = (path or "").strip()
    while "//" in p:
        p = p.replace("//", "/")
    p = p.rstrip("/")
    return p if p else "/"


def _is_nomenclature_list_path(path: str) -> bool:
    return _norm_api_path(path) == "/api/nomenclature"


def _is_users_list_path(path: str) -> bool:
    return _norm_api_path(path) == "/api/users"


ADMIN_LOGIN = "admin"
_LOGIN_RE = re.compile(r"^[a-zA-Z0-9._-]{2,64}$")
_PUBLIC_API_PATHS = frozenset({"/api/auth/login"})
_AUTH_SESSION_TTL_SEC = 24 * 3600
_AUTH_SESSIONS: dict[str, dict] = {}
_AUTH_SESSIONS_LOCK = threading.Lock()
_LOGIN_GUARD_LOCK = threading.Lock()
_LOGIN_FAIL_LIMIT = 5
_LOGIN_BAN_SECONDS = 24 * 3600
_LOGIN_GUARD_PATH = BASE_DIR / "login_guard.json"


def _admin_password() -> str:
    return os.environ.get("PALETLIST_ADMIN_PASSWORD", "1999")


def _deploy_secret() -> str:
    return (os.environ.get("PALETLIST_DEPLOY_SECRET") or "").strip()


def _ssh_allow_secret() -> str:
    # Отдельный секрет предпочтителен; иначе тот же, что для деплоя.
    raw = (os.environ.get("PALETLIST_SSH_ALLOW_SECRET") or "").strip()
    return raw or _deploy_secret()


def _is_public_api_path(path: str) -> bool:
    return _norm_api_path(path) in _PUBLIC_API_PATHS


def _is_auth_me_path(path: str) -> bool:
    return _norm_api_path(path) == "/api/auth/me"


def _is_auth_login_path(path: str) -> bool:
    return _norm_api_path(path) == "/api/auth/login"


def _is_auth_account_path(path: str) -> bool:
    return _norm_api_path(path) == "/api/auth/account"


def _is_updates_path(path: str) -> bool:
    return _norm_api_path(path) == "/api/updates"


def _is_database_backup_path(path: str) -> bool:
    return _norm_api_path(path) == "/api/database/backup"


def _is_database_restore_path(path: str) -> bool:
    return _norm_api_path(path) == "/api/database/restore"


def _is_server_status_path(path: str) -> bool:
    return _norm_api_path(path) == "/api/server-status"


def _is_deploy_path(path: str) -> bool:
    return _norm_api_path(path) == "/api/deploy"


def _is_ssh_allow_path(path: str) -> bool:
    return _norm_api_path(path) == "/api/ssh-allow"


def _is_unique_clients_list_path(path: str) -> bool:
    return _norm_api_path(path) == "/api/unique-clients"


def _is_feedback_threads_list_path(path: str) -> bool:
    return _norm_api_path(path) == "/api/feedback/threads"


def _parse_feedback_thread_detail_id(path: str) -> int | None:
    norm = _norm_api_path(path)
    prefix = "/api/feedback/threads/"
    if not norm.startswith(prefix):
        return None
    tail = norm[len(prefix) :]
    if not tail or "/" in tail:
        return None
    try:
        return int(tail)
    except ValueError:
        return None


def _parse_feedback_thread_messages_id(path: str) -> int | None:
    norm = _norm_api_path(path)
    suffix = "/messages"
    if not norm.startswith("/api/feedback/threads/") or not norm.endswith(suffix):
        return None
    middle = norm[len("/api/feedback/threads/") : -len(suffix)]
    if not middle or "/" in middle:
        return None
    try:
        return int(middle)
    except ValueError:
        return None


def authenticate_app_login(login: str, password: str) -> dict | None:
    login_norm = (login or "").strip()
    pwd = str(password or "")
    if not login_norm or not pwd:
        return None
    if _is_reserved_admin_login(login_norm):
        if pwd == _admin_password():
            return {
                "user_id": 0,
                "login": ADMIN_LOGIN,
                "display_name": "Администратор",
                "is_admin": True,
                "permissions": _admin_permissions(),
            }
        return None
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        row = cur.execute(
            """
            SELECT id, login, password_hash, display_name, department,
                   perm_orders, perm_nomenclature, perm_manage_users, perm_feedback,
                   perm_order_monitoring
            FROM app_users
            WHERE login = ? COLLATE NOCASE
            """,
            (login_norm,),
        ).fetchone()
        con.close()
    if not row:
        return None
    if not _verify_app_password(pwd, row["password_hash"]):
        return None
    return {
        "user_id": int(row["id"]),
        "login": row["login"] or login_norm,
        "display_name": row["display_name"] or "",
        "department": (row["department"] or "") if "department" in row.keys() else "",
        "is_admin": False,
        "permissions": _permissions_from_row(row),
    }


def create_auth_session(user: dict) -> str:
    token = secrets.token_urlsafe(32)
    expires_at = time.time() + _AUTH_SESSION_TTL_SEC
    perms = user.get("permissions")
    if not isinstance(perms, dict):
        perms = _admin_permissions() if user.get("is_admin") else _permissions_from_row({})
    payload = {
        "user_id": int(user.get("user_id", 0)),
        "login": user.get("login") or "",
        "display_name": user.get("display_name") or "",
        "department": user.get("department") or "",
        "is_admin": bool(user.get("is_admin")),
        "permissions": dict(perms),
        "expires_at": expires_at,
    }
    with _AUTH_SESSIONS_LOCK:
        _AUTH_SESSIONS[token] = payload
    return token


def _login_guard_load() -> dict:
    if not _LOGIN_GUARD_PATH.is_file():
        return {"ips": {}, "logins": {}}
    try:
        data = json.loads(_LOGIN_GUARD_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {"ips": {}, "logins": {}}
    if not isinstance(data, dict):
        return {"ips": {}, "logins": {}}
    ips = data.get("ips") if isinstance(data.get("ips"), dict) else {}
    logins = data.get("logins") if isinstance(data.get("logins"), dict) else {}
    return {"ips": ips, "logins": logins}


def _login_guard_save(data: dict) -> None:
    tmp = _LOGIN_GUARD_PATH.with_suffix(".json.tmp")
    payload = json.dumps(data, ensure_ascii=False, indent=2)
    tmp.write_text(payload, encoding="utf-8")
    os.replace(tmp, _LOGIN_GUARD_PATH)


def _login_guard_entry(bucket: dict, key: str) -> dict:
    raw = bucket.get(key)
    if not isinstance(raw, dict):
        return {"fails": 0, "banned_until": 0}
    try:
        fails = int(raw.get("fails") or 0)
    except (TypeError, ValueError):
        fails = 0
    try:
        banned_until = float(raw.get("banned_until") or 0)
    except (TypeError, ValueError):
        banned_until = 0.0
    return {"fails": max(0, fails), "banned_until": max(0.0, banned_until)}


def _login_guard_prune(bucket: dict, now: float) -> None:
    stale = []
    for key, raw in list(bucket.items()):
        entry = _login_guard_entry(bucket, key)
        if entry["banned_until"] and entry["banned_until"] <= now:
            stale.append(key)
        elif entry["fails"] <= 0 and not entry["banned_until"]:
            stale.append(key)
    for key in stale:
        bucket.pop(key, None)


def login_guard_status(ip: str, login: str) -> dict | None:
    """Если IP или логин в бане — вернуть описание блокировки."""
    ip_key = (ip or "").strip()
    login_key = (login or "").strip().casefold()
    now = time.time()
    with _LOGIN_GUARD_LOCK:
        data = _login_guard_load()
        _login_guard_prune(data["ips"], now)
        _login_guard_prune(data["logins"], now)
        checks = []
        if ip_key:
            checks.append(("ip", ip_key, _login_guard_entry(data["ips"], ip_key)))
        if login_key:
            checks.append(
                ("login", login_key, _login_guard_entry(data["logins"], login_key))
            )
        for kind, key, entry in checks:
            until = float(entry.get("banned_until") or 0)
            if until > now:
                left = int(until - now)
                hours = max(1, (left + 3599) // 3600)
                return {
                    "error": "banned",
                    "message": (
                        "Слишком много неверных попыток входа. "
                        f"Доступ заблокирован примерно на {hours} ч."
                    ),
                    "banned_until": int(until),
                    "scope": kind,
                    "key": key,
                }
        _login_guard_save(data)
    return None


def login_guard_register_failure(ip: str, login: str) -> dict | None:
    """Учесть неудачный вход. При 5 ошибках — бан на сутки (IP и/или логин)."""
    ip_key = (ip or "").strip()
    login_key = (login or "").strip().casefold()
    now = time.time()
    banned = None
    with _LOGIN_GUARD_LOCK:
        data = _login_guard_load()
        _login_guard_prune(data["ips"], now)
        _login_guard_prune(data["logins"], now)

        def _bump(bucket: dict, key: str, kind: str) -> None:
            nonlocal banned
            if not key:
                return
            entry = _login_guard_entry(bucket, key)
            if entry["banned_until"] > now:
                banned = {
                    "error": "banned",
                    "message": "Слишком много неверных попыток входа. Доступ заблокирован на сутки.",
                    "banned_until": int(entry["banned_until"]),
                    "scope": kind,
                }
                return
            entry["fails"] = int(entry["fails"]) + 1
            if entry["fails"] >= _LOGIN_FAIL_LIMIT:
                entry["banned_until"] = now + _LOGIN_BAN_SECONDS
                entry["fails"] = _LOGIN_FAIL_LIMIT
                banned = {
                    "error": "banned",
                    "message": (
                        "Слишком много неверных попыток входа. "
                        "Доступ заблокирован на 24 часа."
                    ),
                    "banned_until": int(entry["banned_until"]),
                    "scope": kind,
                }
            bucket[key] = entry

        _bump(data["ips"], ip_key, "ip")
        _bump(data["logins"], login_key, "login")
        _login_guard_save(data)
    return banned


def login_guard_clear_success(ip: str, login: str) -> None:
    """После успешного входа сбросить счётчики для IP и логина."""
    ip_key = (ip or "").strip()
    login_key = (login or "").strip().casefold()
    with _LOGIN_GUARD_LOCK:
        data = _login_guard_load()
        if ip_key:
            data["ips"].pop(ip_key, None)
        if login_key:
            data["logins"].pop(login_key, None)
        _login_guard_save(data)


def resolve_auth_session(token: str | None) -> dict | None:
    if not token:
        return None
    now = time.time()
    with _AUTH_SESSIONS_LOCK:
        session = _AUTH_SESSIONS.get(token)
        if not session:
            return None
        if now >= float(session.get("expires_at") or 0):
            _AUTH_SESSIONS.pop(token, None)
            return None
        return dict(session)


def revoke_auth_session(token: str | None) -> None:
    if not token:
        return
    with _AUTH_SESSIONS_LOCK:
        _AUTH_SESSIONS.pop(token, None)


def _auth_display_name(session: dict | None) -> str:
    if not session:
        return ""
    display_name = _normalize_str(session.get("display_name") or "")
    if display_name:
        return display_name
    login = _normalize_str(session.get("login") or "")
    if login:
        return login
    if session.get("is_admin"):
        return "Администратор"
    return ""


def update_auth_account(session: dict, body: dict, token: str | None) -> dict:
    """Смена логина/пароля текущим пользователем (не admin)."""
    if session.get("is_admin"):
        return {
            "error": "forbidden",
            "message": "Для администратора используйте раздел «Управление пользователями».",
        }
    session_perms = session.get("permissions") or {}
    if session_perms.get("manage_users"):
        return {
            "error": "forbidden",
            "message": "Используйте раздел «Управление пользователями» для смены логина и пароля.",
        }
    user_id = int(session.get("user_id") or 0)
    if user_id <= 0:
        return {"error": "forbidden", "message": "Учётная запись не найдена."}
    new_login = _normalize_str(body.get("login", session.get("login") or ""))
    if not new_login:
        return {"error": "validation", "message": "Укажите логин."}
    if _is_reserved_admin_login(new_login):
        return {
            "error": "validation",
            "message": "Логин «admin» зарезервирован для администратора.",
        }
    if not _LOGIN_RE.match(new_login):
        return {
            "error": "validation",
            "message": "Логин: 2–64 символа (латиница, цифры, . _ -).",
        }
    password_raw = body.get("password")
    new_password = (
        str(password_raw).strip() if password_raw is not None else ""
    )
    department_in_body = "department" in body
    department = (
        _normalize_str(body.get("department", "")) if department_in_body else None
    )
    if department_in_body and len(department or "") > 128:
        return {
            "error": "validation",
            "message": "Подразделение не длиннее 128 символов.",
        }
    if new_password and len(new_password) < 4:
        return {
            "error": "validation",
            "message": "Новый пароль не короче 4 символов.",
        }
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        row = cur.execute(
            """
            SELECT id, login, password_hash, display_name
            FROM app_users WHERE id = ?
            """,
            (user_id,),
        ).fetchone()
        if not row:
            con.close()
            return {"error": "not_found", "message": "Пользователь не найден."}
        password_unchanged = False
        login_changed = (new_login or "").lower() != (row["login"] or "").lower()
        try:
            if new_password:
                if _verify_app_password(new_password, row["password_hash"]):
                    password_unchanged = True
                    cur.execute(
                        "UPDATE app_users SET login = ? WHERE id = ?",
                        (new_login, user_id),
                    )
                else:
                    pwd_hash = _hash_app_password(new_password)
                    cur.execute(
                        """
                        UPDATE app_users
                        SET login = ?, password_hash = ?
                        WHERE id = ?
                        """,
                        (new_login, pwd_hash, user_id),
                    )
            else:
                cur.execute(
                    "UPDATE app_users SET login = ? WHERE id = ?",
                    (new_login, user_id),
                )
            if department_in_body:
                cur.execute(
                    "UPDATE app_users SET department = ? WHERE id = ?",
                    (department or "", user_id),
                )
            row = cur.execute(
                f"{_APP_USER_SELECT} WHERE id = ?",
                (user_id,),
            ).fetchone()
            con.commit()
        except sqlite3.IntegrityError:
            con.close()
            return {
                "error": "duplicate",
                "message": "Пользователь с таким логином уже существует.",
            }
        con.close()
    user_api = _user_row_to_api(row)
    if token:
        with _AUTH_SESSIONS_LOCK:
            stored = _AUTH_SESSIONS.get(token)
            if stored:
                stored["login"] = user_api["login"]
                stored["display_name"] = user_api.get("display_name") or ""
                stored["department"] = user_api.get("department") or ""
    if department_in_body:
        sync_auth_sessions_department_for_user(
            user_id, user_api.get("department") or ""
        )
    if password_unchanged and not login_changed and not department_in_body:
        msg = "Пароль совпадает с текущим — изменений не было."
    elif password_unchanged and login_changed:
        msg = "Логин обновлён. Пароль совпадает с текущим — он не менялся."
    elif password_unchanged and department_in_body:
        msg = "Подразделение сохранено."
    else:
        msg = "Данные аккаунта сохранены."
    return {
        "ok": True,
        "user": user_api,
        "password_unchanged": password_unchanged,
        "message": msg,
    }


def _is_reserved_admin_login(login: str) -> bool:
    return (login or "").strip().lower() == ADMIN_LOGIN


def _hash_app_password(password: str) -> str:
    salt = secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac(
        "sha256", password.encode("utf-8"), bytes.fromhex(salt), 120_000
    )
    return f"pbkdf2:{salt}:{digest.hex()}"


def _verify_app_password(password: str, stored: str) -> bool:
    parts = str(stored or "").split(":")
    if len(parts) != 3 or parts[0] != "pbkdf2":
        return False
    try:
        salt = bytes.fromhex(parts[1])
        expected = bytes.fromhex(parts[2])
    except ValueError:
        return False
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 120_000)
    return secrets.compare_digest(digest, expected)


def _init_app_users_table(cur) -> None:
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS app_users (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          login TEXT NOT NULL COLLATE NOCASE UNIQUE,
          password_hash TEXT NOT NULL DEFAULT '',
          display_name TEXT NOT NULL DEFAULT '',
          created_at TEXT NOT NULL DEFAULT ''
        )
        """
    )
    _migrate_app_users_permissions(cur)


def _migrate_app_users_permissions(cur) -> None:
    cols = {r[1] for r in cur.execute("PRAGMA table_info(app_users)").fetchall()}
    for col, default in (
        ("perm_orders", 1),
        ("perm_nomenclature", 1),
        ("perm_manage_users", 0),
        ("perm_feedback", 1),
        ("perm_order_monitoring", 0),
    ):
        if col not in cols:
            cur.execute(
                f"ALTER TABLE app_users ADD COLUMN {col} INTEGER NOT NULL DEFAULT {default}"
            )
    if "department" not in cols:
        cur.execute(
            "ALTER TABLE app_users ADD COLUMN department TEXT NOT NULL DEFAULT ''"
        )


def _perm_cell_bool(row, col: str, default: int) -> bool:
    keys = row.keys() if hasattr(row, "keys") else ()
    if col not in keys:
        return bool(default)
    raw = row[col]
    if raw is None:
        return bool(default)
    try:
        return bool(int(raw))
    except (TypeError, ValueError):
        return bool(default)


def _permissions_from_row(row) -> dict:
    return {
        "orders": _perm_cell_bool(row, "perm_orders", 1),
        "nomenclature": _perm_cell_bool(row, "perm_nomenclature", 1),
        "manage_users": _perm_cell_bool(row, "perm_manage_users", 0),
        "feedback": _perm_cell_bool(row, "perm_feedback", 1),
        "order_monitoring": _perm_cell_bool(row, "perm_order_monitoring", 0),
    }


def fetch_app_user_permissions(user_id: int) -> dict | None:
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        row = cur.execute(
            f"{_APP_USER_SELECT} WHERE id = ?",
            (int(user_id),),
        ).fetchone()
        con.close()
    if not row:
        return None
    return _permissions_from_row(row)


def sync_auth_sessions_department_for_user(user_id: int, department: str) -> None:
    uid = int(user_id)
    dept = _normalize_str(department)
    with _AUTH_SESSIONS_LOCK:
        for session in _AUTH_SESSIONS.values():
            if int(session.get("user_id") or 0) == uid and not session.get("is_admin"):
                session["department"] = dept


def sync_auth_sessions_permissions_for_user(user_id: int, permissions: dict) -> None:
    uid = int(user_id)
    perms = {
        "orders": bool(permissions.get("orders")),
        "nomenclature": bool(permissions.get("nomenclature")),
        "manage_users": bool(permissions.get("manage_users")),
        "feedback": bool(permissions.get("feedback")),
        "order_monitoring": bool(permissions.get("order_monitoring")),
    }
    with _AUTH_SESSIONS_LOCK:
        for session in _AUTH_SESSIONS.values():
            if int(session.get("user_id") or 0) == uid and not session.get("is_admin"):
                session["permissions"] = dict(perms)


def _admin_permissions() -> dict:
    return {
        "orders": True,
        "nomenclature": True,
        "manage_users": True,
        "feedback": True,
        "order_monitoring": True,
    }


def _parse_permissions_patch(body: dict) -> dict | None:
    raw = body.get("permissions")
    if raw is None:
        return None
    if not isinstance(raw, dict):
        return {}
    out = {}
    for key, col in (
        ("orders", "perm_orders"),
        ("nomenclature", "perm_nomenclature"),
        ("manage_users", "perm_manage_users"),
        ("feedback", "perm_feedback"),
        ("order_monitoring", "perm_order_monitoring"),
    ):
        if key in raw:
            out[col] = 1 if raw[key] else 0
    return out


def _user_row_to_api(row) -> dict:
    return {
        "id": int(row["id"]),
        "login": row["login"] or "",
        "display_name": row["display_name"] or "",
        "department": (row["department"] or "")
        if "department" in row.keys()
        else "",
        "created_at": row["created_at"] or "",
        "permissions": _permissions_from_row(row),
    }


_APP_USER_SELECT = """
    SELECT id, login, display_name, department, created_at,
           perm_orders, perm_nomenclature, perm_manage_users, perm_feedback,
           perm_order_monitoring
    FROM app_users
"""


def fetch_app_user_department_options() -> list[str]:
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        rows = cur.execute(
            """
            SELECT DISTINCT trim(department) AS department
            FROM app_users
            WHERE trim(department) != ''
            ORDER BY department COLLATE NOCASE ASC
            """
        ).fetchall()
        con.close()
    return [str(r["department"] or "").strip() for r in rows if str(r["department"] or "").strip()]


def fetch_app_user_department(user_id: int) -> str:
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        row = cur.execute(
            "SELECT department FROM app_users WHERE id = ?",
            (int(user_id),),
        ).fetchone()
        con.close()
    if not row:
        return ""
    return _normalize_str(row["department"] or "")


def fetch_app_users() -> list[dict]:
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        rows = cur.execute(
            f"""
            {_APP_USER_SELECT}
            WHERE lower(trim(login)) != lower(?)
            ORDER BY login COLLATE NOCASE ASC
            """,
            (ADMIN_LOGIN,),
        ).fetchall()
        con.close()
    return [_user_row_to_api(r) for r in rows]


def create_app_user(body: dict) -> dict:
    login = _normalize_str(body.get("login", ""))
    password = str(body.get("password") or "")
    display_name = _normalize_str(body.get("display_name", ""))
    department = _normalize_str(body.get("department", ""))
    if len(department) > 128:
        return {
            "error": "validation",
            "message": "Подразделение не длиннее 128 символов.",
        }
    if not login:
        return {"error": "validation", "message": "Укажите логин."}
    if _is_reserved_admin_login(login):
        return {
            "error": "validation",
            "message": "Логин «admin» зарезервирован для администратора.",
        }
    if not _LOGIN_RE.match(login):
        return {
            "error": "validation",
            "message": "Логин: 2–64 символа (латиница, цифры, . _ -).",
        }
    if len(password) < 4:
        return {
            "error": "validation",
            "message": "Пароль не короче 4 символов.",
        }
    created_at = datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    pwd_hash = _hash_app_password(password)
    perm_patch = _parse_permissions_patch(body) or {}
    perm_orders = perm_patch.get("perm_orders", 1)
    perm_nomenclature = perm_patch.get("perm_nomenclature", 1)
    perm_manage_users = perm_patch.get("perm_manage_users", 0)
    perm_feedback = perm_patch.get("perm_feedback", 1)
    perm_order_monitoring = perm_patch.get("perm_order_monitoring", 0)
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        try:
            cur.execute(
                """
                INSERT INTO app_users (
                  login, password_hash, display_name, department, created_at,
                  perm_orders, perm_nomenclature, perm_manage_users, perm_feedback,
                  perm_order_monitoring
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    login,
                    pwd_hash,
                    display_name,
                    department,
                    created_at,
                    perm_orders,
                    perm_nomenclature,
                    perm_manage_users,
                    perm_feedback,
                    perm_order_monitoring,
                ),
            )
            uid = int(cur.lastrowid)
            row = cur.execute(
                f"{_APP_USER_SELECT} WHERE id = ?",
                (uid,),
            ).fetchone()
            con.commit()
        except sqlite3.IntegrityError:
            con.close()
            return {
                "error": "duplicate",
                "message": "Пользователь с таким логином уже существует.",
            }
        con.close()
    return {"ok": True, "user": _user_row_to_api(row)}


def update_app_user(user_id: int, body: dict) -> dict:
    perm_patch = _parse_permissions_patch(body)
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        row = cur.execute(
            f"{_APP_USER_SELECT} WHERE id = ?",
            (int(user_id),),
        ).fetchone()
        if not row:
            con.close()
            return {"error": "not_found", "message": "Пользователь не найден."}
        if _is_reserved_admin_login(row["login"]):
            con.close()
            return {
                "error": "forbidden",
                "message": "Учётную запись администратора нельзя изменить здесь.",
            }
        login = _normalize_str(body.get("login", row["login"]))
        display_name = _normalize_str(
            body.get("display_name", row["display_name"] or "")
        )
        department_in_body = "department" in body
        department = (
            _normalize_str(body.get("department", ""))
            if department_in_body
            else (row["department"] or "" if "department" in row.keys() else "")
        )
        if department_in_body and len(department) > 128:
            con.close()
            return {
                "error": "validation",
                "message": "Подразделение не длиннее 128 символов.",
            }
        password = body.get("password")
        if _is_reserved_admin_login(login):
            con.close()
            return {
                "error": "validation",
                "message": "Логин «admin» зарезервирован для администратора.",
            }
        if login and not _LOGIN_RE.match(login):
            con.close()
            return {
                "error": "validation",
                "message": "Логин: 2–64 символа (латиница, цифры, . _ -).",
            }
        if password is not None and str(password) != "" and len(str(password)) < 4:
            con.close()
            return {
                "error": "validation",
                "message": "Пароль не короче 4 символов.",
            }
        pwd_hash = None
        if password is not None and str(password) != "":
            pwd_hash = _hash_app_password(str(password))
        only_permissions = (
            perm_patch is not None
            and len(perm_patch) > 0
            and "login" not in body
            and "display_name" not in body
            and "password" not in body
            and not department_in_body
        )
        only_department = (
            department_in_body
            and "login" not in body
            and "display_name" not in body
            and "password" not in body
            and (perm_patch is None or len(perm_patch) == 0)
        )
        try:
            if only_department:
                cur.execute(
                    "UPDATE app_users SET department = ? WHERE id = ?",
                    (department, int(user_id)),
                )
            elif only_permissions and perm_patch:
                sets = ", ".join(f"{col} = ?" for col in perm_patch)
                cur.execute(
                    f"UPDATE app_users SET {sets} WHERE id = ?",
                    (*perm_patch.values(), int(user_id)),
                )
            elif pwd_hash is not None:
                cur.execute(
                    """
                    UPDATE app_users
                    SET login = ?, display_name = ?, password_hash = ?
                    WHERE id = ?
                    """,
                    (login, display_name, pwd_hash, int(user_id)),
                )
                if department_in_body:
                    cur.execute(
                        "UPDATE app_users SET department = ? WHERE id = ?",
                        (department, int(user_id)),
                    )
                if perm_patch:
                    sets = ", ".join(f"{col} = ?" for col in perm_patch)
                    cur.execute(
                        f"UPDATE app_users SET {sets} WHERE id = ?",
                        (*perm_patch.values(), int(user_id)),
                    )
            else:
                cur.execute(
                    """
                    UPDATE app_users SET login = ?, display_name = ? WHERE id = ?
                    """,
                    (login, display_name, int(user_id)),
                )
                if department_in_body:
                    cur.execute(
                        "UPDATE app_users SET department = ? WHERE id = ?",
                        (department, int(user_id)),
                    )
                if perm_patch:
                    sets = ", ".join(f"{col} = ?" for col in perm_patch)
                    cur.execute(
                        f"UPDATE app_users SET {sets} WHERE id = ?",
                        (*perm_patch.values(), int(user_id)),
                    )
            row = cur.execute(
                f"{_APP_USER_SELECT} WHERE id = ?",
                (int(user_id),),
            ).fetchone()
            con.commit()
        except sqlite3.IntegrityError:
            con.close()
            return {
                "error": "duplicate",
                "message": "Пользователь с таким логином уже существует.",
            }
        con.close()
    user_api = _user_row_to_api(row)
    if perm_patch:
        sync_auth_sessions_permissions_for_user(
            int(user_id), user_api.get("permissions") or {}
        )
    if department_in_body:
        sync_auth_sessions_department_for_user(
            int(user_id), user_api.get("department") or ""
        )
    return {"ok": True, "user": user_api}


def delete_app_user(user_id: int) -> dict:
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        row = cur.execute(
            "SELECT login FROM app_users WHERE id = ?",
            (int(user_id),),
        ).fetchone()
        if not row:
            con.close()
            return {"error": "not_found", "message": "Пользователь не найден."}
        if _is_reserved_admin_login(row["login"]):
            con.close()
            return {
                "error": "forbidden",
                "message": "Учётную запись администратора нельзя удалить.",
            }
        cur.execute("DELETE FROM app_users WHERE id = ?", (int(user_id),))
        con.commit()
        con.close()
    return {"ok": True}


def _parse_user_id_path(path: str) -> int | None:
    prefix = "/api/users/"
    norm = _norm_api_path(path)
    if not norm.startswith(prefix):
        return None
    tail = norm[len(prefix) :].strip("/")
    if not tail or "/" in tail:
        return None
    try:
        return int(tail)
    except ValueError:
        return None


def _normalize_nomenclature_search_text(value: str) -> str:
    """Единый вид строки для поиска: регистр, пробелы, №/No, латинская x и кириллическая х."""
    t = unicodedata.normalize("NFKC", str(value or "")).lower().replace("\u00a0", " ")
    t = t.replace("№", "no")
    t = re.sub(r"[xх]", "х", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _normalize_product_name_for_import_match(value: str) -> str:
    """Ключ для сопоставления наименования из Excel с products.name (без нечёткого подбора)."""
    t = unicodedata.normalize("NFKC", str(value or "")).replace("\u00a0", " ")
    t = _strip_trailing_name_suffix(t)
    t = re.sub(r"№", "No", t, flags=re.IGNORECASE)
    t = re.sub(r"([0-9A-Za-zА-Яа-яЁё])(\()", r"\1 \2", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t.casefold()


def _nomenclature_matches_query(item: dict, query: str) -> bool:
    qn = _normalize_nomenclature_search_text(query)
    if not qn:
        return True
    art = _normalize_nomenclature_search_text(item.get("article", ""))
    name = _normalize_nomenclature_search_text(item.get("name", ""))
    return qn in art or qn in name


def _levenshtein(a: str, b: str) -> int:
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, start=1):
        cur = [i]
        for j, cb in enumerate(b, start=1):
            cur.append(
                min(prev[j] + 1, cur[j - 1] + 1, prev[j - 1] + (0 if ca == cb else 1))
            )
        prev = cur
    return prev[-1]


def _nomenclature_distance_to_item(query_norm: str, item: dict) -> int:
    art = _normalize_nomenclature_search_text(item.get("article", ""))
    name = _normalize_nomenclature_search_text(item.get("name", ""))
    best = 10**9
    if art:
        best = min(best, _levenshtein(query_norm, art))
    if name:
        best = min(best, _levenshtein(query_norm, name))
        for word in re.findall(r"\S{2,}", name):
            best = min(best, _levenshtein(query_norm, word))
    return best


def search_nomenclature(items: list, query: str, limit: int = 5) -> list:
    """Топ limit позиций по релевантности (подстрока, префикс, расстояние Левенштейна)."""
    qn = _normalize_nomenclature_search_text(query)
    if not qn or not items:
        return []
    limit = max(1, min(50, int(limit)))
    scored: list[tuple[int, dict]] = []
    for item in items:
        art = _normalize_nomenclature_search_text(item.get("article", ""))
        name = _normalize_nomenclature_search_text(item.get("name", ""))
        dist = _nomenclature_distance_to_item(qn, item)
        if qn in art or qn in name:
            dist -= 1000
        elif art.startswith(qn) or name.startswith(qn):
            dist -= 500
        scored.append((dist, item))
    scored.sort(
        key=lambda pair: (
            pair[0],
            str(pair[1].get("article", "")).casefold(),
        )
    )
    out: list[dict] = []
    seen: set[int] = set()
    for _, item in scored:
        if len(out) >= limit:
            break
        iid = int(item.get("id") or 0)
        if iid in seen:
            continue
        seen.add(iid)
        out.append(item)
    return out


def _is_orders_list_path(path: str) -> bool:
    return _norm_api_path(path) == "/api/orders"


def _is_orders_import_excel_path(path: str) -> bool:
    return _norm_api_path(path) == "/api/orders/import-excel"


def _is_orders_import_drogeri_excel_path(path: str) -> bool:
    return _norm_api_path(path) == "/api/orders/import-drogeri-excel"


def _is_orders_batches_export_pdf_path(path: str) -> bool:
    return _norm_api_path(path) == "/api/orders/batches-export-pdf"


def _is_arnest_unirus_pallet_sheets_pdf_path(path: str) -> bool:
    return _norm_api_path(path) == "/api/arnest-unirus-pallet-sheets-pdf"


def _is_lab_industries_pallet_sheets_pdf_path(path: str) -> bool:
    return _norm_api_path(path) == "/api/lab-industries-pallet-sheets-pdf"


def _is_lab_sscc_reset_path(path: str) -> bool:
    return _norm_api_path(path) == "/api/lab/sscc-reset"


def _is_arnest_unirus_pallet_sheets_xlsx_path(path: str) -> bool:
    return _norm_api_path(path) == "/api/arnest-unirus-pallet-sheets-xlsx"


def _is_print_arnest_unirus_pallet_sheets_raw_path(path: str) -> bool:
    return _norm_api_path(path) == "/api/print-arnest-unirus-pallet-sheets-raw"


def _is_pallet_printer_raw_ping_path(path: str) -> bool:
    return _norm_api_path(path) == "/api/pallet-printer-raw-ping"


# Префикс Code128 данных паллеты (как в макросе Word / tec-it).
PALLET_BARCODE_PREFIX = "1500000"
MAX_PALLET_SHEET_PDF_PAGES = 500


def _format_pallet_suffix_vba(n: int) -> str:
    """Эквивалент VBA Format(n, \"000\"): минимум три знака с ведущими нулями."""
    if n < 0:
        n = -n
    s = str(n)
    if len(s) < 3:
        return s.zfill(3)
    return s


def build_pallet_barcode_data(pallet_number: str) -> str | None:
    """Строка данных Code128: 1500000 + суффикс номера паллеты; None если номер пустой."""
    raw = str(pallet_number).strip()
    if not raw:
        return None
    try:
        n = int(float(raw.replace(",", ".")))
    except ValueError:
        digits = re.sub(r"\D", "", raw)
        if not digits:
            return None
        try:
            n = int(digits)
        except ValueError:
            return None
    return PALLET_BARCODE_PREFIX + _format_pallet_suffix_vba(n)


def render_code128_barcode_png(barcode_data: str, *, write_text: bool = True) -> bytes:
    """Code128 в PNG (python-barcode + Pillow). При write_text=False только полосы — подпись в PDF отдельно."""
    if not HAVE_CODE128_BARCODE or Code128 is None or ImageWriter is None:
        raise RuntimeError(
            "Не установлен python-barcode[images] (pip install \"python-barcode[images]\")"
        )
    payload = str(barcode_data)
    if not payload.strip():
        raise ValueError("Пустые данные для штрих-кода")
    writer = ImageWriter(format="PNG", dpi=300)
    buf = io.BytesIO()
    Code128(payload, writer=writer).write(buf, options={"write_text": write_text})
    out = buf.getvalue()
    if not out or _png_ihdr_pixel_size(out) is None:
        raise RuntimeError("Пустой или некорректный PNG штрих-кода")
    return out


def render_gs1_128_barcode_png(barcode_data: str, *, write_text: bool = True) -> bytes:
    """GS1-128 (UCC/EAN-128) в PNG; при write_text=False подпись рисуется в PDF отдельно."""
    if not HAVE_GS1_128_BARCODE or Gs1_128 is None or ImageWriter is None:
        raise RuntimeError(
            "Не установлен python-barcode[images] с поддержкой GS1-128 "
            '(pip install "python-barcode[images]")'
        )
    payload = str(barcode_data)
    if not payload.strip():
        raise ValueError("Пустые данные для штрих-кода GS1-128")
    writer = ImageWriter(format="PNG", dpi=300)
    buf = io.BytesIO()
    Gs1_128(payload, writer=writer).write(buf, options={"write_text": write_text})
    out = buf.getvalue()
    if not out or _png_ihdr_pixel_size(out) is None:
        raise RuntimeError("Пустой или некорректный PNG штрих-кода GS1-128")
    return out


def _png_ihdr_pixel_size(png: bytes) -> tuple[int, int] | None:
    """Ширина и высота растра из PNG IHDR (первый чанк)."""
    if len(png) < 24 or png[:8] != b"\x89PNG\r\n\x1a\n":
        return None
    if png[12:16] != b"IHDR":
        return None
    w = int.from_bytes(png[16:20], "big")
    h = int.from_bytes(png[20:24], "big")
    if w <= 0 or h <= 0:
        return None
    return w, h


def _gif_logical_screen_size(gif: bytes) -> tuple[int, int] | None:
    """Логический экран GIF87a/GIF89a (ширина и высота в пикселях)."""
    if len(gif) < 10 or gif[:6] not in (b"GIF87a", b"GIF89a"):
        return None
    w = int.from_bytes(gif[6:8], "little")
    h = int.from_bytes(gif[8:10], "little")
    if w <= 0 or h <= 0:
        return None
    return w, h


def _barcode_raster_pixel_size(raw: bytes) -> tuple[int, int] | None:
    """Размеры растра штрих-кода (PNG локальной генерации или GIF)."""
    z = _png_ihdr_pixel_size(raw)
    if z:
        return z
    return _gif_logical_screen_size(raw)


# A4 и поля как в типовом документе Word (≈2 см слева/справа).
_ARNEST_A4_W_MM = 210.0
_ARNEST_SIDE_MARGIN_MM = 20.0
# Строка 1: высота растра полос из соотношения сторон её PNG и _ARNEST_BARCODE_HEIGHT_SCALE.
# Строка 8: ширина _ARNEST_LINE8_BARCODE_W_MM, высота из PNG × scale.
# Строка 14 (паллета): та же ширина; высота принудительно как у строки 8, чтобы полосы были одинаковой высоты.
_ARNEST_BARCODE_HEIGHT_SCALE = 0.5
_ARNEST_BARCODE_CAPTION_GAP_MM = 1.5  # между низом растра полос и подписью (PDF)
_ARNEST_BARCODE_CAPTION_LINE_H_MM = 6.0
_ARNEST_BARCODE_CAPTION_FONT_MAX_PT = 11.0
_ARNEST_BARCODE_CAPTION_FONT_MIN_PT = 7.0
_ARNEST_LINE2_GAP_MM = 10.0  # только между строкой 1 (штрих-код) и строкой 2
_ARNEST_TEXT_LINE_GAP_MM = 5.0  # между всеми остальными строками (текст и нижний штрих-код)
_ARNEST_LINE4_NAME_EXTRA_GAP_MM = 4.0  # доп. зазор после блока строки 4 (наименование может занять 2+ строки)
_ARNEST_TEXT_FONT_PT = 12.0  # все текстовые элементы паллетного листа (не штрих-код)
_ARNEST_LINE2_TEXT_H_MM = 7.0
_ARNEST_LINE2_ARTICLE_MAX_W_MM = 105.0
_ARNEST_TAB_MM = 12.5  # одна стандартная табуляция (~1,25 см) между артикулом и наименованием в строке 4
_ARNEST_ASEPTICA_LABEL = "ASEPTICA"
_ARNEST_LINE3_DESCRIPTION_LABEL = "Description"
_ARNEST_LINE5_GTIN_LABEL = "GTIN CODE"
_ARNEST_LINE6_MGT_DATE_LABEL = "Mgt. Date"
_ARNEST_LINE6_EXPIRY_DATE_LABEL = "Expiry Date"
_ARNEST_LINE6_CROSS_WEIGHT_LABEL = "Cross Weight(KG)"
_ARNEST_LINE8_BARCODE_W_MM = 45.0
_ARNEST_LINE9_PALLET_QTY_LABEL = "Pallet Quantity"
_ARNEST_LINE9_UNITS_PC_LABEL = "Units PC"
_ARNEST_LINE9_TI_HI_LABEL = "TI*HI"
_ARNEST_LINE10_KOR_LABEL = "кор."
_ARNEST_LINE12_TOTAL_QUANTITY_LABEL = "Total Quantity"
_ARNEST_LINE15_BATCH_NUMBER_LABEL = "Batch Number"
_ARNEST_LINE14_TO_BATCH_LABEL_GAP_MM = 0.0  # штрих-код паллеты (стр. 14) → «Batch Number»
_ARNEST_LINE15_LABEL_TO_BATCH_GAP_MM = (
    _ARNEST_TEXT_LINE_GAP_MM / 2.0
)  # «Batch Number» → номер партии


def _arnest_regular_font_path() -> Path | None:
    """Обычный Calibri / запасной DejaVu Sans / Liberation для наименования в строке 4."""
    windir = Path(os.environ.get("WINDIR", r"C:\Windows"))
    win_fonts = windir / "Fonts"
    candidates = [
        BASE_DIR / "fonts" / "calibri.ttf",
        BASE_DIR / "fonts" / "Calibri.ttf",
        win_fonts / "calibri.ttf",
        win_fonts / "CALIBRI.TTF",
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("/usr/share/fonts/TTF/DejaVuSans.ttf"),
        Path("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"),
    ]
    for c in candidates:
        if c.is_file():
            return c
    return None


def _arnest_line2_bold_font_path() -> Path | None:
    """Жирный TTF: Calibri Bold при наличии, иначе типичный системный запасной (DejaVu на Linux)."""
    windir = Path(os.environ.get("WINDIR", r"C:\Windows"))
    win_fonts = windir / "Fonts"
    candidates = [
        BASE_DIR / "fonts" / "calibrib.ttf",
        BASE_DIR / "fonts" / "Calibrib.ttf",
        win_fonts / "calibrib.ttf",
        win_fonts / "CALIBRIB.TTF",
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        Path("/usr/share/fonts/TTF/DejaVuSans-Bold.ttf"),
        Path("/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"),
    ]
    for c in candidates:
        if c.is_file():
            return c
    return None


def _arnest_pdf_register_text_fonts(pdf: FPDF) -> str | None:
    """Семейство PLCalibri: обычный (наименование) и жирный (артикул, ASEPTICA, Description)."""
    bold_p = _arnest_line2_bold_font_path()
    if bold_p is None:
        return "no_line2_font"
    reg_p = _arnest_regular_font_path()
    try:
        pdf.add_font("PLCalibri", "B", str(bold_p))
        if reg_p is not None:
            pdf.add_font("PLCalibri", "", str(reg_p))
        else:
            pdf.add_font("PLCalibri", "", str(bold_p))
    except OSError:
        return "no_line2_font"
    return None


def _arnest_clip_text_to_width_mm(pdf: FPDF, text: str, max_mm: float) -> str:
    """Подрезает строку по ширине (мм) с многоточием, шрифт уже выбран."""
    t = text or ""
    if pdf.get_string_width(t) <= max_mm:
        return t
    ell = "…"
    if pdf.get_string_width(ell) > max_mm:
        return ""
    n = len(t)
    while n > 0:
        cand = t[:n] + ell
        if pdf.get_string_width(cand) <= max_mm:
            return cand
        n -= 1
    return ell


def _arnest_barcode_caption_font_pt(pdf: FPDF, text: str, max_w_mm: float) -> float:
    """Размер шрифта подписи под Code128: уменьшаем, пока строка влезает в max_w_mm."""
    t = str(text) if text is not None else ""
    lo = _ARNEST_BARCODE_CAPTION_FONT_MIN_PT
    hi = _ARNEST_BARCODE_CAPTION_FONT_MAX_PT
    pt = hi
    step = 0.5
    while pt >= lo:
        pdf.set_font("PLCalibri", "", pt)
        if pdf.get_string_width(t) <= max_w_mm:
            return pt
        pt -= step
    pdf.set_font("PLCalibri", "", lo)
    return lo


def _arnest_draw_code128_caption_below(
    pdf: FPDF,
    x: float,
    w_mm: float,
    y_barcode_bottom: float,
    text: str,
    *,
    body_caption_style: bool = False,
) -> float:
    """Подпись под штрих-кодом (отдельно от PNG). Возвращает занятую высоту: зазор + строка.

    body_caption_style: как цифры на листе (партия и т.д.) — 12 pt жирный, без уменьшения кегля.
    """
    gap = _ARNEST_BARCODE_CAPTION_GAP_MM
    s = str(text) if text is not None else ""
    if body_caption_style:
        h_line = _ARNEST_LINE2_TEXT_H_MM
        fs = _ARNEST_TEXT_FONT_PT
        pdf.set_font("PLCalibri", "B", fs)
        draw = _arnest_clip_text_to_width_mm(pdf, s, w_mm) if s else ""
        pdf.set_xy(x, y_barcode_bottom + gap)
        pdf.cell(w_mm, h_line, draw or s, align="C")
        return gap + h_line
    h_line = _ARNEST_BARCODE_CAPTION_LINE_H_MM
    fs = _arnest_barcode_caption_font_pt(pdf, text, w_mm)
    pdf.set_font("PLCalibri", "", fs)
    draw = s if pdf.get_string_width(s) <= w_mm else _arnest_clip_text_to_width_mm(pdf, s, w_mm)
    pdf.set_xy(x, y_barcode_bottom + gap)
    pdf.cell(w_mm, h_line, draw or s, align="C")
    return gap + h_line


def _arnest_pallet_pdf_error_message(code: str) -> str:
    return {
        "no_fpdf": "На сервере не установлен пакет fpdf2 (pip install fpdf2).",
        "validation": "Укажите число страниц от 1 до 500 (по одной на паллету).",
        "validation_pallets": "Передайте непустой массив pallets (не более 500 паллет).",
        "pdf_build": "Не удалось сформировать PDF.",
        "no_barcode": (
            "На сервере не установлен python-barcode[images] "
            '(pip install "python-barcode[images]").'
        ),
        "barcode_fetch": "Не удалось сформировать изображение штрих-кода.",
        "no_line2_font": (
            "Не найдены TTF для текста паллетного PDF (жирный обязателен). На сервере: apt install fonts-dejavu-core "
            "или положите calibri.ttf и calibrib.ttf в каталог fonts/ рядом с api_server.py."
        ),
        "no_openpyxl": 'На сервере не установлен openpyxl (pip install openpyxl).',
        "xlsx_build": "Не удалось сформировать Excel-файл.",
    }.get(code, "Ошибка генерации PDF.")


def _arnest_first_digit_run(article: str) -> str | None:
    m = re.search(r"\d+", article or "")
    return m.group(0) if m else None


def _arnest_yymmdd_digits(raw: str) -> str | None:
    d = re.sub(r"\D", "", raw or "")
    if len(d) >= 6:
        d = d[-6:]
    if len(d) != 6 or not d.isdigit():
        return None
    return d


def _arnest_yymmdd_for_barcode(raw: str) -> str | None:
    """Шесть цифр из поля → в штрих-коде пары AB·CD·EF дают EF·CD·AB (напр. 123456 → 563412)."""
    d = _arnest_yymmdd_digits(raw)
    if not d:
        return None
    return d[4:6] + d[2:4] + d[0:2]


def _arnest_six_digits_display_ddmmyy(raw: str) -> str | None:
    """Шесть цифр → «ДД.ММ.ГГГГ» для строки 7: сначала ДДММГГ, при невалидной дате — ГГММДД.

    Если оба разбора невалидны (напр. несуществующий месяц вроде «22»), возвращается None —
    в PDF подставляется «—»; выдумывать дату не нужно.
    """
    d = re.sub(r"\D", "", raw or "")
    if len(d) >= 6:
        d = d[-6:]
    if len(d) != 6 or not d.isdigit():
        return None

    dd = int(d[0:2])
    mm = int(d[2:4])
    yy = int(d[4:6])
    year = 2000 + yy
    try:
        datetime.date(year, mm, dd)
        return f"{dd:02d}.{mm:02d}.{year:04d}"
    except ValueError:
        pass

    yy2 = int(d[0:2])
    mm2 = int(d[2:4])
    dd2 = int(d[4:6])
    yfull = 2000 + yy2
    try:
        datetime.date(yfull, mm2, dd2)
        return f"{dd2:02d}.{mm2:02d}.{yfull:04d}"
    except ValueError:
        return None


def _arnest_coerce_date_string(val) -> str | None:
    if val is None or isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float):
            if not math.isfinite(val):
                return None
            if abs(val - round(val)) < 1e-9:
                s = str(int(round(val)))
            else:
                s = str(val).strip()
        else:
            s = str(val)
    else:
        s = str(val).strip()
    if not s or s.lower() == "none":
        return None
    return s


def _arnest_unirus_date_raw(row: dict, *keys: str) -> str:
    """Дата из полей верхнего уровня паллеты (учёт JSON null)."""
    for key in keys:
        if key not in row:
            continue
        s = _arnest_coerce_date_string(row[key])
        if s:
            return s
    return ""


def _arnest_date_raw_from_slots(row: dict, *slot_keys: str) -> str:
    """Дата из slots[] (если в объекте паллеты не продублированы наверх)."""
    slots = row.get("slots")
    if not isinstance(slots, list):
        return ""
    for slot in slots:
        if not isinstance(slot, dict):
            continue
        for sk in slot_keys:
            if sk not in slot:
                continue
            s = _arnest_coerce_date_string(slot.get(sk))
            if s:
                return s
    return ""


def _arnest_mfg_date_raw(row: dict) -> str:
    s = _arnest_unirus_date_raw(
        row,
        "manufacturing_date_raw",
        "manufacturingDateRaw",
        "unirusMfgDate",
    )
    if s:
        return s
    return _arnest_date_raw_from_slots(
        row, "unirusMfgDate", "manufacturingDateRaw", "manufacturing_date_raw"
    )


def _arnest_exp_date_raw(row: dict) -> str:
    s = _arnest_unirus_date_raw(
        row,
        "expiry_date_raw",
        "expiryDateRaw",
        "unirusExpiryDate",
    )
    if s:
        return s
    return _arnest_date_raw_from_slots(
        row, "unirusExpiryDate", "expiryDateRaw", "expiry_date_raw"
    )


def _arnest_format_weight_kg_ru(raw) -> str:
    """Вес в кг для PDF: одна десятичная, запятая как десятичный разделитель."""
    try:
        v = float(raw)
    except (TypeError, ValueError):
        v = 0.0
    if not math.isfinite(v) or v < 0:
        v = 0.0
    s = f"{v:.1f}"
    return s.replace(".", ",")


def _arnest_format_boxes_qty_for_barcode(raw) -> str:
    """Количество коробок для штрих-кода строки 8."""
    try:
        v = float(raw)
    except (TypeError, ValueError):
        v = 0.0
    if not math.isfinite(v) or v < 0:
        v = 0.0
    if abs(v - round(v)) < 1e-9:
        return str(int(round(v)))
    return f"{v:.1f}".rstrip("0").rstrip(".")


def _arnest_remainder_boxes_star_one(boxes_f: float, rl: int) -> str:
    """Остаток коробок после полных рядов: всего − (ряд × полные ряды), затем «*1» (напр. 5*1)."""
    if rl <= 0 or not math.isfinite(boxes_f) or boxes_f < 0:
        return "—*1"
    full_rows = int(math.floor(boxes_f / rl + 1e-9))
    rem = max(0.0, boxes_f - rl * full_rows)
    if abs(rem - round(rem)) < 1e-9:
        rem_s = str(int(round(rem)))
    else:
        rem_s = f"{rem:.1f}".rstrip("0").rstrip(".")
    return f"{rem_s}*1"


def _arnest_line_barcode_data(row: dict) -> tuple[str | None, str | None]:
    """Строка Code128: артикул (цифры), партия, дата изготовления и срок в кодировке для сканера (пары цифр дат переставлены)."""
    article = str(row.get("article", "")).strip()
    part_art = _arnest_first_digit_run(article)
    if not part_art:
        return None, "в артикуле нет цифр"
    batch = str(row.get("batch_number", row.get("batchNumber", ""))).strip()
    if not batch:
        return None, "не указана партия"
    mfg = _arnest_yymmdd_for_barcode(_arnest_mfg_date_raw(row))
    if not mfg:
        return None, "дата изготовления: нужны 6 цифр ДДММГГ"
    exp = _arnest_yymmdd_for_barcode(_arnest_exp_date_raw(row))
    if not exp:
        return None, "срок годности: нужны 6 цифр ДДММГГ"
    return f"{part_art} {batch} {mfg} {exp}", None


def build_arnest_unirus_pallet_sheets_pdf_with_barcodes(
    pallets: list[dict],
) -> tuple[bytes | None, str, str]:
    """Один PDF: 1, 8, 14 — Code128 (только полосы в PNG, читаемая строка под ними в PDF); 2–16 — текст 12 pt; 14 — паллета слева (если номер); 15–16 — партия у колонки «кор.»."""
    if not HAVE_FPDF or FPDF is None or Align is None:
        return None, "no_fpdf", ""
    if not HAVE_CODE128_BARCODE:
        return None, "no_barcode", _arnest_pallet_pdf_error_message("no_barcode")
    pallet_rows = [p for p in pallets if isinstance(p, dict)]
    if sort_assemble_pallets_by_number is not None:
        pallet_rows = sort_assemble_pallets_by_number(pallet_rows)
    n = len(pallet_rows)
    if n < 1 or n > MAX_PALLET_SHEET_PDF_PAGES:
        return None, "validation_pallets", ""
    try:
        pdf = FPDF(orientation="P", unit="mm", format="A4")
        pdf.set_auto_page_break(False)
        font_err = _arnest_pdf_register_text_fonts(pdf)
        if font_err:
            return None, font_err, _arnest_pallet_pdf_error_message(font_err)
        barcode_w_mm = _ARNEST_A4_W_MM - 2 * _ARNEST_SIDE_MARGIN_MM
        y_top = 14.0
        for idx, row in enumerate(pallet_rows, start=1):
            data, err_detail = _arnest_line_barcode_data(row)
            if err_detail:
                return None, "validation_pallet", f"Паллета {idx}: {err_detail}."
            try:
                png = render_code128_barcode_png(data, write_text=False)
            except Exception as exc:
                return (
                    None,
                    "barcode_fetch",
                    f"Не удалось сформировать штрих-код для паллеты {idx}: {exc}",
                )
            dims = _barcode_raster_pixel_size(png)
            if not dims:
                return None, "pdf_build", ""
            pw, ph = dims
            h_base_mm = barcode_w_mm * (ph / pw)
            h_mm = h_base_mm * _ARNEST_BARCODE_HEIGHT_SCALE
            pdf.add_page()
            pdf.image(
                io.BytesIO(png),
                x=Align.C,
                y=y_top,
                w=barcode_w_mm,
                h=h_mm,
                keep_aspect_ratio=False,
            )
            cap1_h = _arnest_draw_code128_caption_below(
                pdf, _ARNEST_SIDE_MARGIN_MM, barcode_w_mm, y_top + h_mm, data
            )
            y2 = y_top + h_mm + cap1_h + _ARNEST_LINE2_GAP_MM
            fs = _ARNEST_TEXT_FONT_PT
            h_txt = _ARNEST_LINE2_TEXT_H_MM
            left_w = _ARNEST_LINE2_ARTICLE_MAX_W_MM
            art_full = str(row.get("article", "")).strip()
            pdf.set_font("PLCalibri", "BU", fs)
            art_show = (
                _arnest_clip_text_to_width_mm(pdf, art_full, left_w) if art_full else "—"
            )
            if not art_show:
                art_show = "—"
            pdf.set_xy(_ARNEST_SIDE_MARGIN_MM, y2)
            pdf.cell(left_w, h_txt, art_show)
            pdf.set_font("PLCalibri", "B", fs)
            ase = _ARNEST_ASEPTICA_LABEL
            w_ase = pdf.get_string_width(ase)
            pdf.set_xy((_ARNEST_A4_W_MM - w_ase) / 2.0, y2)
            pdf.cell(w_ase, h_txt, ase)
            y3 = y2 + h_txt + _ARNEST_TEXT_LINE_GAP_MM
            content_w = _ARNEST_A4_W_MM - 2 * _ARNEST_SIDE_MARGIN_MM
            pdf.set_font("PLCalibri", "B", fs)
            pdf.set_xy(_ARNEST_SIDE_MARGIN_MM, y3)
            pdf.cell(content_w, h_txt, _ARNEST_LINE3_DESCRIPTION_LABEL, align="L")
            y4 = y3 + h_txt + _ARNEST_TEXT_LINE_GAP_MM
            x0 = _ARNEST_SIDE_MARGIN_MM
            max_r = _ARNEST_A4_W_MM - _ARNEST_SIDE_MARGIN_MM
            art_l4 = str(row.get("article", "")).strip() or "—"
            name_raw = str(row.get("name", row.get("product_name", ""))).strip() or "—"
            pdf.set_font("PLCalibri", "B", fs)
            w_art0 = pdf.get_string_width(art_l4)
            max_art = max(10.0, max_r - x0 - _ARNEST_TAB_MM - 25.0)
            art_l4_draw = (
                _arnest_clip_text_to_width_mm(pdf, art_l4, max_art)
                if w_art0 > max_art
                else art_l4
            )
            w_art = pdf.get_string_width(art_l4_draw)
            pdf.set_xy(x0, y4)
            pdf.cell(w_art, h_txt, art_l4_draw)
            x_name = x0 + w_art + _ARNEST_TAB_MM
            pdf.set_font("PLCalibri", "", fs)
            rem_w = max(0.0, max_r - x_name)
            name_txt = name_raw if name_raw else "—"
            name_block_h = h_txt
            if rem_w > 0.5:
                pdf.set_xy(x_name, y4)
                wm = WrapMode.WORD if WrapMode is not None else None
                if wm is not None:
                    pdf.multi_cell(
                        rem_w,
                        h_txt,
                        name_txt,
                        border=0,
                        align="L",
                        wrapmode=wm,
                    )
                else:
                    pdf.multi_cell(rem_w, h_txt, name_txt, border=0, align="L")
                name_block_h = max(h_txt, pdf.get_y() - y4)
            else:
                pdf.set_xy(x_name, y4)
                pdf.cell(1.0, h_txt, _arnest_clip_text_to_width_mm(pdf, name_txt, 1.0) or "—")
            y5 = (
                y4
                + max(h_txt, name_block_h)
                + _ARNEST_LINE4_NAME_EXTRA_GAP_MM
                + _ARNEST_TEXT_LINE_GAP_MM
            )
            pdf.set_font("PLCalibri", "B", fs)
            pdf.set_xy(x0, y5)
            pdf.cell(content_w, h_txt, _ARNEST_LINE5_GTIN_LABEL, align="L")
            y6 = y5 + h_txt + _ARNEST_TEXT_LINE_GAP_MM
            tab = _ARNEST_TAB_MM
            mgt = _ARNEST_LINE6_MGT_DATE_LABEL
            exp = _ARNEST_LINE6_EXPIRY_DATE_LABEL
            cwk = _ARNEST_LINE6_CROSS_WEIGHT_LABEL
            pdf.set_font("PLCalibri", "", fs)
            w_mgt = pdf.get_string_width(mgt)
            w_exp = pdf.get_string_width(exp)
            pdf.set_xy(x0, y6)
            pdf.cell(w_mgt, h_txt, mgt)
            x_exp = x0 + w_mgt + tab + pdf.get_string_width("  ")
            pdf.set_xy(x_exp, y6)
            pdf.cell(w_exp, h_txt, exp)
            x_cwk = x_exp + w_exp + 3 * tab
            pdf.set_xy(x_cwk, y6)
            rem_cwk = max(0.0, max_r - x_cwk)
            if rem_cwk > 0:
                cwk_draw = _arnest_clip_text_to_width_mm(pdf, cwk, rem_cwk)
                pdf.cell(rem_cwk, h_txt, cwk_draw or cwk)
            y7 = y6 + h_txt + _ARNEST_TEXT_LINE_GAP_MM
            mfg_raw = _arnest_mfg_date_raw(row)
            exp_raw = _arnest_exp_date_raw(row)
            mfg_txt = _arnest_six_digits_display_ddmmyy(mfg_raw) or "—"
            exp_txt = _arnest_six_digits_display_ddmmyy(exp_raw) or "—"
            w_kg_str = _arnest_format_weight_kg_ru(row.get("pallet_weight_kg", 0))
            pdf.set_font("PLCalibri", "B", fs)
            # Выравниваем строку 7 по тем же колонкам, что и строку 6.
            x_e7 = x_exp
            x_w7 = x_cwk
            rem_m7 = max(0.0, x_e7 - x0)
            rem_e7 = max(0.0, x_w7 - x_e7)
            rem_w7 = max(0.0, max_r - x_w7)
            pdf.set_xy(x0, y7)
            mfg_draw = _arnest_clip_text_to_width_mm(pdf, mfg_txt, rem_m7) if rem_m7 > 0 else ""
            pdf.cell(rem_m7, h_txt, mfg_draw or mfg_txt)
            pdf.set_xy(x_e7, y7)
            exp_draw = _arnest_clip_text_to_width_mm(pdf, exp_txt, rem_e7) if rem_e7 > 0 else ""
            pdf.cell(rem_e7, h_txt, exp_draw or exp_txt)
            pdf.set_xy(x_w7, y7)
            if rem_w7 > 0:
                w_kg_draw = _arnest_clip_text_to_width_mm(pdf, w_kg_str, rem_w7)
                pdf.cell(rem_w7, h_txt, w_kg_draw or w_kg_str)
            y8 = y7 + h_txt + _ARNEST_TEXT_LINE_GAP_MM
            boxes_qty = _arnest_format_boxes_qty_for_barcode(
                row.get("pallet_boxes_qty", row.get("pallet_qty", 0))
            )
            try:
                png_boxes = render_code128_barcode_png(boxes_qty, write_text=False)
            except Exception as exc:
                return (
                    None,
                    "barcode_fetch",
                    f"Не удалось сформировать штрих-код количества коробок для паллеты {idx}: {exc}",
                )
            dims_boxes = _barcode_raster_pixel_size(png_boxes)
            if not dims_boxes:
                return None, "pdf_build", ""
            pw2, ph2 = dims_boxes
            h_boxes_mm = (
                _ARNEST_LINE8_BARCODE_W_MM * (ph2 / pw2) * _ARNEST_BARCODE_HEIGHT_SCALE
            )
            pdf.image(
                io.BytesIO(png_boxes),
                x=Align.C,
                y=y8,
                w=_ARNEST_LINE8_BARCODE_W_MM,
                h=h_boxes_mm,
                keep_aspect_ratio=False,
            )
            cap8_h = _arnest_draw_code128_caption_below(
                pdf,
                _ARNEST_SIDE_MARGIN_MM,
                barcode_w_mm,
                y8 + h_boxes_mm,
                boxes_qty,
                body_caption_style=True,
            )
            y9 = y8 + h_boxes_mm + cap8_h + _ARNEST_TEXT_LINE_GAP_MM
            pdf.set_font("PLCalibri", "", fs)
            pdf.set_xy(x0, y9)
            pdf.cell(content_w, h_txt, _ARNEST_LINE9_PALLET_QTY_LABEL, align="L")
            pdf.set_font("PLCalibri", "", fs)
            x_units = x_w7
            pdf.set_xy(x_units, y9)
            pdf.cell(max(0.0, max_r - x_units), h_txt, _ARNEST_LINE9_UNITS_PC_LABEL)
            x_tihi = x_units + pdf.get_string_width(_ARNEST_LINE9_UNITS_PC_LABEL) + tab
            pdf.set_xy(x_tihi, y9)
            pdf.cell(max(0.0, max_r - x_tihi), h_txt, _ARNEST_LINE9_TI_HI_LABEL)
            y10 = y9 + h_txt + _ARNEST_TEXT_LINE_GAP_MM
            boxes_str = _arnest_format_boxes_qty_for_barcode(row.get("pallet_boxes_qty", 0))
            rl = int(row.get("row_layout", 0) or 0)
            try:
                boxes_f = float(row.get("pallet_boxes_qty", 0) or 0)
            except (TypeError, ValueError):
                boxes_f = 0.0
            if rl > 0:
                full_rows = int(math.floor(boxes_f / rl + 1e-9))
                ti_hi_str = f"{rl}*{full_rows}"
            else:
                ti_hi_str = "—"
            kor = _ARNEST_LINE10_KOR_LABEL
            pdf.set_font("PLCalibri", "B", fs)
            rem_l10 = max(0.0, x_units - x0 - 1.0)
            pdf.set_xy(x0, y10)
            if rem_l10 > 0:
                boxes_draw = _arnest_clip_text_to_width_mm(pdf, boxes_str, rem_l10)
                pdf.cell(rem_l10, h_txt, boxes_draw or boxes_str)
            else:
                pdf.cell(pdf.get_string_width(boxes_str), h_txt, boxes_str)
            pdf.set_font("PLCalibri", "B", fs)
            pdf.set_xy(x_units, y10)
            pdf.cell(pdf.get_string_width(kor), h_txt, kor)
            pdf.set_font("PLCalibri", "B", fs)
            pdf.set_xy(x_tihi, y10)
            rem_t10 = max(0.0, max_r - x_tihi)
            if rem_t10 > 0:
                ti_draw = _arnest_clip_text_to_width_mm(pdf, ti_hi_str, rem_t10)
                pdf.cell(rem_t10, h_txt, ti_draw or ti_hi_str)
            else:
                pdf.cell(pdf.get_string_width(ti_hi_str), h_txt, ti_hi_str)
            y11 = y10 + h_txt + _ARNEST_TEXT_LINE_GAP_MM
            rem_star = _arnest_remainder_boxes_star_one(boxes_f, rl)
            pdf.set_font("PLCalibri", "B", fs)
            pdf.set_xy(x_units, y11)
            pdf.cell(pdf.get_string_width(kor), h_txt, kor)
            pdf.set_font("PLCalibri", "B", fs)
            pdf.set_xy(x_tihi, y11)
            rem_t11 = max(0.0, max_r - x_tihi)
            if rem_t11 > 0:
                r11 = _arnest_clip_text_to_width_mm(pdf, rem_star, rem_t11)
                pdf.cell(rem_t11, h_txt, r11 or rem_star)
            else:
                pdf.cell(pdf.get_string_width(rem_star), h_txt, rem_star)
            y12 = y11 + h_txt + _ARNEST_TEXT_LINE_GAP_MM
            pdf.set_font("PLCalibri", "", fs)
            pdf.set_xy(x0, y12)
            pdf.cell(content_w, h_txt, _ARNEST_LINE12_TOTAL_QUANTITY_LABEL, align="L")
            y13 = y12 + h_txt + _ARNEST_TEXT_LINE_GAP_MM
            units_str = _arnest_format_boxes_qty_for_barcode(
                row.get("units_pc", row.get("unitsPc", 0))
            )
            pdf.set_font("PLCalibri", "B", fs)
            pdf.set_xy(x0, y13)
            rem_13 = content_w
            if rem_13 > 0:
                u13 = _arnest_clip_text_to_width_mm(pdf, units_str, rem_13)
                pdf.cell(rem_13, h_txt, u13 or units_str)
            else:
                pdf.cell(pdf.get_string_width(units_str), h_txt, units_str)
            y14 = y13 + h_txt + _ARNEST_TEXT_LINE_GAP_MM
            pallet_bc = build_pallet_barcode_data(
                str(row.get("pallet_number", row.get("palletNumber", "")))
            )
            y_after_line14 = y14
            if pallet_bc:
                try:
                    png_pn = render_code128_barcode_png(pallet_bc, write_text=False)
                except Exception as exc:
                    return (
                        None,
                        "barcode_fetch",
                        f"Не удалось сформировать штрих-код номера паллеты {idx}: {exc}",
                    )
                # Как у штрих-кода коробок (стр. 8): короткий Code128 паллеты иначе получался ниже.
                h_pallet_mm = h_boxes_mm
                pdf.image(
                    io.BytesIO(png_pn),
                    x=x0,
                    y=y14,
                    w=_ARNEST_LINE8_BARCODE_W_MM,
                    h=h_pallet_mm,
                    keep_aspect_ratio=False,
                )
                cap14_h = _arnest_draw_code128_caption_below(
                    pdf,
                    x0,
                    _ARNEST_LINE8_BARCODE_W_MM,
                    y14 + h_pallet_mm,
                    pallet_bc,
                    body_caption_style=True,
                )
                y_after_line14 = (
                    y14 + h_pallet_mm + cap14_h + _ARNEST_LINE14_TO_BATCH_LABEL_GAP_MM
                )
            y15 = y_after_line14
            rem_bn = max(0.0, max_r - x_units)
            pdf.set_font("PLCalibri", "", fs)
            pdf.set_xy(x_units, y15)
            pdf.cell(rem_bn, h_txt, _ARNEST_LINE15_BATCH_NUMBER_LABEL, align="L")
            y16 = y15 + h_txt + _ARNEST_LINE15_LABEL_TO_BATCH_GAP_MM
            batch_disp = str(
                row.get("batch_number", row.get("batchNumber", ""))
            ).strip() or "—"
            pdf.set_font("PLCalibri", "B", fs)
            pdf.set_xy(x_units, y16)
            if rem_bn > 0:
                b16 = _arnest_clip_text_to_width_mm(pdf, batch_disp, rem_bn)
                pdf.cell(rem_bn, h_txt, b16 or batch_disp)
            else:
                pdf.cell(pdf.get_string_width(batch_disp), h_txt, batch_disp)
        raw = pdf.output(dest="S")
    except Exception:
        return None, "pdf_build", ""
    if isinstance(raw, (bytes, bytearray)):
        return bytes(raw), "", ""
    if isinstance(raw, str):
        return raw.encode("latin-1"), "", ""
    return None, "pdf_build", ""


def build_blank_arnest_unirus_pallet_sheets_pdf(page_count: int) -> tuple[bytes | None, str]:
    """Один PDF: page_count пустых страниц A4 (пока без содержимого)."""
    if not HAVE_FPDF or FPDF is None:
        return None, "no_fpdf"
    if page_count < 1 or page_count > MAX_PALLET_SHEET_PDF_PAGES:
        return None, "validation"
    try:
        pdf = FPDF(orientation="P", unit="mm", format="A4")
        pdf.set_auto_page_break(False)
        for _ in range(page_count):
            pdf.add_page()
        raw = pdf.output(dest="S")
    except Exception:
        return None, "pdf_build"
    if isinstance(raw, (bytes, bytearray)):
        return bytes(raw), ""
    if isinstance(raw, str):
        return raw.encode("latin-1"), ""
    return None, "pdf_build"


def _batches_export_pdf_page_bottom(pdf: FPDF, margin: float, footer_h: float) -> float:
    return pdf.h - margin - footer_h


def _batches_export_pdf_row_height(
    pdf: FPDF,
    w_name: float,
    w_batch: float,
    line_h: float,
    name: str,
    batches: str,
) -> float:
    name_s = str(name or "—").strip() or "—"
    batch_s = str(batches or "").strip() or "—"
    h_name = pdf.multi_cell(w_name, line_h, name_s, dry_run=True, output="HEIGHT")
    h_batch = pdf.multi_cell(w_batch, line_h, batch_s, dry_run=True, output="HEIGHT")
    return max(line_h, h_name, h_batch)


def _batches_export_pdf_draw_table_header(
    pdf: FPDF,
    col_name: float,
    col_batch: float,
) -> None:
    pdf.set_font("PLCalibri", "B", 10)
    pdf.set_fill_color(228, 238, 234)
    pdf.cell(col_name, 7, "Наименование", border=1, fill=True)
    pdf.cell(col_batch, 7, "Партия", border=1, fill=True, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("PLCalibri", "", 10)


def _batches_export_pdf_draw_block_intro(
    pdf: FPDF,
    ship: str,
    client: str,
    col_name: float,
    col_batch: float,
) -> None:
    pdf.set_font("PLCalibri", "B", 11)
    pdf.cell(0, 6, f"Дата отгрузки: {ship}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"Клиент: {client}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    _batches_export_pdf_draw_table_header(pdf, col_name, col_batch)


def _batches_export_pdf_draw_row(
    pdf: FPDF,
    x: float,
    y: float,
    w_name: float,
    w_batch: float,
    line_h: float,
    name: str,
    batches: str,
) -> float:
    """Две ячейки в строке; возвращает высоту строки (мм)."""
    name_s = str(name or "—").strip() or "—"
    batch_s = str(batches or "").strip() or "—"
    h_row = _batches_export_pdf_row_height(pdf, w_name, w_batch, line_h, name_s, batch_s)
    pdf.rect(x, y, w_name, h_row)
    pdf.rect(x + w_name, y, w_batch, h_row)
    pdf.set_xy(x, y)
    pdf.multi_cell(w_name, line_h, name_s, border=0)
    pdf.set_xy(x + w_name, y)
    pdf.multi_cell(w_batch, line_h, batch_s, border=0)
    return h_row


class _OrdersBatchesExportPdf(FPDF):
    _FOOTER_H_MM = 10.0

    def footer(self) -> None:
        self.set_y(-self._FOOTER_H_MM)
        self.set_font("PLCalibri", "", 9)
        self.cell(0, 8, f"Страница {self.page_no()} из {{nb}}", align="C")


def build_orders_batches_export_pdf_bytes(body: dict) -> tuple[bytes | None, str | None, str | None]:
    """PDF: дата отгрузки, клиент, наименования и партии по отфильтрованным заказам (тело с клиента)."""
    orders_raw = body.get("orders")
    if not isinstance(orders_raw, list) or len(orders_raw) == 0:
        return None, "validation", "Нет заказов для выгрузки."
    if not HAVE_FPDF or FPDF is None:
        return None, "no_fpdf", _arnest_pallet_pdf_error_message("no_fpdf")
    try:
        margin = 12.0
        footer_h = _OrdersBatchesExportPdf._FOOTER_H_MM
        pdf = _OrdersBatchesExportPdf(orientation="P", unit="mm", format="A4")
        font_err = _arnest_pdf_register_text_fonts(pdf)
        if font_err:
            return None, font_err, _arnest_pallet_pdf_error_message(font_err)
        pdf.alias_nb_pages()
        pdf.set_auto_page_break(auto=False)
        pdf.set_margins(margin, margin, margin)
        pdf.add_page()
        page_w = pdf.w - 2 * margin
        col_name = page_w * 0.62
        col_batch = page_w - col_name
        line_h = 6.0
        x0 = pdf.l_margin
        page_bottom = _batches_export_pdf_page_bottom(pdf, margin, footer_h)
        block_intro_h = 21.0

        pdf.set_font("PLCalibri", "B", 14)
        pdf.cell(0, 10, "Партии по заказам", new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.ln(3)

        wrote_any = False
        for block in orders_raw:
            if not isinstance(block, dict):
                continue
            lines_raw = block.get("lines")
            if not isinstance(lines_raw, list) or not lines_raw:
                continue
            lines = [ln for ln in lines_raw if isinstance(ln, dict)]
            if not lines:
                continue

            ship = str(block.get("ship_date") or "—").strip() or "—"
            client = str(block.get("client") or "—").strip() or "—"
            pdf.set_font("PLCalibri", "", 10)
            row_items: list[tuple[str, str, float]] = []
            for ln in lines:
                name = str(ln.get("name") or "—").strip() or "—"
                batches = str(ln.get("batches") or "").strip() or "—"
                row_h = _batches_export_pdf_row_height(
                    pdf, col_name, col_batch, line_h, name, batches
                )
                row_items.append((name, batches, row_h))

            gap = 4.0 if wrote_any else 0.0
            body_h = block_intro_h + sum(r[2] for r in row_items)
            need_h = gap + body_h
            y = pdf.get_y()
            if y + need_h > page_bottom:
                if body_h <= page_bottom - margin:
                    pdf.add_page()
                else:
                    first_row_h = row_items[0][2] if row_items else 0.0
                    min_start_h = gap + block_intro_h + first_row_h
                    if y + min_start_h > page_bottom:
                        pdf.add_page()

            if wrote_any:
                pdf.ln(4)
            wrote_any = True

            _batches_export_pdf_draw_block_intro(pdf, ship, client, col_name, col_batch)
            for name, batches, row_h in row_items:
                y = pdf.get_y()
                if y + row_h > page_bottom:
                    pdf.add_page()
                    _batches_export_pdf_draw_block_intro(pdf, ship, client, col_name, col_batch)
                    y = pdf.get_y()
                h_row = _batches_export_pdf_draw_row(
                    pdf, x0, y, col_name, col_batch, line_h, name, batches
                )
                pdf.set_xy(x0, y + h_row)

        if not wrote_any:
            return None, "validation", "Нет заказов с указанными партиями."

        raw = pdf.output(dest="S")
    except Exception:
        return None, "pdf_build", "Не удалось сформировать PDF."
    if isinstance(raw, (bytes, bytearray)):
        return bytes(raw), None, None
    if isinstance(raw, str):
        return raw.encode("latin-1"), None, None
    return None, "pdf_build", "Не удалось сформировать PDF."


def build_arnest_unirus_pallet_sheets_pdf_bytes(body: dict) -> tuple[bytes | None, str | None, str | None]:
    """Собрать PDF в памяти (тот же ввод, что у POST /api/arnest-unirus-pallet-sheets-pdf).
    Возвращает (pdf_bytes, err_code, err_detail); err_code None при успехе."""
    pallets_raw = body.get("pallets")
    if isinstance(pallets_raw, list) and len(pallets_raw) > 0:
        blob, err, err_detail = build_arnest_unirus_pallet_sheets_pdf_with_barcodes(pallets_raw)
        if err:
            return None, err, err_detail or _arnest_pallet_pdf_error_message(err)
        return blob, None, None
    if isinstance(pallets_raw, list) and len(pallets_raw) == 0:
        return None, "validation_pallets", _arnest_pallet_pdf_error_message("validation_pallets")
    raw = body.get("page_count", body.get("pages", 0))
    try:
        page_count = int(raw)
    except (TypeError, ValueError):
        page_count = 0
    blob, err = build_blank_arnest_unirus_pallet_sheets_pdf(page_count)
    if err:
        return None, err, _arnest_pallet_pdf_error_message(err)
    return blob, None, None


def build_arnest_unirus_pallet_sheets_xlsx_bytes(
    body: dict,
) -> tuple[bytes | None, str | None, str | None]:
    """Сводный XLSX только в памяти (на диск не пишется).

    JSON: { \"pallets\": [...] } — как для PDF; опционально ship_date (строка), order_id (число) — для шапки листа.
    """
    if not HAVE_OPENPYXL or Workbook is None or get_column_letter is None:
        return None, "no_openpyxl", _arnest_pallet_pdf_error_message("no_openpyxl")
    pallets_raw = body.get("pallets")
    ship_date_raw = str(body.get("ship_date") or "").strip()
    order_id_raw = body.get("order_id")
    try:
        order_id_int = int(order_id_raw) if order_id_raw is not None and str(order_id_raw).strip() != "" else None
    except (TypeError, ValueError):
        order_id_int = None
    if not isinstance(pallets_raw, list) or len(pallets_raw) == 0:
        return None, "validation_pallets", _arnest_pallet_pdf_error_message("validation_pallets")
    if len(pallets_raw) > MAX_PALLET_SHEET_PDF_PAGES:
        return None, "validation_pallets", _arnest_pallet_pdf_error_message("validation_pallets")

    validated: list[dict] = []
    for idx, row in enumerate(pallets_raw, start=1):
        if not isinstance(row, dict):
            return None, "validation_pallet", f"Паллета {idx}: ожидался объект с полями."
        _, err_detail = _arnest_line_barcode_data(row)
        if err_detail:
            return None, "validation_pallet", f"Паллета {idx}: {err_detail}."
        validated.append(row)

    if sort_assemble_pallets_by_number is not None:
        validated = sort_assemble_pallets_by_number(validated)

    try:
        wb = Workbook()
        ws = wb.active
        if ws is None:
            return None, "xlsx_build", _arnest_pallet_pdf_error_message("xlsx_build")
        ws.title = "Паллеты"

        thin = Side(style="thin", color="5A6578")
        grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_fill = PatternFill("solid", fgColor="147A6E")
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14, color="153045")
        zebra_fill = PatternFill("solid", fgColor="F0F4F8")

        title_parts = ["Арнест Юнирусь — свод по паллетам"]
        if ship_date_raw:
            title_parts.append(f"отгрузка: {ship_date_raw}")
        if order_id_int is not None:
            title_parts.append(f"заказ №{order_id_int}")
        title_text = " · ".join(title_parts)

        last_col = 9
        last_letter = get_column_letter(last_col)
        ws.merge_cells(f"A1:{last_letter}1")
        tcell = ws["A1"]
        tcell.value = title_text
        tcell.font = title_font
        tcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 28

        headers = (
            "Номер паллеты",
            "Артикул",
            "Наименование",
            "Партия",
            "Дата изготовления",
            "Срок годности",
            "Вес брутто, кг",
            "Коробок",
            "Штук",
        )
        hdr_row = 3
        for col_idx, text in enumerate(headers, start=1):
            c = ws.cell(row=hdr_row, column=col_idx, value=text)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = grid_border

        widths = (14, 18, 40, 16, 14, 14, 14, 11, 11)
        for i, wch in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = wch

        data_start = 4
        total_weight = 0.0
        for i, row in enumerate(validated, start=1):
            r = data_start + i - 1
            mfg_raw = _arnest_mfg_date_raw(row)
            exp_raw = _arnest_exp_date_raw(row)
            mfg_disp = _arnest_six_digits_display_ddmmyy(mfg_raw) or "—"
            exp_disp = _arnest_six_digits_display_ddmmyy(exp_raw) or "—"
            try:
                w_kg = float(row.get("pallet_weight_kg", 0) or 0)
            except (TypeError, ValueError):
                w_kg = 0.0
            if not math.isfinite(w_kg) or w_kg < 0:
                w_kg = 0.0
            total_weight += w_kg
            try:
                boxes_f = float(row.get("pallet_boxes_qty", row.get("pallet_qty", 0)) or 0)
            except (TypeError, ValueError):
                boxes_f = 0.0
            try:
                units_f = float(row.get("units_pc", row.get("unitsPc", 0)) or 0)
            except (TypeError, ValueError):
                units_f = 0.0
            boxes_disp = _arnest_format_boxes_qty_for_barcode(boxes_f)
            units_disp = _arnest_format_boxes_qty_for_barcode(units_f)
            pallet_no = str(row.get("pallet_number", row.get("palletNumber", "")) or "").strip()
            article = str(row.get("article", "") or "").strip()
            name = str(row.get("name", row.get("product_name", "")) or "").strip()
            batch = str(row.get("batch_number", row.get("batchNumber", "")) or "").strip()

            vals = (
                pallet_no,
                article,
                name,
                batch,
                mfg_disp,
                exp_disp,
                w_kg,
                boxes_disp,
                units_disp,
            )
            for col_idx, val in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=col_idx, value=val)
                cell.border = grid_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if i % 2 == 0:
                    cell.fill = zebra_fill
            ws.cell(row=r, column=7).number_format = "0.000"

        last_data = data_start + len(validated) - 1
        sum_row = last_data + 1
        ws.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=6)
        sum_label = ws.cell(row=sum_row, column=1, value="Итого вес брутто, кг:")
        sum_label.font = Font(bold=True)
        sum_label.alignment = Alignment(horizontal="right", vertical="center")
        sum_label.border = grid_border
        c_w = ws.cell(row=sum_row, column=7, value=round(total_weight, 3))
        c_w.font = Font(bold=True)
        c_w.number_format = "0.000"
        c_w.border = grid_border
        ws.cell(row=sum_row, column=8, value="").border = grid_border
        ws.cell(row=sum_row, column=9, value="").border = grid_border

        ws.freeze_panes = ws.cell(row=data_start, column=1)
        ws.auto_filter.ref = f"A{hdr_row}:{last_letter}{last_data}"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), None, None
    except Exception as exc:
        return None, "xlsx_build", f"{_arnest_pallet_pdf_error_message('xlsx_build')} ({exc})"


def _wrap_pdf_pjl_a4_tray1(pdf: bytes) -> bytes:
    """PJL перед/после PDF для JetDirect: A4, лоток 1, язык PDF."""
    uel = b"\x1b%-12345X"
    crlf = b"\r\n"
    pjl = (
        b'@PJL JOB NAME = "Paletlist pallets"'
        + crlf
        + b"@PJL SET PAPER = A4"
        + crlf
        + b"@PJL SET MEDIASOURCE = TRAY1"
        + crlf
        + b"@PJL ENTER LANGUAGE = PDF"
        + crlf
    )
    end = crlf + b"@PJL EOJ" + crlf + uel
    return uel + pjl + pdf + end


def _get_raw_printer_host_port() -> tuple[str, int | None, str | None]:
    """Хост и порт JetDirect; при ошибке порта — (host, None, текст ошибки)."""
    host = (os.environ.get("PALLET_RAW_PRINT_HOST") or "192.168.1.196").strip()
    port_raw = (os.environ.get("PALLET_RAW_PRINT_PORT") or "9100").strip()
    try:
        return host, int(port_raw), None
    except ValueError:
        return host, None, "Некорректный PALLET_RAW_PRINT_PORT"


def check_raw_pallet_printer_reachable() -> tuple[bool, str, str]:
    """Проверка TCP до принтера без отправки данных на печать.
    Возвращает (успех, \"host:port\", сообщение для пользователя)."""
    host, port, cfg_err = _get_raw_printer_host_port()
    addr = f"{host}:{port}" if port is not None else f"{host}:?"
    if cfg_err:
        return False, addr, cfg_err
    ping_timeout = float((os.environ.get("PALLET_RAW_PRINT_PING_TIMEOUT_SEC") or "5").strip() or "5")
    try:
        with socket.create_connection((host, port), timeout=ping_timeout):
            pass
    except OSError as exc:
        return False, addr, str(exc)
    return True, addr, "TCP-соединение установлено (данные на печать не отправлялись)."


def send_pdf_to_raw_lan_printer(pdf: bytes) -> tuple[bool, str, int]:
    """Отправка PDF на принтер по TCP:9100 (сервер должен видеть IP принтера в LAN).
    Обёртка PJL A4/лоток1 — по умолчанию включена (PALLET_RAW_PRINT_PJL=0 — только сырые байты PDF)."""
    host, port, cfg_err = _get_raw_printer_host_port()
    if cfg_err or port is None:
        return False, cfg_err or "Ошибка настройки порта принтера", 0
    timeout = float((os.environ.get("PALLET_RAW_PRINT_TIMEOUT_SEC") or "25").strip() or "25")
    use_pjl = (os.environ.get("PALLET_RAW_PRINT_PJL", "1") or "1").strip().lower() not in (
        "0",
        "false",
        "no",
        "off",
    )
    payload = _wrap_pdf_pjl_a4_tray1(pdf) if use_pjl else pdf
    try:
        with socket.create_connection((host, port), timeout=timeout) as sock:
            sock.sendall(payload)
    except OSError as exc:
        return False, f"{host}:{port}: {exc}", 0
    return True, f"{host}:{port}", len(payload)


def _parse_orders_detail_id(path: str):
    """Для /api/orders/12 возвращает 12; для списка или чужих путей — None."""
    p = _norm_api_path(path)
    prefix = "/api/orders/"
    if not p.startswith(prefix):
        return None
    tail = p[len(prefix) :]
    if not tail or "/" in tail:
        return None
    try:
        return int(tail)
    except ValueError:
        return None


def _parse_orders_packing_sheets_html_id(path: str) -> int | None:
    """Для GET /api/orders/12/packing-sheets-html возвращает 12."""
    p = _norm_api_path(path)
    prefix = "/api/orders/"
    suffix = "/packing-sheets-html"
    if not p.startswith(prefix) or not p.endswith(suffix):
        return None
    mid = p[len(prefix) : -len(suffix)]
    if not mid or "/" in mid:
        return None
    try:
        return int(mid)
    except ValueError:
        return None


def _parse_orders_packing_sheets_pdf_id(path: str) -> int | None:
    """Для GET /api/orders/12/packing-sheets-pdf возвращает 12."""
    p = _norm_api_path(path)
    prefix = "/api/orders/"
    suffix = "/packing-sheets-pdf"
    if not p.startswith(prefix) or not p.endswith(suffix):
        return None
    mid = p[len(prefix) : -len(suffix)]
    if not mid or "/" in mid:
        return None
    try:
        return int(mid)
    except ValueError:
        return None


def _parse_orders_lab_ship_id(path: str) -> int | None:
    """Для POST /api/orders/12/lab-ship возвращает 12."""
    p = _norm_api_path(path)
    prefix = "/api/orders/"
    suffix = "/lab-ship"
    if not p.startswith(prefix) or not p.endswith(suffix):
        return None
    mid = p[len(prefix) : -len(suffix)]
    if not mid or "/" in mid:
        return None
    try:
        return int(mid)
    except ValueError:
        return None


def _parse_orders_assemble_presence_id(path: str) -> int | None:
    """Для /api/orders/12/assemble-presence возвращает 12."""
    p = _norm_api_path(path)
    prefix = "/api/orders/"
    suffix = "/assemble-presence"
    if not p.startswith(prefix) or not p.endswith(suffix):
        return None
    mid = p[len(prefix) : -len(suffix)]
    if not mid or "/" in mid:
        return None
    try:
        return int(mid)
    except ValueError:
        return None


def _parse_orders_assemble_sync_id(path: str) -> int | None:
    """Для GET /api/orders/12/assemble-sync возвращает 12."""
    p = _norm_api_path(path)
    prefix = "/api/orders/"
    suffix = "/assemble-sync"
    if not p.startswith(prefix) or not p.endswith(suffix):
        return None
    mid = p[len(prefix) : -len(suffix)]
    if not mid or "/" in mid:
        return None
    try:
        return int(mid)
    except ValueError:
        return None


def _parse_order_messages_id(path: str) -> int | None:
    """Для /api/orders/12/messages возвращает 12."""
    p = _norm_api_path(path)
    prefix = "/api/orders/"
    suffix = "/messages"
    if not p.startswith(prefix) or not p.endswith(suffix):
        return None
    mid = p[len(prefix) : -len(suffix)]
    if not mid or "/" in mid:
        return None
    try:
        return int(mid)
    except ValueError:
        return None


def _parse_order_message_detail_id(path: str) -> tuple[int, int] | None:
    """Для /api/orders/12/messages/34 возвращает (12, 34)."""
    p = _norm_api_path(path)
    prefix = "/api/orders/"
    marker = "/messages/"
    if not p.startswith(prefix) or marker not in p:
        return None
    rest = p[len(prefix) :]
    order_part, msg_part = rest.split(marker, 1)
    if not order_part or "/" in order_part or not msg_part or "/" in msg_part:
        return None
    try:
        return int(order_part), int(msg_part)
    except ValueError:
        return None


def _path_allowed_for_order_monitoring(path: str, method: str) -> bool:
    """Только просмотр заказов и чат — для perm order_monitoring без orders."""
    m = (method or "GET").upper()
    if _is_orders_list_path(path) and m == "GET":
        return True
    if _parse_orders_detail_id(path) is not None and m == "GET":
        return True
    if _parse_order_messages_id(path) is not None and m in ("GET", "POST"):
        return True
    if _parse_order_message_detail_id(path) is not None and m == "DELETE":
        return True
    return False


def _path_needs_orders_permission(path: str) -> bool:
    if (
        _is_orders_list_path(path)
        or _is_orders_import_excel_path(path)
        or _is_orders_import_drogeri_excel_path(path)
        or _is_orders_batches_export_pdf_path(path)
    ):
        return True
    if _parse_orders_detail_id(path) is not None:
        return True
    if _parse_orders_packing_sheets_html_id(path) is not None:
        return True
    if _parse_orders_packing_sheets_pdf_id(path) is not None:
        return True
    if _parse_orders_lab_ship_id(path) is not None:
        return True
    if _parse_orders_assemble_presence_id(path) is not None:
        return True
    if _parse_orders_assemble_sync_id(path) is not None:
        return True
    if _parse_order_messages_id(path) is not None:
        return True
    if _parse_order_message_detail_id(path) is not None:
        return True
    return False


def _path_needs_nomenclature_permission(path: str) -> bool:
    norm = _norm_api_path(path)
    return norm == "/api/nomenclature" or norm.startswith("/api/nomenclature/")


def _presence_label_from_user_agent(user_agent: str) -> str:
    u = (user_agent or "").lower()
    if "ipad" in u or "tablet" in u:
        return "планшет"
    if "mobile" in u or "android" in u or "iphone" in u:
        return "телефон"
    return "компьютер"


def _prune_assemble_presence_locked(now: float) -> None:
    cutoff = now - ASSEMBLE_PRESENCE_TTL_SEC
    for oid in list(_ASSEMBLE_PRESENCE.keys()):
        bucket = _ASSEMBLE_PRESENCE[oid]
        for cid in list(bucket.keys()):
            if bucket[cid]["ts"] < cutoff:
                del bucket[cid]
        if not bucket:
            del _ASSEMBLE_PRESENCE[oid]


def fetch_order_assemble_sync_fields(
    order_id: int, since_rev: int | None = None
) -> dict | None:
    """Лёгкий снимок сборки для live-sync (без позиций заказа).

    Если since_rev совпадает с текущей ревизией — возвращает unchanged без assemble_state.
    """
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        row = cur.execute(
            """
            SELECT assemble_state, assemble_revision, assemble_state_updated_at,
                   assembled_percent, order_readiness,
                   COALESCE(lab_sscc_shipped, 0) AS lab_sscc_shipped
            FROM orders WHERE id = ?
            """,
            (order_id,),
        ).fetchone()
        con.close()
    if not row:
        return None
    current_rev = max(0, int(row["assemble_revision"] or 0))
    updated_at = (row["assemble_state_updated_at"] or "").strip()
    readiness = _order_readiness_from_row(row)
    base = {
        "assemble_revision": current_rev,
        "assemble_state_updated_at": updated_at,
        "order_readiness": readiness,
        "assembled_percent": max(0, min(100, int(row["assembled_percent"] or 0))),
    }
    if since_rev is not None and int(since_rev) == current_rev:
        return {**base, "unchanged": True}
    return {
        **base,
        "unchanged": False,
        "assemble_state": _assemble_state_cell_to_api(row["assemble_state"]),
    }


def touch_assemble_presence(
    order_id: int,
    client_id: str,
    user_agent: str = "",
    *,
    since_rev: int | None = None,
    include_assemble_sync: bool = False,
) -> dict:
    """Отметить присутствие в сборке; опционально вернуть снимок assemble_state."""
    cid = (client_id or "").strip()
    if not cid or len(cid) > 80:
        return {"error": "validation", "message": "Укажите client_id (до 80 символов)."}
    now = time.time()
    label = _presence_label_from_user_agent(user_agent)
    with PRESENCE_LOCK:
        _prune_assemble_presence_locked(now)
        bucket = _ASSEMBLE_PRESENCE.setdefault(int(order_id), {})
        bucket[cid] = {"ts": now, "label": label}
        others = [
            {"client_id": k, "label": v.get("label") or "устройство"}
            for k, v in bucket.items()
            if k != cid
        ]
    out: dict = {"ok": True, "order_id": int(order_id), "others": others}
    if include_assemble_sync:
        sync = fetch_order_assemble_sync_fields(int(order_id), since_rev)
        if sync is None:
            return {"error": "not_found", "message": "Заказ не найден."}
        out.update(sync)
    return out


def leave_assemble_presence(order_id: int, client_id: str) -> dict:
    cid = (client_id or "").strip()
    if not cid:
        return {"error": "validation", "message": "Укажите client_id."}
    with PRESENCE_LOCK:
        bucket = _ASSEMBLE_PRESENCE.get(int(order_id))
        if bucket and cid in bucket:
            del bucket[cid]
        if bucket is not None and not bucket:
            del _ASSEMBLE_PRESENCE[int(order_id)]
    return {"ok": True}


def get_connection():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA foreign_keys = ON")
    return con


_DB_BACKUP_REQUIRED_TABLES = (
    "products",
    "orders",
    "order_items",
    "app_users",
)
_DB_RESTORE_MAX_BYTES = 200 * 1024 * 1024


def export_database_bytes():
    """Согласованный снимок warehouse.db для скачивания администратором."""
    with DB_LOCK:
        if not DB_PATH.is_file():
            return None, {
                "error": "not_found",
                "message": "Файл базы данных не найден на сервере.",
            }
        con = sqlite3.connect(str(DB_PATH))
        try:
            con.execute("PRAGMA wal_checkpoint(TRUNCATE)")
        except sqlite3.Error:
            pass
        finally:
            con.close()
        return DB_PATH.read_bytes(), None


def cleanup_database_backup_files(keep_backup_name: str | None = None) -> list[str]:
    """Оставить только warehouse.db и один бэкап перед загрузкой (если указан)."""
    keep = {"warehouse.db"}
    if keep_backup_name:
        keep.add(Path(keep_backup_name).name)
    removed: list[str] = []
    for path in BASE_DIR.glob("warehouse*"):
        name = path.name
        if name in keep:
            continue
        # Только файлы-копии БД рядом с приложением
        if not (
            name.startswith("warehouse.before-restore-")
            or name.startswith("warehouse.bad-before-fix-")
            or name.startswith("warehouse.restore-")
            or name.startswith("warehouse.from-")
            or name.endswith(".tmp")
            or (name.startswith("warehouse.") and name.endswith(".db") and name != "warehouse.db")
        ):
            continue
        try:
            if path.is_file():
                path.unlink()
                removed.append(name)
        except OSError:
            pass
    return removed


def restore_database_from_bytes(data: bytes) -> dict:
    """Заменить warehouse.db загруженным файлом (с бэкапом текущей БД)."""
    if not isinstance(data, (bytes, bytearray)):
        return {"error": "validation", "message": "Файл не передан."}
    raw = bytes(data)
    if len(raw) < 100:
        return {"error": "validation", "message": "Файл слишком маленький."}
    if len(raw) > _DB_RESTORE_MAX_BYTES:
        return {
            "error": "validation",
            "message": "Файл слишком большой (лимит 200 МБ).",
        }
    if not raw.startswith(b"SQLite format 3\x00"):
        return {
            "error": "validation",
            "message": "Нужен файл SQLite базы данных (.db).",
        }

    fd, tmp_name = tempfile.mkstemp(prefix="paletlist-restore-", suffix=".db")
    tmp_path = Path(tmp_name)
    staging = None
    bak_name = ""
    try:
        os.close(fd)
        tmp_path.write_bytes(raw)
        con = sqlite3.connect(str(tmp_path))
        try:
            tables = {
                str(r[0])
                for r in con.execute(
                    "SELECT name FROM sqlite_master WHERE type='table'"
                ).fetchall()
            }
            missing = [t for t in _DB_BACKUP_REQUIRED_TABLES if t not in tables]
            if missing:
                return {
                    "error": "validation",
                    "message": "В файле нет нужных таблиц: " + ", ".join(missing),
                }
            check = con.execute("PRAGMA integrity_check").fetchone()
            if not check or str(check[0]).lower() != "ok":
                return {
                    "error": "validation",
                    "message": "Файл базы данных повреждён.",
                }
        finally:
            con.close()

        stamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
        bak_name = f"warehouse.before-restore-{stamp}.db"
        bak_path = DB_PATH.with_name(bak_name)
        staging = DB_PATH.with_name(f"warehouse.restore-{stamp}.tmp")
        with DB_LOCK:
            if DB_PATH.is_file():
                shutil.copy2(DB_PATH, bak_path)
            else:
                bak_name = ""
            shutil.copy2(tmp_path, staging)
            os.replace(staging, DB_PATH)
            staging = None
            cleanup_database_backup_files(bak_name or None)
    except OSError as exc:
        return {"error": "io", "message": f"Не удалось заменить БД: {exc}"}
    except sqlite3.Error as exc:
        return {"error": "validation", "message": f"Некорректный файл БД: {exc}"}
    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except OSError:
            pass
        if staging is not None:
            try:
                staging.unlink(missing_ok=True)
            except OSError:
                pass

    try:
        init_db()
    except Exception as exc:  # noqa: BLE001
        return {
            "error": "database",
            "message": f"БД заменена, но миграция не завершилась: {exc}",
            "backup": bak_name,
        }
    return {"ok": True, "backup": bak_name, "size": len(raw)}


def _is_valid_ipv4(ip: str) -> bool:
    parts = (ip or "").strip().split(".")
    if len(parts) != 4:
        return False
    try:
        return all(0 <= int(p) <= 255 for p in parts)
    except ValueError:
        return False


def _secrets_match(provided: str | None, expected: str) -> bool:
    if not expected or not provided:
        return False
    a = str(provided).strip().encode("utf-8")
    b = str(expected).strip().encode("utf-8")
    if len(a) != len(b):
        return False
    return secrets.compare_digest(a, b)


def allow_ssh_ip(ip: str) -> dict:
    """Разрешить SSH с IP (UFW), сохранив закреплённые адреса."""
    ip = (ip or "").strip()
    if not _is_valid_ipv4(ip):
        return {"error": "validation", "message": "Некорректный IPv4-адрес."}
    if ip in ("0.0.0.0", "127.0.0.1"):
        return {"error": "validation", "message": "Этот адрес нельзя добавлять."}

    pinned_path = Path("/etc/paletlist/ssh-pinned-ips.txt")
    pinned: list[str] = []
    if pinned_path.is_file():
        for line in pinned_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and _is_valid_ipv4(line):
                if line not in pinned:
                    pinned.append(line)
    allowed = list(pinned)
    if ip not in allowed:
        allowed.append(ip)
    allowed_set = set(allowed)

    def _run(args: list[str]) -> subprocess.CompletedProcess:
        return subprocess.run(
            args,
            capture_output=True,
            text=True,
            timeout=30,
            check=False,
        )

    try:
        # Add needed IPs first (safe even if previous SSH rules were wiped).
        for addr in allowed:
            proc = _run(
                [
                    "ufw",
                    "allow",
                    "from",
                    addr,
                    "to",
                    "any",
                    "port",
                    "22",
                    "proto",
                    "tcp",
                    "comment",
                    "SSH allowed",
                ]
            )
            if proc.returncode != 0:
                detail = (proc.stderr or proc.stdout or "").strip().lower()
                if "existing" not in detail and "skipping" not in detail:
                    return {
                        "error": "firewall",
                        "message": (proc.stderr or proc.stdout or "").strip()
                        or f"Не удалось добавить {addr}",
                    }

        # Remove only stale SSH source IPs not in allow-list.
        for _ in range(40):
            status = _run(["ufw", "status", "numbered"])
            stale_num = None
            for line in (status.stdout or "").splitlines():
                m = re.match(r"^\[\s*(\d+)\]\s+22/tcp\b(.*)$", line.strip())
                if not m:
                    continue
                rest = m.group(2)
                ip_match = re.search(r"\b(\d{1,3}(?:\.\d{1,3}){3})\b", rest)
                src_ip = ip_match.group(1) if ip_match else None
                if src_ip and src_ip not in allowed_set:
                    stale_num = int(m.group(1))
                    break
                if not src_ip and "Anywhere" in rest:
                    stale_num = int(m.group(1))
                    break
            if stale_num is None:
                break
            _run(["ufw", "--force", "delete", str(stale_num)])

        _run(["ufw", "allow", "80/tcp"])
        _run(["ufw", "allow", "443/tcp"])
    except (OSError, subprocess.TimeoutExpired) as exc:
        return {"error": "firewall", "message": f"Не удалось обновить UFW: {exc}"}

    return {
        "ok": True,
        "allowed_ip": ip,
        "ssh_ips": allowed,
        "message": f"SSH разрешён для {ip}. Можно подключаться.",
    }


def start_background_deploy() -> dict:
    """Запустить деплой в фоне (git pull + restart), чтобы ответ успел уйти."""
    script = Path("/usr/local/sbin/paletlist-deploy")
    if not script.is_file():
        # fallback to repo script after pull is impossible if not deployed yet
        alt = BASE_DIR / "scripts" / "pull-on-server.sh"
        if alt.is_file():
            script = alt
        else:
            return {
                "error": "config",
                "message": "Скрипт деплоя не найден на сервере.",
            }
    log_path = Path("/var/log/paletlist-deploy.log")

    def _run() -> None:
        time.sleep(0.4)
        try:
            with log_path.open("a", encoding="utf-8") as log:
                log.write(
                    f"\n===== deploy {datetime.datetime.now(datetime.timezone.utc).isoformat()} =====\n"
                )
                subprocess.run(
                    ["bash", str(script)],
                    stdout=log,
                    stderr=subprocess.STDOUT,
                    check=False,
                )
        except Exception as exc:  # noqa: BLE001
            try:
                with log_path.open("a", encoding="utf-8") as log:
                    log.write(f"deploy failed: {exc}\n")
            except OSError:
                pass

    threading.Thread(target=_run, daemon=True).start()
    return {
        "ok": True,
        "started": True,
        "message": "Деплой запущен. Лог: /var/log/paletlist-deploy.log",
    }


def _read_proc_meminfo() -> dict:
    result = {}
    try:
        for line in Path("/proc/meminfo").read_text(encoding="utf-8").splitlines():
            if ":" not in line:
                continue
            key, raw = line.split(":", 1)
            parts = raw.strip().split()
            if not parts:
                continue
            try:
                kb = int(parts[0])
            except ValueError:
                continue
            result[key] = kb * 1024
    except OSError:
        pass
    return result


def _read_proc_stat_cpu():
    try:
        line = Path("/proc/stat").read_text(encoding="utf-8").splitlines()[0]
    except (OSError, IndexError):
        return None
    parts = line.split()
    if not parts or parts[0] != "cpu":
        return None
    try:
        vals = [int(x) for x in parts[1:]]
    except ValueError:
        return None
    if len(vals) < 4:
        return None
    idle = vals[3] + (vals[4] if len(vals) > 4 else 0)
    total = sum(vals)
    return idle, total


def _cpu_usage_percent(sample_sec: float = 0.12) -> float | None:
    first = _read_proc_stat_cpu()
    if not first:
        return None
    time.sleep(max(0.05, min(0.5, sample_sec)))
    second = _read_proc_stat_cpu()
    if not second:
        return None
    idle1, total1 = first
    idle2, total2 = second
    d_total = total2 - total1
    d_idle = idle2 - idle1
    if d_total <= 0:
        return 0.0
    used = 100.0 * (1.0 - (d_idle / d_total))
    return round(max(0.0, min(100.0, used)), 1)


def _loadavg() -> list[float]:
    try:
        raw = Path("/proc/loadavg").read_text(encoding="utf-8").split()
        return [round(float(raw[i]), 2) for i in range(3)]
    except (OSError, ValueError, IndexError):
        try:
            return [round(float(x), 2) for x in os.getloadavg()]
        except (OSError, AttributeError):
            return []


def _uptime_seconds() -> float | None:
    try:
        return float(Path("/proc/uptime").read_text(encoding="utf-8").split()[0])
    except (OSError, ValueError, IndexError):
        return None


def _systemctl_state(unit: str) -> dict:
    def _run(args: list[str]) -> str:
        try:
            proc = subprocess.run(
                args,
                capture_output=True,
                text=True,
                timeout=3,
                check=False,
            )
            return (proc.stdout or "").strip() or (proc.stderr or "").strip()
        except (OSError, subprocess.TimeoutExpired):
            return "unknown"

    active = _run(["systemctl", "is-active", unit]) or "unknown"
    enabled = _run(["systemctl", "is-enabled", unit]) or "unknown"
    return {"unit": unit, "active": active, "enabled": enabled}


def _listening_ports_summary() -> list[dict]:
    ports = []
    try:
        proc = subprocess.run(
            ["ss", "-tlnH"],
            capture_output=True,
            text=True,
            timeout=3,
            check=False,
        )
        lines = (proc.stdout or "").splitlines()
    except (OSError, subprocess.TimeoutExpired):
        lines = []
    seen = set()
    for line in lines:
        parts = line.split()
        if len(parts) < 4:
            continue
        local = parts[3]
        if ":" not in local:
            continue
        port_s = local.rsplit(":", 1)[-1]
        try:
            port = int(port_s)
        except ValueError:
            continue
        if port in seen:
            continue
        seen.add(port)
        ports.append({"port": port, "address": local})
    ports.sort(key=lambda row: row["port"])
    return ports[:40]


def _disk_usage_for(path: str) -> dict | None:
    try:
        usage = shutil.disk_usage(path)
    except OSError:
        return None
    total = int(usage.total)
    used = int(usage.used)
    free = int(usage.free)
    pct = round((used / total) * 100.0, 1) if total else 0.0
    return {
        "path": path,
        "total_bytes": total,
        "used_bytes": used,
        "free_bytes": free,
        "used_percent": pct,
    }


def _database_status_snapshot() -> dict:
    info = {
        "path": str(DB_PATH),
        "exists": DB_PATH.is_file(),
        "size_bytes": 0,
        "counts": {},
    }
    if not DB_PATH.is_file():
        return info
    try:
        info["size_bytes"] = int(DB_PATH.stat().st_size)
    except OSError:
        pass
    try:
        with DB_LOCK:
            con = get_connection()
            try:
                for table in (
                    "products",
                    "orders",
                    "order_items",
                    "app_users",
                    "feedback_threads",
                ):
                    try:
                        info["counts"][table] = int(
                            con.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
                        )
                    except sqlite3.Error:
                        info["counts"][table] = None
            finally:
                con.close()
    except sqlite3.Error as exc:
        info["error"] = str(exc)
    return info


def collect_server_status() -> dict:
    """Снимок ключевых показателей сервера для админ-мониторинга."""
    mem = _read_proc_meminfo()
    mem_total = int(mem.get("MemTotal") or 0)
    mem_available = int(mem.get("MemAvailable") or mem.get("MemFree") or 0)
    mem_used = max(0, mem_total - mem_available) if mem_total else 0
    mem_pct = round((mem_used / mem_total) * 100.0, 1) if mem_total else 0.0
    swap_total = int(mem.get("SwapTotal") or 0)
    swap_free = int(mem.get("SwapFree") or 0)
    swap_used = max(0, swap_total - swap_free) if swap_total else 0

    hostname = socket.gethostname()
    try:
        uname = os.uname()
        os_summary = f"{uname.sysname} {uname.release}"
    except AttributeError:
        os_summary = "unknown"

    cpu_count = os.cpu_count() or 0
    return {
        "ok": True,
        "generated_at": datetime.datetime.now(datetime.timezone.utc).strftime(
            "%Y-%m-%dT%H:%M:%SZ"
        ),
        "hostname": hostname,
        "os": os_summary,
        "uptime_seconds": _uptime_seconds(),
        "loadavg": _loadavg(),
        "cpu": {
            "count": cpu_count,
            "usage_percent": _cpu_usage_percent(),
        },
        "memory": {
            "total_bytes": mem_total,
            "used_bytes": mem_used,
            "available_bytes": mem_available,
            "used_percent": mem_pct,
            "swap_total_bytes": swap_total,
            "swap_used_bytes": swap_used,
        },
        "disks": [
            row
            for row in (
                _disk_usage_for("/"),
                _disk_usage_for(str(BASE_DIR)),
            )
            if row
        ],
        "services": [
            _systemctl_state("nginx"),
            _systemctl_state("paletlist-api.service"),
            _systemctl_state("ssh"),
        ],
        "listening_ports": _listening_ports_summary(),
        "database": _database_status_snapshot(),
        "process_count": _process_count(),
    }


def _process_count() -> int | None:
    try:
        return sum(1 for p in Path("/proc").iterdir() if p.name.isdigit())
    except OSError:
        return None


def _table_columns(cur, table: str):
    return {row[1] for row in cur.execute(f"PRAGMA table_info({table})").fetchall()}


def _migrate_products(cur):
    cols = _table_columns(cur, "products")
    if "pieces_in_box" not in cols:
        cur.execute(
            "ALTER TABLE products ADD COLUMN pieces_in_box INTEGER NOT NULL DEFAULT 0"
        )
    if "row_layout" not in cols:
        cur.execute(
            "ALTER TABLE products ADD COLUMN row_layout INTEGER NOT NULL DEFAULT 0"
        )
    if "max_rows" not in cols:
        cur.execute("ALTER TABLE products ADD COLUMN max_rows INTEGER NOT NULL DEFAULT 0")
    if "box_weight" not in cols:
        cur.execute(
            "ALTER TABLE products ADD COLUMN box_weight REAL NOT NULL DEFAULT 0"
        )
    if "volume_ml" not in cols:
        cur.execute(
            "ALTER TABLE products ADD COLUMN volume_ml REAL NOT NULL DEFAULT 0"
        )
    if "volume_unit" not in cols:
        cur.execute(
            "ALTER TABLE products ADD COLUMN volume_unit TEXT NOT NULL DEFAULT 'ml'"
        )
    if "pieces_per_set" not in cols:
        cur.execute(
            "ALTER TABLE products ADD COLUMN pieces_per_set INTEGER NOT NULL DEFAULT 1"
        )
    if "sets_in_box" not in cols:
        cur.execute(
            "ALTER TABLE products ADD COLUMN sets_in_box INTEGER NOT NULL DEFAULT 1"
        )
        for row in cur.execute(
            "SELECT id, pieces_in_box, pieces_per_set FROM products"
        ).fetchall():
            pid = int(row["id"])
            pib = int(row["pieces_in_box"] or 0)
            pps = max(1, int(row["pieces_per_set"] or 1))
            if pps <= 1 or pib <= 0:
                sib = 1
            else:
                sib = pib // pps
                if sib < 1:
                    sib = 1
                if pps > 0 and pib % pps != 0:
                    sib = max(1, int(round(float(pib) / pps)))
            cur.execute(
                "UPDATE products SET sets_in_box = ? WHERE id = ?",
                (sib, pid),
            )
        for row in cur.execute(
            "SELECT id, pieces_in_box, sets_in_box FROM products"
        ).fetchall():
            pid = int(row["id"])
            pib = int(row["pieces_in_box"] or 0)
            sib = max(1, int(row["sets_in_box"] or 1))
            if sib <= 1 or pib <= 0:
                pps = 1
            elif pib % sib == 0:
                pps = max(1, pib // sib)
            else:
                pps = max(1, int(round(float(pib) / sib)))
            cur.execute(
                "UPDATE products SET pieces_per_set = ? WHERE id = ?",
                (pps, pid),
            )


# Хвосты наименований: «, шт» / «, набор» и варианты (запятая ASCII/полношир./„ , опц. точка после шт).
_NAME_SUFFIX_TRAIL_PATTERNS = (
    re.compile(r"\s*[\u002C\uFF0C\u201A]\s*набор\.?\s*$", re.IGNORECASE),
    re.compile(r"\s*[\u002C\uFF0C\u201A]\s*шт\.?\s*$", re.IGNORECASE),
)


def _strip_trailing_name_suffix(name: str) -> str:
    """Убирает хвост «, шт» / «, набор» (и типичные варианты написания) в конце наименования."""
    s = unicodedata.normalize("NFKC", name or "").rstrip()
    changed = True
    while changed and s:
        changed = False
        for pat in _NAME_SUFFIX_TRAIL_PATTERNS:
            nxt = pat.sub("", s).rstrip()
            if nxt != s:
                s = nxt
                changed = True
                break
    return s


def _migrate_strip_trailing_name_suffixes(cur):
    """Идемпотентно чистит products.name и order_items.name от суффиксов «, шт» / «, набор»."""
    for row in cur.execute("SELECT id, name FROM products"):
        old = row["name"] if row["name"] is not None else ""
        new = _strip_trailing_name_suffix(old)
        if new != old:
            cur.execute(
                "UPDATE products SET name = ? WHERE id = ?",
                (new, int(row["id"])),
            )
    if not cur.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name='order_items' LIMIT 1"
    ).fetchone():
        return
    for row in cur.execute("SELECT id, name FROM order_items"):
        old = row["name"] if row["name"] is not None else ""
        new = _strip_trailing_name_suffix(old)
        if new != old:
            cur.execute(
                "UPDATE order_items SET name = ? WHERE id = ?",
                (new, int(row["id"])),
            )


def _normalize_packaging(pieces_in_box: int, sets_in_box: int):
    """Возвращает (pib, sib, pps) или dict с error/message при ошибке валидации."""
    pib = max(0, int(pieces_in_box))
    sib = max(1, int(sets_in_box))
    if sib <= 1:
        return (pib, sib, 1)
    if pib <= 0:
        return (pib, sib, 1)
    if pib % sib != 0:
        return {
            "error": "validation",
            "message": "Штук в коробке должно делиться на количество наборов без остатка.",
        }
    pps = max(1, pib // sib)
    return (pib, sib, pps)


def _computed_pieces_per_set_for_api(pib: int, sib: int):
    sib = max(1, int(sib))
    pib = int(pib)
    if sib <= 1 or pib <= 0:
        return 1
    if pib % sib == 0:
        return max(1, pib // sib)
    return round(pib / sib, 2)


def init_db():
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS products (
              id INTEGER PRIMARY KEY,
              article TEXT NOT NULL DEFAULT '',
              name TEXT NOT NULL DEFAULT '',
              pieces_in_box INTEGER NOT NULL DEFAULT 0,
              sets_in_box INTEGER NOT NULL DEFAULT 1,
              pieces_per_set INTEGER NOT NULL DEFAULT 1,
              row_layout INTEGER NOT NULL DEFAULT 0,
              max_rows INTEGER NOT NULL DEFAULT 0,
              box_weight REAL NOT NULL DEFAULT 0,
              volume_ml REAL NOT NULL DEFAULT 0
            )
            """
        )
        _migrate_products(cur)
        cur.execute(
            """
            DELETE FROM products
            WHERE TRIM(COALESCE(article, '')) = ''
              AND TRIM(COALESCE(name, '')) = ''
            """
        )
        count = cur.execute("SELECT COUNT(*) FROM products").fetchone()[0]
        if count == 0 and JSON_SEED_PATH.exists():
            seed_items = json.loads(JSON_SEED_PATH.read_text(encoding="utf-8"))
            rows = []
            for item in seed_items:
                try:
                    pib = int(item.get("pieces_in_box") or 0)
                    sib = max(1, int(item.get("sets_in_box") or 1))
                    norm = _normalize_packaging(pib, sib)
                    if isinstance(norm, dict):
                        pib, sib, pps = pib, 1, 1
                    else:
                        pib, sib, pps = norm
                    rows.append(
                        (
                            int(item.get("id")),
                            (item.get("article") or "").strip(),
                            _strip_trailing_name_suffix((item.get("name") or "").strip()),
                            pib,
                            sib,
                            pps,
                            int(item.get("row_layout") or 0),
                            int(item.get("max_rows") or 0),
                            float(item.get("box_weight") or 0),
                            float(item.get("volume_ml") or 0),
                        )
                    )
                except (TypeError, ValueError):
                    continue
            if rows:
                cur.executemany(
                    """
                    INSERT INTO products (
                      id, article, name,
                      pieces_in_box, sets_in_box, pieces_per_set, row_layout, max_rows, box_weight, volume_ml
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    rows,
                )
        _init_orders_table(cur)
        _init_app_users_table(cur)
        _init_feedback_tables(cur)
        _init_order_chat_tables(cur)
        _init_unique_clients_table(cur)
        _migrate_strip_trailing_name_suffixes(cur)
        con.commit()
        con.close()


def _init_feedback_tables(cur) -> None:
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS feedback_threads (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          user_id INTEGER NOT NULL,
          subject TEXT NOT NULL DEFAULT '',
          created_at TEXT NOT NULL DEFAULT '',
          updated_at TEXT NOT NULL DEFAULT ''
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS feedback_messages (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          thread_id INTEGER NOT NULL,
          author_user_id INTEGER NOT NULL DEFAULT 0,
          author_name TEXT NOT NULL DEFAULT '',
          body TEXT NOT NULL DEFAULT '',
          created_at TEXT NOT NULL DEFAULT '',
          FOREIGN KEY (thread_id) REFERENCES feedback_threads(id) ON DELETE CASCADE
        )
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_feedback_threads_user
        ON feedback_threads(user_id)
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_feedback_messages_thread
        ON feedback_messages(thread_id)
        """
    )
    feedback_cols = {
        r[1] for r in cur.execute("PRAGMA table_info(feedback_threads)").fetchall()
    }
    if "is_closed" not in feedback_cols:
        cur.execute(
            "ALTER TABLE feedback_threads ADD COLUMN is_closed INTEGER NOT NULL DEFAULT 0"
        )
    if "user_last_read_at" not in feedback_cols:
        cur.execute(
            "ALTER TABLE feedback_threads ADD COLUMN user_last_read_at TEXT NOT NULL DEFAULT ''"
        )
    if "admin_last_read_at" not in feedback_cols:
        cur.execute(
            "ALTER TABLE feedback_threads ADD COLUMN admin_last_read_at TEXT NOT NULL DEFAULT ''"
        )
    if "task_status" not in feedback_cols:
        cur.execute(
            "ALTER TABLE feedback_threads ADD COLUMN task_status TEXT NOT NULL DEFAULT 'in_progress'"
        )
        cur.execute(
            """
            UPDATE feedback_threads
            SET task_status = 'resolved'
            WHERE is_closed = 1
            """
        )


def _init_order_chat_tables(cur) -> None:
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS order_messages (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          order_id INTEGER NOT NULL,
          author_user_id INTEGER NOT NULL DEFAULT 0,
          author_name TEXT NOT NULL DEFAULT '',
          body TEXT NOT NULL DEFAULT '',
          created_at TEXT NOT NULL DEFAULT '',
          FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS order_chat_reads (
          order_id INTEGER NOT NULL,
          user_id INTEGER NOT NULL DEFAULT 0,
          last_read_at TEXT NOT NULL DEFAULT '',
          PRIMARY KEY (order_id, user_id),
          FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE
        )
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_order_messages_order
        ON order_messages(order_id)
        """
    )


def _init_unique_clients_table(cur) -> None:
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS unique_clients (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          display_name TEXT NOT NULL,
          client_key TEXT NOT NULL UNIQUE,
          delivery_type TEXT NOT NULL,
          pallet_kind TEXT NOT NULL DEFAULT 'generic',
          sort_order INTEGER NOT NULL DEFAULT 0,
          active INTEGER NOT NULL DEFAULT 1
        )
        """
    )
    seeds = (
        ("Арнест Юнирусь", "арнест юнирусь", "Арнест Юнирусь", "arnest", 10),
        ("ЛАБ Индастриз", "лаб индастриз", "ЛАБ Индастриз", "lab", 20),
        ("Дрогери Ритейл", "дрогери ритейл", "Дрогери Ритейл", "generic", 30),
    )
    for display_name, client_key, delivery_type, pallet_kind, sort_order in seeds:
        cur.execute(
            """
            INSERT OR IGNORE INTO unique_clients (
              display_name, client_key, delivery_type, pallet_kind, sort_order
            ) VALUES (?, ?, ?, ?, ?)
            """,
            (display_name, client_key, delivery_type, pallet_kind, sort_order),
        )
    _reload_unique_clients_cache(cur)


def _init_orders_table(cur):
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS orders (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          ship_date TEXT NOT NULL DEFAULT '',
          client TEXT NOT NULL DEFAULT '',
          assembled_percent INTEGER NOT NULL DEFAULT 0,
          names TEXT NOT NULL DEFAULT '',
          extra_info TEXT NOT NULL DEFAULT ''
        )
        """
    )
    n = cur.execute("SELECT COUNT(*) FROM orders").fetchone()[0]
    if n == 0:
        samples = [
            (
                "15.05.2026",
                "ООО «Розница Юг»",
                72,
                "Салфетки САЛДЕЗ; Aqua Joy гель Mango 1 л",
                "Паллеты в зоне Б, проверить маркировку",
            ),
            (
                "18.05.2026",
                "ИП Ким А.С.",
                35,
                "Спрей Шаума Kids; Bambolina масло 250 мл",
                "Частичная отгрузка, звонок за 2 ч",
            ),
            (
                "22.05.2026",
                "Сеть «Косметика Плюс»",
                0,
                "Adaly двухфазное средство 120 мл (×24 короба)",
                "Новый заказ, сбор не начат",
            ),
        ]
        cur.executemany(
            """
            INSERT INTO orders (
              ship_date, client, assembled_percent, names, extra_info
            ) VALUES (?, ?, ?, ?, ?)
            """,
            samples,
        )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS order_items (
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          order_id INTEGER NOT NULL,
          product_id INTEGER,
          article TEXT NOT NULL DEFAULT '',
          name TEXT NOT NULL DEFAULT '',
          quantity REAL NOT NULL DEFAULT 0,
          unit TEXT NOT NULL DEFAULT 'piece',
          FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE
        )
        """
    )
    order_cols = {r[1] for r in cur.execute("PRAGMA table_info(orders)").fetchall()}
    if "assemble_state" not in order_cols:
        cur.execute(
            "ALTER TABLE orders ADD COLUMN assemble_state TEXT NOT NULL DEFAULT ''"
        )
    if "assemble_revision" not in order_cols:
        cur.execute(
            "ALTER TABLE orders ADD COLUMN assemble_revision INTEGER NOT NULL DEFAULT 0"
        )
    if "assemble_state_updated_at" not in order_cols:
        cur.execute(
            "ALTER TABLE orders ADD COLUMN assemble_state_updated_at TEXT NOT NULL DEFAULT ''"
        )
    if "buyer_order_mode" not in order_cols:
        cur.execute(
            "ALTER TABLE orders ADD COLUMN buyer_order_mode TEXT NOT NULL DEFAULT ''"
        )
    if "buyer_order" not in order_cols:
        cur.execute(
            "ALTER TABLE orders ADD COLUMN buyer_order TEXT NOT NULL DEFAULT ''"
        )
    if "total_order_quantity" not in order_cols:
        cur.execute(
            "ALTER TABLE orders ADD COLUMN total_order_quantity REAL"
        )
    item_cols = {r[1] for r in cur.execute("PRAGMA table_info(order_items)").fetchall()}
    if "buyer_order" not in item_cols:
        cur.execute(
            "ALTER TABLE order_items ADD COLUMN buyer_order TEXT NOT NULL DEFAULT ''"
        )
    if "total_order_quantity" not in item_cols:
        cur.execute(
            "ALTER TABLE order_items ADD COLUMN total_order_quantity REAL"
        )
    if "lab_sscc_ai_seq" not in order_cols:
        cur.execute("ALTER TABLE orders ADD COLUMN lab_sscc_ai_seq INTEGER")
    if "lab_sscc_seq_start" not in order_cols:
        cur.execute("ALTER TABLE orders ADD COLUMN lab_sscc_seq_start INTEGER")
    if "lab_sscc_shipped" not in order_cols:
        cur.execute(
            "ALTER TABLE orders ADD COLUMN lab_sscc_shipped INTEGER NOT NULL DEFAULT 0"
        )
    if "lab_sscc_pallet_count" not in order_cols:
        cur.execute("ALTER TABLE orders ADD COLUMN lab_sscc_pallet_count INTEGER")
    if "last_modified_by" not in order_cols:
        cur.execute(
            "ALTER TABLE orders ADD COLUMN last_modified_by TEXT NOT NULL DEFAULT ''"
        )
    order_cols = {r[1] for r in cur.execute("PRAGMA table_info(orders)").fetchall()}
    if "client_city" not in order_cols:
        cur.execute(
            "ALTER TABLE orders ADD COLUMN client_city TEXT NOT NULL DEFAULT ''"
        )
    if "last_edited_by" not in order_cols:
        cur.execute(
            "ALTER TABLE orders ADD COLUMN last_edited_by TEXT NOT NULL DEFAULT ''"
        )
        if "last_modified_by" in order_cols:
            cur.execute(
                """
                UPDATE orders
                SET last_edited_by = last_modified_by
                WHERE COALESCE(last_edited_by, '') = ''
                  AND COALESCE(last_modified_by, '') != ''
                """
            )
    if "last_assembled_by" not in order_cols:
        cur.execute(
            "ALTER TABLE orders ADD COLUMN last_assembled_by TEXT NOT NULL DEFAULT ''"
        )
    order_cols = {r[1] for r in cur.execute("PRAGMA table_info(orders)").fetchall()}
    if "order_readiness" not in order_cols:
        cur.execute(
            "ALTER TABLE orders ADD COLUMN order_readiness TEXT NOT NULL DEFAULT 'assembling'"
        )
        cur.execute(
            """
            UPDATE orders
            SET order_readiness = 'assembled'
            WHERE assembled_percent >= 100
              AND COALESCE(order_readiness, 'assembling') = 'assembling'
            """
        )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS app_settings (
          key TEXT PRIMARY KEY,
          value TEXT NOT NULL DEFAULT ''
        )
        """
    )
    if not cur.execute(
        "SELECT 1 FROM app_settings WHERE key = 'lab_sscc_last_shipped' LIMIT 1"
    ).fetchone():
        cur.execute(
            "INSERT INTO app_settings (key, value) VALUES ('lab_sscc_last_shipped', '11')"
        )


_LAB_SSCC_LAST_SHIPPED_KEY = "lab_sscc_last_shipped"
_LAB_CLIENT_SQL = "LOWER(TRIM(REPLACE(client, '  ', ' '))) = 'лаб индастриз'"


def _lab_sscc_get_last_shipped(cur) -> int:
    row = cur.execute(
        "SELECT value FROM app_settings WHERE key = ? LIMIT 1",
        (_LAB_SSCC_LAST_SHIPPED_KEY,),
    ).fetchone()
    if not row:
        return 0
    try:
        return max(0, int(str(row[0]).strip()))
    except ValueError:
        return 0


def _lab_sscc_set_last_shipped(cur, value: int) -> None:
    stored = max(0, int(value))
    cur.execute(
        """
        INSERT INTO app_settings (key, value) VALUES (?, ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value
        """,
        (_LAB_SSCC_LAST_SHIPPED_KEY, str(stored)),
    )


def reset_lab_sscc_pallet_counter(
    *,
    last_shipped: int | None = None,
    next_pallet: int | None = None,
) -> dict:
    """Сброс глобального счётчика SSCC и пересчёт неотгруженных заказов ЛАБ."""
    from packing_sheets_lab_industries import lab_sscc_next_after

    if next_pallet is not None:
        last = max(0, int(next_pallet) - 1)
    elif last_shipped is not None:
        last = max(0, int(last_shipped))
    else:
        last = 11
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        _lab_sscc_set_last_shipped(cur, last)
        for row in _lab_sscc_unshipped_lab_order_rows(cur):
            cur.execute(
                """
                UPDATE orders
                SET lab_sscc_seq_start = NULL, lab_sscc_pallet_count = NULL
                WHERE id = ?
                """,
                (int(row["id"]),),
            )
        _lab_sscc_sync_all_unshipped_seq_starts(cur)
        con.commit()
        con.close()
    return {
        "ok": True,
        "last_shipped": last,
        "next_pallet_sscc": lab_sscc_next_after(last),
    }


def get_lab_sscc_last_shipped() -> int:
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        last = _lab_sscc_get_last_shipped(cur)
        con.close()
        return last


def _count_assemble_pallets(assemble_state) -> int:
    if not isinstance(assemble_state, dict):
        return 0
    pallets = assemble_state.get("pallets")
    if not isinstance(pallets, list):
        return 0
    return sum(1 for p in pallets if isinstance(p, dict))


def _lab_sscc_pallet_count_for_order_row(row) -> int:
    """Число паллет — только из актуальной сборки (не из устаревшего lab_sscc_pallet_count)."""
    return _count_assemble_pallets(_assemble_state_cell_to_api(row["assemble_state"]))


def _normalize_client_key(client: str) -> str:
    return re.sub(r"\s+", " ", str(client or "").strip()).lower()


_UNIQUE_CLIENTS_BY_KEY: dict[str, dict] = {}


def _reload_unique_clients_cache(cur) -> None:
    global _UNIQUE_CLIENTS_BY_KEY
    rows = cur.execute(
        """
        SELECT display_name, client_key, delivery_type, pallet_kind, sort_order
        FROM unique_clients
        WHERE active = 1
        ORDER BY sort_order, id
        """
    ).fetchall()
    _UNIQUE_CLIENTS_BY_KEY = {
        str(r["client_key"] or ""): {
            "display_name": str(r["display_name"] or ""),
            "delivery_type": str(r["delivery_type"] or ""),
            "pallet_kind": str(r["pallet_kind"] or "generic").strip().lower(),
            "sort_order": int(r["sort_order"] or 0),
        }
        for r in rows
    }


def _unique_client_record(client: str) -> dict | None:
    return _UNIQUE_CLIENTS_BY_KEY.get(_normalize_client_key(client))


def fetch_unique_clients() -> dict:
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        if not _UNIQUE_CLIENTS_BY_KEY:
            _reload_unique_clients_cache(cur)
        con.close()
    clients = [
        {
            "display_name": rec["display_name"],
            "client_key": key,
            "delivery_type": rec["delivery_type"],
            "pallet_kind": rec["pallet_kind"],
        }
        for key, rec in sorted(
            _UNIQUE_CLIENTS_BY_KEY.items(),
            key=lambda item: (item[1]["sort_order"], item[0]),
        )
    ]
    return {"clients": clients}


def _lab_sscc_unshipped_lab_order_rows(cur):
    rows = cur.execute(
        """
        SELECT id, client, assemble_state, lab_sscc_pallet_count
        FROM orders
        WHERE COALESCE(lab_sscc_shipped, 0) = 0
        ORDER BY id ASC
        """
    ).fetchall()
    return [r for r in rows if _is_lab_industries_client(str(r["client"] or ""))]


def _lab_sscc_pallet_count_for_order_id(
    cur, order_id: int, pallet_overrides: dict[int, int] | None = None
) -> int:
    if pallet_overrides and int(order_id) in pallet_overrides:
        return max(0, int(pallet_overrides[int(order_id)]))
    row = cur.execute(
        "SELECT assemble_state, lab_sscc_pallet_count FROM orders WHERE id = ?",
        (int(order_id),),
    ).fetchone()
    if not row:
        return 0
    return _lab_sscc_pallet_count_for_order_row(row)


def _lab_sscc_count_physical_pallets_from_pdf_rows(pallets_raw: list) -> int:
    seen: set[int | str] = set()
    for i, p in enumerate(pallets_raw):
        if not isinstance(p, dict):
            continue
        pi = p.get("lab_sscc_pallet_index")
        if pi is not None and str(pi).strip() != "":
            try:
                seen.add(int(pi))
                continue
            except (TypeError, ValueError):
                pass
        pn = str(p.get("pallet_number") or "").strip()
        seen.add(pn if pn else f"__row_{i}")
    return len(seen)


def _lab_sscc_compute_seq_start(cur, order_id: int) -> int:
    """Стартовый SSCC для одного заказа (после полного sync)."""
    row = cur.execute(
        "SELECT lab_sscc_seq_start FROM orders WHERE id = ?",
        (int(order_id),),
    ).fetchone()
    if row and row[0] is not None:
        return max(1, int(row[0]))
    _lab_sscc_sync_all_unshipped_seq_starts(cur)
    row = cur.execute(
        "SELECT lab_sscc_seq_start FROM orders WHERE id = ?",
        (int(order_id),),
    ).fetchone()
    if row and row[0] is not None:
        return max(1, int(row[0]))
    from packing_sheets_lab_industries import lab_sscc_next_after

    return lab_sscc_next_after(_lab_sscc_get_last_shipped(cur))


def _lab_sscc_persist_seq_start(cur, order_id: int, seq_start: int) -> None:
    row = cur.execute(
        "SELECT assemble_state, lab_sscc_pallet_count FROM orders WHERE id = ?",
        (int(order_id),),
    ).fetchone()
    pallet_n = _lab_sscc_pallet_count_for_order_row(row) if row else 0
    if pallet_n > 0:
        cur.execute(
            """
            UPDATE orders SET lab_sscc_seq_start = ?, lab_sscc_pallet_count = ?
            WHERE id = ?
            """,
            (int(seq_start), pallet_n, int(order_id)),
        )
    else:
        cur.execute(
            "UPDATE orders SET lab_sscc_seq_start = ? WHERE id = ?",
            (int(seq_start), int(order_id)),
        )


def _lab_sscc_sync_all_unshipped_seq_starts(
    cur, pallet_overrides: dict[int, int] | None = None
) -> None:
    """Распределить SSCC по цепочке неотгруженных заказов ЛАБ (id ASC), без дублей seq_start."""
    from packing_sheets_lab_industries import lab_sscc_last_for_order, lab_sscc_next_after

    overrides = {int(k): max(0, int(v)) for k, v in (pallet_overrides or {}).items()}
    next_start = lab_sscc_next_after(_lab_sscc_get_last_shipped(cur))
    for row in _lab_sscc_unshipped_lab_order_rows(cur):
        oid = int(row["id"])
        pallet_n = _lab_sscc_pallet_count_for_order_id(cur, oid, overrides)
        if pallet_n <= 0:
            cur.execute(
                "UPDATE orders SET lab_sscc_seq_start = NULL WHERE id = ?",
                (oid,),
            )
            continue
        seq_start = next_start
        _lab_sscc_persist_seq_start(cur, oid, seq_start)
        next_start = lab_sscc_next_after(lab_sscc_last_for_order(seq_start, pallet_n))


def _lab_sscc_unshipped_seq_start_map(
    cur, pallet_overrides: dict[int, int] | None = None
) -> dict[int, int]:
    _lab_sscc_sync_all_unshipped_seq_starts(cur, pallet_overrides)
    out: dict[int, int] = {}
    for row in _lab_sscc_unshipped_lab_order_rows(cur):
        oid = int(row["id"])
        r = cur.execute(
            "SELECT lab_sscc_seq_start FROM orders WHERE id = ?",
            (oid,),
        ).fetchone()
        if r and r[0] is not None:
            out[oid] = max(1, int(r[0]))
    return out


def get_or_assign_lab_sscc_seq_start(order_id: int) -> int:
    """Актуальный старт SSCC для заказа ЛАБ (пересчёт цепочки неотгруженных)."""
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        row = cur.execute(
            """
            SELECT lab_sscc_seq_start, client, COALESCE(lab_sscc_shipped, 0) AS lab_sscc_shipped
            FROM orders WHERE id = ?
            """,
            (int(order_id),),
        ).fetchone()
        if not row:
            con.close()
            return 1
        if not _is_lab_industries_client(str(row["client"] or "")):
            con.close()
            return 1
        if int(row["lab_sscc_shipped"] or 0):
            seq = max(1, int(row["lab_sscc_seq_start"] or 1))
            con.close()
            return seq
        seq_map = _lab_sscc_unshipped_seq_start_map(cur)
        seq = seq_map.get(int(order_id))
        if seq is None:
            seq = _lab_sscc_compute_seq_start(cur, int(order_id))
        con.commit()
        con.close()
        return seq


def confirm_lab_order_shipment(order_id: int) -> dict:
    """Отгрузка заказа ЛАБ: сохранить последний SSCC-номер паллеты."""
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        row = cur.execute(
            """
            SELECT client, assemble_state, lab_sscc_seq_start,
                   COALESCE(lab_sscc_shipped, 0) AS lab_sscc_shipped
            FROM orders WHERE id = ?
            """,
            (int(order_id),),
        ).fetchone()
        if not row:
            con.close()
            return {"error": "not_found", "message": "Заказ не найден."}
        if not _is_lab_industries_client(str(row["client"] or "")):
            con.close()
            return {
                "error": "lab_client",
                "message": "Отгрузка SSCC только для клиента «ЛАБ Индастриз».",
            }
        if int(row["lab_sscc_shipped"] or 0):
            con.close()
            return {
                "error": "already_shipped",
                "message": "Заказ уже отмечен как отгруженный.",
            }
        st = _assemble_state_cell_to_api(row["assemble_state"])
        pallet_count = _count_assemble_pallets(st)
        if pallet_count <= 0:
            con.close()
            return {
                "error": "no_pallets",
                "message": "В сборке нет паллет — укажите паллеты перед отгрузкой.",
            }
        override = {int(order_id): pallet_count}
        _lab_sscc_sync_all_unshipped_seq_starts(cur, override)
        seq_row = cur.execute(
            "SELECT lab_sscc_seq_start FROM orders WHERE id = ?",
            (int(order_id),),
        ).fetchone()
        if seq_row and seq_row[0] is not None:
            seq_start = max(1, int(seq_row[0]))
        else:
            seq_start = _lab_sscc_compute_seq_start(cur, int(order_id))
        from packing_sheets_lab_industries import lab_sscc_last_for_order

        last_used = lab_sscc_last_for_order(seq_start, pallet_count)
        _lab_sscc_set_last_shipped(cur, last_used)
        cur.execute(
            """
            UPDATE orders
            SET lab_sscc_shipped = 1,
                lab_sscc_pallet_count = ?,
                lab_sscc_seq_start = ?
            WHERE id = ?
            """,
            (pallet_count, seq_start, int(order_id)),
        )
        _lab_sscc_sync_all_unshipped_seq_starts(cur)
        con.commit()
        con.close()
    return {
        "ok": True,
        "last_shipped": last_used,
        "pallet_count": pallet_count,
        "seq_start": seq_start,
    }


_MAX_ASSEMBLE_STATE_JSON_BYTES = 2 * 1024 * 1024


def _assemble_state_cell_to_api(value) -> dict | None:
    """JSON из БД → объект для API или None."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        obj = json.loads(s)
    except json.JSONDecodeError:
        return None
    if not isinstance(obj, dict) or not isinstance(obj.get("pallets"), list):
        return None
    return obj


def _normalize_assemble_slot_for_compare(slot) -> dict:
    if not isinstance(slot, dict):
        return {}
    li = slot.get("lineIndex")
    line_index: int | str = ""
    if li not in ("", None):
        try:
            n = int(li)
            if n >= 0:
                line_index = n
        except (TypeError, ValueError):
            line_index = ""
    mode = "rows" if slot.get("mode") == "rows" else "direct"
    direct_unit = "piece" if slot.get("directUnit") == "piece" else "box"
    return {
        "lineIndex": line_index,
        "batchNumber": _normalize_str(str(slot.get("batchNumber") or "")),
        "unirusMfgDate": _normalize_str(str(slot.get("unirusMfgDate") or "")),
        "unirusExpiryDate": _normalize_str(str(slot.get("unirusExpiryDate") or "")),
        "labExpiryDate": _normalize_str(str(slot.get("labExpiryDate") or "")),
        "mode": mode,
        "directUnit": direct_unit,
        "directQty": slot.get("directQty"),
        "fullRows": slot.get("fullRows"),
        "partialBoxes": slot.get("partialBoxes"),
    }


def _normalize_assemble_state_for_compare(value) -> dict:
    obj = value
    if isinstance(value, str):
        obj = _assemble_state_cell_to_api(value)
    if not isinstance(obj, dict):
        return {"optional_batch_enabled": False, "pallets": []}
    pallets_out = []
    for pal in obj.get("pallets") or []:
        if not isinstance(pal, dict):
            continue
        slots = [
            _normalize_assemble_slot_for_compare(s)
            for s in (pal.get("slots") or [])
            if isinstance(s, dict)
        ]
        slots.sort(
            key=lambda s: (
                str(s.get("lineIndex", "")),
                str(s.get("batchNumber", "")),
            )
        )
        pallets_out.append(
            {
                "id": pal.get("id"),
                "palletNumber": _normalize_str(
                    str(pal.get("palletNumber") or pal.get("pallet_number") or "")
                ),
                "slots": slots,
            }
        )
    pallets_out.sort(
        key=lambda p: (
            p.get("id") if p.get("id") is not None else 0,
            p.get("palletNumber") or "",
        )
    )
    return {
        "optional_batch_enabled": bool(obj.get("optional_batch_enabled")),
        "pallets": pallets_out,
    }


def _assemble_states_equal(left, right) -> bool:
    return _normalize_assemble_state_for_compare(left) == _normalize_assemble_state_for_compare(
        right
    )


def _order_item_save_signature(item_tuple) -> tuple:
    pid, article, name, qty, unit, line_buyer, line_total_qty = item_tuple
    qty_f = float(qty) if qty is not None else 0.0
    total_qty = (
        float(line_total_qty) if line_total_qty is not None else None
    )
    return (
        pid,
        article or "",
        name or "",
        qty_f,
        unit or "",
        line_buyer or "",
        total_qty,
    )


def _fetch_order_item_signatures(cur, order_id: int) -> list[tuple]:
    rows = cur.execute(
        """
        SELECT product_id, article, name, quantity, unit, buyer_order,
               total_order_quantity
        FROM order_items
        WHERE order_id = ?
        ORDER BY id
        """,
        (order_id,),
    ).fetchall()
    return [
        _order_item_save_signature(
            (
                row["product_id"],
                row["article"],
                row["name"],
                row["quantity"],
                row["unit"],
                row["buyer_order"],
                row["total_order_quantity"],
            )
        )
        for row in rows
    ]


def _build_order_line_index_remap(
    old_sigs: list[tuple], new_sigs: list[tuple]
) -> dict[int, int | None]:
    """Сопоставление старого индекса строки заказа с новым (None — строка удалена)."""
    remap: dict[int, int | None] = {}
    used_new: set[int] = set()

    for i, osig in enumerate(old_sigs):
        if i < len(new_sigs) and new_sigs[i] == osig and i not in used_new:
            remap[i] = i
            used_new.add(i)

    for i, osig in enumerate(old_sigs):
        if i in remap:
            continue
        matched = False
        for j, nsig in enumerate(new_sigs):
            if j in used_new:
                continue
            if osig == nsig:
                remap[i] = j
                used_new.add(j)
                matched = True
                break
        if not matched:
            remap[i] = None

    return remap


def _remap_assemble_state_line_indices(
    assemble_state_raw: str, index_remap: dict[int, int | None]
) -> tuple[str, bool]:
    """Пересчитать lineIndex в assemble_state после изменения позиций заказа."""
    if not assemble_state_raw or not index_remap:
        return assemble_state_raw, False
    obj = _assemble_state_cell_to_api(assemble_state_raw)
    if not obj:
        return assemble_state_raw, False

    changed = False
    for pal in obj.get("pallets") or []:
        if not isinstance(pal, dict):
            continue
        old_slots = pal.get("slots") or []
        new_slots = []
        for slot in old_slots:
            if not isinstance(slot, dict):
                continue
            li = slot.get("lineIndex")
            if li in ("", None):
                new_slots.append(slot)
                continue
            try:
                old_idx = int(li)
            except (TypeError, ValueError):
                new_slots.append(slot)
                continue
            if old_idx not in index_remap:
                new_slots.append(slot)
                continue
            new_idx = index_remap[old_idx]
            if new_idx is None:
                changed = True
                continue
            if new_idx != old_idx:
                slot = dict(slot)
                slot["lineIndex"] = new_idx
                changed = True
            new_slots.append(slot)
        if len(new_slots) != len(old_slots):
            changed = True
        pal["slots"] = new_slots

    if not changed:
        return assemble_state_raw, False
    return json.dumps(obj, ensure_ascii=False, separators=(",", ":")), True


def _assemble_state_has_meaningful_pallets(pallets) -> bool:
    """Пустой шаблон (один паллет без слотов) не считается сохранённой сборкой."""
    if not isinstance(pallets, list) or not pallets:
        return False
    for pal in pallets:
        if not isinstance(pal, dict):
            continue
        if str(pal.get("palletNumber") or pal.get("pallet_number") or "").strip():
            return True
        slots = pal.get("slots")
        if not isinstance(slots, list):
            continue
        for slot in slots:
            if not isinstance(slot, dict):
                continue
            li = slot.get("lineIndex")
            if li not in ("", None):
                try:
                    if int(li) >= 0:
                        return True
                except (TypeError, ValueError):
                    pass
            for key in (
                "batchNumber",
                "unirusMfgDate",
                "unirusExpiryDate",
                "labExpiryDate",
            ):
                if str(slot.get(key) or "").strip():
                    return True
    return len(pallets) > 1


_ORDER_READINESS_ASSEMBLING = "assembling"
_ORDER_READINESS_ASSEMBLED = "assembled"
_ORDER_READINESS_SHIPPED = "shipped"
_VALID_ORDER_READINESS = frozenset(
    {
        _ORDER_READINESS_ASSEMBLING,
        _ORDER_READINESS_ASSEMBLED,
        _ORDER_READINESS_SHIPPED,
    }
)


def _normalize_order_readiness(
    value, *, default: str = _ORDER_READINESS_ASSEMBLING
) -> str:
    v = (value or "").strip().lower()
    if v in _VALID_ORDER_READINESS:
        return v
    return default


def _order_readiness_from_row(row) -> str:
    if row is not None and "order_readiness" in row.keys():
        return _normalize_order_readiness(row["order_readiness"])
    pct = max(0, min(100, int(row["assembled_percent"] or 0))) if row else 0
    return (
        _ORDER_READINESS_ASSEMBLED
        if pct >= 100
        else _ORDER_READINESS_ASSEMBLING
    )


def _order_is_shipment_locked(row) -> bool:
    return _order_readiness_from_row(row) == _ORDER_READINESS_SHIPPED


def _order_shipment_locked_error() -> dict:
    return {
        "error": "order_shipped",
        "message": "Заказ отгружен — изменения сборки и редактирование заказа недоступны.",
    }


def _readiness_after_assembly_patch(
    current_readiness: str,
    *,
    explicit_readiness: str | None,
    old_pct: int,
    apct: int,
    content_changed: bool,
) -> str:
    if explicit_readiness is not None:
        return explicit_readiness
    if content_changed and current_readiness == _ORDER_READINESS_ASSEMBLED:
        return _ORDER_READINESS_ASSEMBLING
    if (
        content_changed
        and apct >= 100
        and old_pct < 100
        and current_readiness == _ORDER_READINESS_ASSEMBLING
    ):
        return _ORDER_READINESS_ASSEMBLED
    return current_readiness


def patch_order_assembly(
    order_id: int, body: dict, *, modified_by: str = ""
) -> dict:
    """Частичное обновление: assembled_percent, assemble_state и/или order_readiness."""
    if not isinstance(body, dict):
        return {"error": "validation", "message": "Ожидался JSON-объект."}
    has_pct = "assembled_percent" in body
    has_state = "assemble_state" in body
    has_readiness = "order_readiness" in body
    if not has_pct and not has_state and not has_readiness:
        return {
            "error": "validation",
            "message": "Передайте assembled_percent, assemble_state и/или order_readiness.",
        }
    new_readiness = None
    if has_readiness:
        raw_readiness = (body.get("order_readiness") or "").strip().lower()
        if raw_readiness not in _VALID_ORDER_READINESS:
            return {
                "error": "validation",
                "message": "order_readiness: допустимы assembling, assembled, shipped.",
            }
        new_readiness = raw_readiness
    expected_rev = None
    if "expected_assemble_revision" in body:
        try:
            expected_rev = int(body["expected_assemble_revision"])
        except (TypeError, ValueError):
            return {
                "error": "validation",
                "message": "expected_assemble_revision: нужно целое число.",
            }
    new_pct = None
    if has_pct:
        try:
            new_pct = max(0, min(100, int(body["assembled_percent"])))
        except (TypeError, ValueError):
            return {"error": "validation", "message": "assembled_percent: нужно целое 0…100."}
    new_state_json = None
    new_state_pallets = None
    if has_state:
        st = body["assemble_state"]
        if st is None:
            new_state_json = ""
            new_state_pallets = []
        elif not isinstance(st, dict):
            return {"error": "validation", "message": "assemble_state: ожидался объект или null."}
        else:
            try:
                new_state_json = json.dumps(st, ensure_ascii=False)
            except (TypeError, ValueError):
                return {"error": "validation", "message": "assemble_state: не удалось сериализовать в JSON."}
            if len(new_state_json.encode("utf-8")) > _MAX_ASSEMBLE_STATE_JSON_BYTES:
                return {
                    "error": "validation",
                    "message": "assemble_state слишком большой.",
                }
            pallets = st.get("pallets")
            if not isinstance(pallets, list):
                return {
                    "error": "validation",
                    "message": "assemble_state.pallets: ожидался массив.",
                }
            new_state_pallets = pallets
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        row = cur.execute(
            """
            SELECT id, assembled_percent, assemble_state, assemble_revision,
                   assemble_state_updated_at, client, order_readiness,
                   COALESCE(lab_sscc_shipped, 0) AS lab_sscc_shipped
            FROM orders WHERE id = ?
            """,
            (order_id,),
        ).fetchone()
        if not row:
            con.close()
            return {"error": "not_found", "message": "Заказ не найден."}
        current_readiness = _order_readiness_from_row(row)
        if (has_pct or has_state) and current_readiness == _ORDER_READINESS_SHIPPED:
            con.close()
            return _order_shipment_locked_error()
        if has_readiness and not has_pct and not has_state:
            if new_readiness == current_readiness:
                con.close()
                return {
                    "ok": True,
                    "id": int(order_id),
                    "order_readiness": current_readiness,
                    "unchanged": True,
                }
            cur.execute(
                "UPDATE orders SET order_readiness = ? WHERE id = ?",
                (new_readiness, order_id),
            )
            con.commit()
            con.close()
            return {
                "ok": True,
                "id": int(order_id),
                "order_readiness": new_readiness,
            }
        current_rev = max(0, int(row["assemble_revision"] or 0))
        if expected_rev is not None and expected_rev != current_rev:
            con.close()
            return {
                "error": "conflict",
                "message": "Сборка была изменена на другом устройстве. Обновите страницу или загрузите данные с сервера.",
                "id": int(order_id),
                "assemble_revision": current_rev,
                "assembled_percent": max(
                    0, min(100, int(row["assembled_percent"] or 0))
                ),
                "assemble_state": _assemble_state_cell_to_api(row["assemble_state"]),
                "assemble_state_updated_at": (row["assemble_state_updated_at"] or "").strip(),
            }
        old_pct = max(0, min(100, int(row["assembled_percent"] or 0)))
        old_ajson = row["assemble_state"] or ""
        apct = new_pct if new_pct is not None else old_pct
        if new_state_json is not None:
            ajson = new_state_json
        else:
            ajson = old_ajson
        content_changed = not (
            apct == old_pct and _assemble_states_equal(old_ajson, ajson)
        )
        effective_readiness = _readiness_after_assembly_patch(
            current_readiness,
            explicit_readiness=new_readiness,
            old_pct=old_pct,
            apct=apct,
            content_changed=content_changed,
        )
        if (
            not content_changed
            and effective_readiness == current_readiness
        ):
            con.close()
            return {
                "ok": True,
                "id": int(order_id),
                "assemble_revision": current_rev,
                "assemble_state_updated_at": (row["assemble_state_updated_at"] or "").strip(),
                "assembled_percent": old_pct,
                "order_readiness": current_readiness,
                "unchanged": True,
            }
        new_rev = current_rev + 1
        updated_at = datetime.datetime.now(datetime.timezone.utc).strftime(
            "%Y-%m-%dT%H:%M:%SZ"
        )
        modifier = _normalize_str(modified_by)
        cur.execute(
            """
            UPDATE orders
            SET assembled_percent = ?, assemble_state = ?,
                assemble_revision = ?, assemble_state_updated_at = ?,
                last_assembled_by = CASE WHEN ? != '' THEN ? ELSE last_assembled_by END,
                order_readiness = ?
            WHERE id = ?
            """,
            (apct, ajson, new_rev, updated_at, modifier, modifier, effective_readiness, order_id),
        )
        lab_sscc_seq_by_order: dict[int, int] = {}
        if has_state and _is_lab_industries_client(
            str(row["client"] or "")
        ) and not int(row["lab_sscc_shipped"] or 0):
            override: dict[int, int] = {}
            if new_state_pallets is not None:
                override[int(order_id)] = len(
                    [p for p in new_state_pallets if isinstance(p, dict)]
                )
            lab_sscc_seq_by_order = _lab_sscc_unshipped_seq_start_map(cur, override)
        con.commit()
        con.close()
    out = {
        "ok": True,
        "id": int(order_id),
        "assemble_revision": new_rev,
        "assemble_state_updated_at": updated_at,
        "assembled_percent": apct,
        "order_readiness": effective_readiness,
    }
    if lab_sscc_seq_by_order:
        out["lab_sscc_seq_by_order"] = {
            str(oid): seq for oid, seq in lab_sscc_seq_by_order.items()
        }
        out["lab_sscc_seq_start"] = lab_sscc_seq_by_order.get(int(order_id))
    return out


def _format_ship_date_storage(value: str) -> str:
    v = (value or "").strip()
    if len(v) == 10 and v[4] == "-" and v[7] == "-":
        return f"{v[8:10]}.{v[5:7]}.{v[0:4]}"
    return v


def _format_order_names_summary(items):
    unit_ru = {"box": "коробок", "set": "наборов", "piece": "шт"}
    parts = []
    for row in items:
        _pid, article, name, qty, unit = row[0], row[1], row[2], row[3], row[4]
        label = (name or article or "—").strip() or "—"
        u = unit_ru.get(unit, "шт")
        q = int(qty) if qty == int(qty) else qty
        parts.append(f"{label} × {q} {u}")
    return "; ".join(parts)


_ORDER_NAMES_LINE_RE = re.compile(
    r"^(.+?)\s*×\s*([\d.,]+)\s*(коробок|наборов|шт)\s*$",
    re.UNICODE,
)


def _synthetic_items_from_order_names(names: str):
    """Если в order_items нет строк, собираем позиции из сводки names (демо/старые заказы)."""
    if not names or not str(names).strip():
        return []
    out = []
    for part in str(names).split(";"):
        seg = part.strip()
        if not seg:
            continue
        m = _ORDER_NAMES_LINE_RE.match(seg)
        if m:
            label = m.group(1).strip()
            try:
                qty = float(m.group(2).replace(",", "."))
            except (TypeError, ValueError):
                qty = 1.0
            unit = {"коробок": "box", "наборов": "set", "шт": "piece"}.get(m.group(3), "piece")
            out.append(
                {
                    "product_id": None,
                    "article": "",
                    "name": label,
                    "quantity": qty,
                    "unit": unit,
                    "pieces_in_box": 0,
                    "sets_in_box": 1,
                    "pieces_per_set": 1,
                    "row_layout": 0,
                    "max_rows": 0,
                    "box_weight": 0.0,
                }
            )
        else:
            out.append(
                {
                    "product_id": None,
                    "article": "",
                    "name": seg,
                    "quantity": 1.0,
                    "unit": "piece",
                    "pieces_in_box": 0,
                    "sets_in_box": 1,
                    "pieces_per_set": 1,
                    "row_layout": 0,
                    "max_rows": 0,
                    "box_weight": 0.0,
                }
            )
    return out


def _build_product_name_import_lookup(cur) -> dict[str, int]:
    """Нормализованное наименование → id (первое при дубликатах ключей)."""
    lookup: dict[str, int] = {}
    for row in cur.execute("SELECT id, name FROM products"):
        key = _normalize_product_name_for_import_match(row["name"] or "")
        if key and key not in lookup:
            lookup[key] = int(row["id"])
    return lookup


def _try_resolve_product_id(cur, article: str, name: str, name_lookup: dict[str, int] | None = None):
    """Совпадение наименования или артикула с номенклатурой (точное и нормализованное имя)."""
    name_n = _normalize_str(name)
    art_n = _normalize_str(article)
    if name_n:
        row = cur.execute(
            "SELECT id FROM products WHERE name = ? LIMIT 1",
            (name_n,),
        ).fetchone()
        if row:
            return int(row["id"])
        norm_key = _normalize_product_name_for_import_match(name)
        if norm_key:
            if name_lookup is not None:
                pid = name_lookup.get(norm_key)
                if pid is not None:
                    return pid
            else:
                for prow in cur.execute("SELECT id, name FROM products"):
                    if _normalize_product_name_for_import_match(prow["name"] or "") == norm_key:
                        return int(prow["id"])
    if art_n:
        row = cur.execute(
            "SELECT id FROM products WHERE article = ? LIMIT 1",
            (art_n,),
        ).fetchone()
        if row:
            return int(row["id"])
    return None


def _is_lab_industries_client(client: str) -> bool:
    rec = _unique_client_record(client)
    return bool(rec and rec.get("pallet_kind") == "lab")


def _is_drogeri_retail_client(client: str) -> bool:
    return _normalize_client_key(client) == "дрогери ритейл"


def _parse_item_total_order_quantity(raw: dict, lab_client: bool) -> float | None | dict:
    """Общее количество заказа по позиции (ЛАБ), для паллетного листа."""
    if not lab_client:
        return None
    val_raw = raw.get("total_order_quantity")
    if val_raw is None or str(val_raw).strip() == "":
        return None
    try:
        val = float(str(val_raw).replace(",", "."))
    except (TypeError, ValueError):
        return {
            "error": "validation",
            "message": "Укажите корректное общее количество заказа по позиции.",
        }
    if val < 0:
        return {
            "error": "validation",
            "message": "Общее количество заказа не может быть отрицательным.",
        }
    return val


def _resolve_normalized_product_ids(normalized, cur, *, allow_unresolved: bool = False):
    """Подставляет product_id и канонические article/name из БД, если id не был передан."""
    out = []
    for pid, article, name, qty, unit, buyer_order, line_total_qty in normalized:
        if pid is None:
            rid = _try_resolve_product_id(cur, article, name)
            if rid is None:
                if allow_unresolved and name:
                    out.append(
                        (None, article, name, qty, unit, buyer_order, line_total_qty)
                    )
                    continue
                return None, {
                    "error": "validation",
                    "message": "Для позиций без привязки к номенклатуре выберите товар в поле «Наименование» из подсказки или приведите название к точному совпадению с номенклатурой.",
                }
            row = cur.execute(
                "SELECT article, name FROM products WHERE id = ?",
                (rid,),
            ).fetchone()
            if not row:
                return None, {
                    "error": "validation",
                    "message": "Товар не найден в номенклатуре.",
                }
            article = _normalize_str(row["article"])
            name = _normalize_str(row["name"])
            pid = rid
        if not name and not article:
            return None, {
                "error": "validation",
                "message": "Для позиции не указано наименование.",
            }
        out.append((pid, article, name, qty, unit, buyer_order, line_total_qty))
    return out, None


def _normalize_order_items_body(body: dict, *, allow_zero_quantity: bool = False):
    """Общая валидация позиций для создания и обновления заказа."""
    ship_raw = body.get("ship_date", "")
    client_raw = body.get("client", "")
    ship = _format_ship_date_storage(ship_raw)
    client_n = _normalize_str(client_raw)
    if not ship or not client_n:
        return {
            "error": "validation",
            "message": "Укажите дату отгрузки и клиента.",
        }
    lab_client = _is_lab_industries_client(client_n)
    drogeri_client = _is_drogeri_retail_client(client_n)
    buyer_order_mode = ""
    order_buyer_order = ""
    if lab_client:
        mode_raw = str(body.get("buyer_order_mode") or "multiple").strip().lower()
        buyer_order_mode = (
            "multiple" if mode_raw in ("multiple", "multi", "several") else "single"
        )
        order_buyer_order = _normalize_str(body.get("buyer_order") or "")
        if buyer_order_mode == "single" and not order_buyer_order:
            return {
                "error": "validation",
                "message": "Укажите заказ покупателя.",
            }
    elif drogeri_client:
        buyer_order_mode = "multiple"
    client_city = _normalize_str(body.get("client_city") or "") if drogeri_client else ""
    raw_items = body.get("items")
    if not isinstance(raw_items, list) or not raw_items:
        return {
            "error": "validation",
            "message": "Добавьте хотя бы одну позицию заказа.",
        }
    normalized = []
    for raw in raw_items:
        if not isinstance(raw, dict):
            continue
        try:
            qty = float(raw.get("quantity", 0))
        except (TypeError, ValueError):
            qty = 0
        unit = (raw.get("unit") or "piece").strip().lower()
        if unit not in ("box", "set", "piece"):
            unit = "piece"
        name = _strip_trailing_name_suffix(_normalize_str(raw.get("name", "")))
        article = _normalize_str(raw.get("article", ""))
        pid_raw = raw.get("product_id")
        pid = None
        if pid_raw is not None and str(pid_raw).strip() != "":
            try:
                pid = int(pid_raw)
            except (TypeError, ValueError):
                pid = None
        if qty <= 0 and not allow_zero_quantity:
            return {
                "error": "validation",
                "message": "Укажите количество больше нуля по каждой позиции.",
            }
        if pid is None and not name and not article:
            return {
                "error": "validation",
                "message": "Для каждой позиции укажите наименование или выберите товар из подсказки.",
            }
        if pid is not None and not name and not article:
            return {
                "error": "validation",
                "message": "Для позиции не указано наименование.",
            }
        line_buyer = _normalize_str(raw.get("buyer_order") or "")
        if lab_client and buyer_order_mode == "multiple":
            if not line_buyer:
                return {
                    "error": "validation",
                    "message": "Укажите заказ покупателя по каждой позиции.",
                }
        elif drogeri_client:
            if not line_buyer:
                return {
                    "error": "validation",
                    "message": "Укажите номер заказа по каждой позиции.",
                }
        elif lab_client:
            line_buyer = ""
        line_total_qty = _parse_item_total_order_quantity(raw, lab_client)
        if isinstance(line_total_qty, dict):
            return line_total_qty
        normalized.append((pid, article, name, qty, unit, line_buyer, line_total_qty))
    if not normalized:
        return {
            "error": "validation",
            "message": "Нет корректных позиций в заказе.",
        }
    return {
        "ok": True,
        "ship": ship,
        "client": client_n,
        "normalized": normalized,
        "buyer_order_mode": buyer_order_mode,
        "buyer_order": order_buyer_order,
        "client_city": client_city,
    }


ORDER_DELIVERY_TYPES = (
    "Обычный",
    "Деловые Линии",
    "Самовывоз",
    "Письмо",
    "Арнест Юнирусь",
    "ЛАБ Индастриз",
    "Дрогери Ритейл",
)


def _is_arnest_unirus_client(client: str) -> bool:
    rec = _unique_client_record(client)
    return bool(rec and rec.get("pallet_kind") == "arnest")


def _normalize_delivery_type(value: str) -> str:
    t = _normalize_str(value)
    if not t:
        return "Обычный"
    aliases = {
        "обычный": "Обычный",
        "деловые линии": "Деловые Линии",
        "самовывоз": "Самовывоз",
        "письмо": "Письмо",
        "арнест юнирусь": "Арнест Юнирусь",
        "лаб индастриз": "ЛАБ Индастриз",
        "дрогери ритейл": "Дрогери Ритейл",
    }
    mapped = aliases.get(t.lower())
    if mapped:
        return mapped
    if t in ORDER_DELIVERY_TYPES:
        return t
    return "Обычный"


def _delivery_type_for_client(client: str, requested: str = "") -> str:
    rec = _unique_client_record(client)
    if rec:
        return rec.get("delivery_type") or _normalize_delivery_type(requested)
    return _normalize_delivery_type(requested)


def insert_order_with_items(body: dict, *, modified_by: str = ""):
    pack = _normalize_order_items_body(body)
    if pack.get("error"):
        return pack
    ship = pack["ship"]
    client_n = pack["client"]
    delivery_type = _delivery_type_for_client(client_n, body.get("extra_info", ""))
    modifier = _normalize_str(modified_by)
    normalized = pack["normalized"]
    buyer_order_mode = pack.get("buyer_order_mode") or ""
    order_buyer_order = pack.get("buyer_order") or ""
    client_city = pack.get("client_city") or ""
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        normalized, res_err = _resolve_normalized_product_ids(normalized, cur)
        if res_err:
            con.close()
            return res_err
        names_summary = _format_order_names_summary(normalized)
        cur.execute(
            """
            INSERT INTO orders (
              ship_date, client, assembled_percent, names, extra_info,
              buyer_order_mode, buyer_order, client_city, last_edited_by
            )
            VALUES (?, ?, 0, ?, ?, ?, ?, ?, ?)
            """,
            (
                ship,
                client_n,
                names_summary,
                delivery_type,
                buyer_order_mode,
                order_buyer_order,
                client_city,
                modifier,
            ),
        )
        oid = cur.lastrowid
        for pid, article, name, qty, unit, line_buyer, line_total_qty in normalized:
            cur.execute(
                """
                INSERT INTO order_items (
                  order_id, product_id, article, name, quantity, unit, buyer_order,
                  total_order_quantity
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (oid, pid, article, name, qty, unit, line_buyer, line_total_qty),
            )
        if _is_lab_industries_client(client_n):
            _lab_sscc_sync_all_unshipped_seq_starts(cur)
        con.commit()
        con.close()
    return {"ok": True, "id": int(oid)}


_IMPORT_SKIP_META_PREFIX = "__import_skip__:"


def _encode_import_skip_extra_info(
    skipped: int,
    baseline: int,
    user_text: str = "",
    skipped_names: list | None = None,
) -> str:
    if skipped <= 0:
        return (user_text or "").strip()
    names_out: list[str] = []
    if skipped_names:
        for raw in skipped_names:
            n = _normalize_str(str(raw or ""))
            if n:
                names_out.append(n)
    payload: dict = {
        "skipped": int(skipped),
        "baseline": int(baseline),
        "names": names_out,
    }
    meta = _IMPORT_SKIP_META_PREFIX + json.dumps(payload, ensure_ascii=False)
    user_text = (user_text or "").strip()
    if user_text:
        return meta + "\n" + user_text
    return meta


def _parse_import_skip_extra_info(extra_info: str) -> tuple[int, int, list[str], str]:
    raw = extra_info or ""
    if not raw.startswith(_IMPORT_SKIP_META_PREFIX):
        return 0, 0, [], raw
    rest = raw[len(_IMPORT_SKIP_META_PREFIX) :]
    nl = rest.find("\n")
    if nl >= 0:
        json_part = rest[:nl]
        user_text = rest[nl + 1 :]
    else:
        json_part = rest
        user_text = ""
    try:
        data = json.loads(json_part)
        skipped = max(0, int(data.get("skipped", 0)))
        baseline = max(0, int(data.get("baseline", 0)))
        names_raw = data.get("names")
        if isinstance(names_raw, list):
            skipped_names = [
                _normalize_str(str(x)) for x in names_raw if _normalize_str(str(x))
            ]
        else:
            skipped_names = []
    except (json.JSONDecodeError, TypeError, ValueError):
        return 0, 0, [], raw
    return skipped, baseline, skipped_names, user_text


def _order_list_fields_from_extra_info(extra_info: str, client: str = "") -> dict:
    skipped, _baseline, skipped_names, clean = _parse_import_skip_extra_info(extra_info)
    clean = _delivery_type_for_client(client, clean)
    return {
        "extra_info": clean,
        "import_skipped_lines": skipped,
        "import_skipped_names": skipped_names,
    }


def _order_editor_fields_from_row(row) -> dict:
    keys = row.keys() if hasattr(row, "keys") else []
    edited = ""
    if "last_edited_by" in keys:
        edited = (row["last_edited_by"] or "").strip()
    elif "last_modified_by" in keys:
        edited = (row["last_modified_by"] or "").strip()
    assembled = ""
    if "last_assembled_by" in keys:
        assembled = (row["last_assembled_by"] or "").strip()
    return {
        "last_edited_by": edited,
        "last_assembled_by": assembled,
    }


_EXCEL_ORDER_COL_CLIENT = "C"
_EXCEL_ORDER_COL_PRODUCT = "I"
_EXCEL_ORDER_COL_QTY = "J"
_EXCEL_ORDER_COL_UNIT = "K"
_EXCEL_ORDER_COL_SHIP = "M"
_EXCEL_ORDER_COL_DELIVERY = "L"

_EXCEL_DELIVERY_RE_SAMOVYVOZ = re.compile(
    r"(?<![\w])самовывоз(?![\w])",
    re.IGNORECASE | re.UNICODE,
)
_EXCEL_DELIVERY_RE_DELOVYE = re.compile(
    r"(?<![\w])деловые\s+линии(?![\w])",
    re.IGNORECASE | re.UNICODE,
)


def _excel_delivery_extra_info_from_texts(texts: list) -> str:
    """Доп. информация заказа по комментариям из столбца L (целые слова/фразы)."""
    parts = []
    for raw in texts or []:
        t = _normalize_str(str(raw or ""))
        if t:
            parts.append(t)
    combined = " ".join(parts)
    if not combined.strip():
        return "Обычный"
    if _EXCEL_DELIVERY_RE_SAMOVYVOZ.search(combined):
        return "Самовывоз"
    if _EXCEL_DELIVERY_RE_DELOVYE.search(combined):
        return "Деловые Линии"
    return "Обычный"


def _excel_row_delivery_text(ws, row: int) -> str:
    raw = _excel_cell(ws, _EXCEL_ORDER_COL_DELIVERY, row)
    if raw is None or str(raw).strip() == "":
        return ""
    return _normalize_str(str(raw))


def _excel_cell(ws, col: str, row: int):
    return ws[f"{col}{row}"].value


def _excel_row_is_empty(ws, row: int) -> bool:
    for col in (
        _EXCEL_ORDER_COL_CLIENT,
        _EXCEL_ORDER_COL_PRODUCT,
        _EXCEL_ORDER_COL_QTY,
        _EXCEL_ORDER_COL_UNIT,
        _EXCEL_ORDER_COL_SHIP,
    ):
        val = _excel_cell(ws, col, row)
        if val is not None and str(val).strip() != "":
            return False
    return True


def _excel_value_to_ship_iso(value) -> str:
    """Дата отгрузки из ячейки M: число/дата Excel или фрагмент в тексте («Отгрузка 15.03.2026»)."""
    if value is None or str(value).strip() == "":
        return ""
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    s = re.sub(r"\s+", " ", str(value).strip().replace("\n", " ").replace("\r", " "))
    if not s:
        return ""

    def _iso_from_ymd(y: int, mo: int, d: int) -> str:
        try:
            return datetime.date(y, mo, d).isoformat()
        except ValueError:
            return ""

    m_iso = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m_iso:
        found = _iso_from_ymd(
            int(m_iso.group(1)), int(m_iso.group(2)), int(m_iso.group(3))
        )
        if found:
            return found

    m_dmy = re.search(r"(\d{1,2})[./](\d{1,2})[./](\d{2,4})", s)
    if m_dmy:
        d, mo, y = int(m_dmy.group(1)), int(m_dmy.group(2)), int(m_dmy.group(3))
        if y < 100:
            y += 2000
        return _iso_from_ymd(y, mo, d)

    return ""


def _parse_localized_number_string(raw: str) -> float | None:
    """Число из текста ячейки: 1500, 1 500, 1.500,5, 1,500.25, «10 шт»."""
    s = str(raw or "").strip()
    if not s:
        return None
    s = (
        s.replace("\u00a0", " ")
        .replace("\u202f", " ")
        .replace("\u2212", "-")
    )
    s = re.sub(r"\s+", " ", s)
    # убрать валюту/единицы, оставить цифры и разделители
    s = re.sub(r"[^\d,.\-+\s]", "", s).strip()
    if not s or s in ("-", "+"):
        return None

    def _to_float(token: str) -> float | None:
        token = token.strip().replace(" ", "")
        if not token or token in ("-", "+"):
            return None
        if "," in token and "." in token:
            if token.rfind(",") > token.rfind("."):
                token = token.replace(".", "").replace(",", ".")
            else:
                token = token.replace(",", "")
        elif "," in token:
            parts = [p.strip() for p in token.split(",")]
            if len(parts) == 2:
                # Одна запятая — десятичный разделитель (800,000 → 800; 10,5 → 10.5).
                int_part = parts[0].replace(" ", "").replace(".", "")
                frac_part = parts[1].replace(" ", "")
                if not int_part:
                    int_part = "0"
                token = f"{int_part}.{frac_part}" if frac_part else int_part
            else:
                # Несколько запятых — группировка тысяч (1,234,567).
                token = "".join(p.replace(" ", "") for p in parts)
        elif "." in token:
            parts = token.split(".")
            if len(parts) > 2 and all(len(p) == 3 for p in parts[1:-1]):
                token = "".join(parts[:-1]) + "." + parts[-1]
            elif len(parts) == 2 and len(parts[1]) == 3 and len(parts[0]) <= 4:
                token = parts[0] + parts[1]
        try:
            return float(token)
        except (TypeError, ValueError):
            return None

    direct = _to_float(s)
    if direct is not None and math.isfinite(direct):
        return direct

    for chunk in re.findall(
        r"-?\d{1,3}(?:[ \u00a0\u202f]\d{3})*(?:[.,]\d+)?|-?\d+(?:[.,]\d+)?",
        s,
    ):
        parsed = _to_float(chunk)
        if parsed is not None and math.isfinite(parsed):
            return parsed
    return None


def _parse_excel_quantity(value) -> tuple[float, bool]:
    """Количество из колонки J: (значение ≥ 0, распознано ли число)."""
    if value is None:
        return 0.0, False
    if isinstance(value, bool):
        return 0.0, False
    if isinstance(value, (int, float)):
        try:
            q = float(value)
        except (TypeError, ValueError):
            return 0.0, False
        if not math.isfinite(q):
            return 0.0, False
        return max(0.0, q), True
    try:
        from decimal import Decimal

        if isinstance(value, Decimal):
            q = float(value)
            if not math.isfinite(q):
                return 0.0, False
            return max(0.0, q), True
    except ImportError:
        pass

    s = str(value).strip()
    if not s:
        return 0.0, False
    parsed = _parse_localized_number_string(s)
    if parsed is None or not math.isfinite(parsed):
        return 0.0, False
    return max(0.0, parsed), True


def _parse_excel_unit(value) -> str:
    raw = str(value or "").strip().lower().replace("\u00a0", " ")
    if not raw:
        return "piece"
    compact = re.sub(r"\s+", "", raw)
    if compact in ("шт", "штука", "штуки", "штук", "piece", "pcs", "pc"):
        return "piece"
    if compact in ("наб", "набор", "наборы", "наборов", "set", "sets"):
        return "set"
    if compact in (
        "кор",
        "кор.",
        "короб",
        "коробка",
        "коробки",
        "коробок",
        "box",
        "boxes",
    ):
        return "box"
    if "набор" in raw:
        return "set"
    if "короб" in raw or raw.startswith("кор"):
        return "box"
    if "шт" in raw or "штук" in raw:
        return "piece"
    return "piece"


_DROGERI_EXCEL_CLIENT = "Дрогери Ритейл"


def _drogeri_excel_header_key(value) -> str:
    s = str(value or "").strip().lower().replace("ё", "е")
    s = s.replace("\u00a0", " ").replace("\u202f", " ")
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[^\wа-я]", "", s)
    return s


def _drogeri_excel_buyer_order_header_score(value) -> int:
    """Чем выше — тем предпочтительнее колонка «Номер заказа» (не «в 1С»)."""
    key = _drogeri_excel_header_key(value)
    if not key or "номерзаказа" not in key:
        return -1
    if key == "номерзаказа":
        return 100
    if "1с" in key or "в1с" in key:
        return 10
    return 50


def _drogeri_excel_match_header(value) -> str | None:
    key = _drogeri_excel_header_key(value)
    if not key:
        return None
    if "номерзаказа" in key:
        return "buyer_order"
    if "наименование" in key:
        return "name"
    if key in ("колво", "количество") or key.startswith("колво") or key.startswith("количество"):
        return "qty"
    if key in ("едизм", "единицаизмерения") or key.startswith("едизм") or "единицаизмерения" in key:
        return "unit"
    return None


def _excel_cell_at(ws, col_idx: int, row: int):
    return ws[f"{get_column_letter(col_idx)}{row}"].value


def _find_drogeri_excel_column_map(ws) -> tuple[dict[str, tuple[int, int]], list[str]]:
    """Колонки по подписям: поле → (номер колонки, строка заголовка)."""
    found: dict[str, tuple[int, int]] = {}
    buyer_order_best: tuple[int, int, int] = (-1, 0, 0)  # score, col, row
    max_row = min(int(ws.max_row or 0), 120)
    max_col = min(int(ws.max_column or 0), 60) if ws.max_column else 60
    if max_row <= 0:
        return {}, ["Лист Excel пустой."]
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell_val = _excel_cell_at(ws, col, row)
            kind = _drogeri_excel_match_header(cell_val)
            if not kind:
                continue
            if kind == "buyer_order":
                score = _drogeri_excel_buyer_order_header_score(cell_val)
                if score > buyer_order_best[0]:
                    buyer_order_best = (score, col, row)
                continue
            if kind not in found:
                found[kind] = (col, row)
    if buyer_order_best[0] >= 0:
        found["buyer_order"] = (buyer_order_best[1], buyer_order_best[2])
    errors: list[str] = []
    labels = {
        "buyer_order": "Номер заказа",
        "name": "Наименование",
        "qty": "Кол-во / Количество",
        "unit": "Ед. изм. / Единица измерения",
    }
    for field, label in labels.items():
        if field not in found:
            errors.append(f"Не найдена колонка «{label}».")
    return found, errors


def _drogeri_excel_data_row_empty(ws, row: int, columns: dict[str, tuple[int, int]]) -> bool:
    for field in ("name", "qty", "buyer_order", "unit"):
        col, _ = columns[field]
        val = _excel_cell_at(ws, col, row)
        if val is not None and str(val).strip() != "":
            return False
    return True


def _parse_drogeri_excel_buyer_order(value) -> str:
    """Номер заказа — только целое число (ячейка с цифрами, без букв вроде ПО/ЦБА)."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, (int, float)):
        try:
            num = float(value)
        except (TypeError, ValueError):
            return ""
        if not math.isfinite(num) or num < 0:
            return ""
        if num != int(num):
            return ""
        return str(int(num))
    s = str(value).strip()
    if not s:
        return ""
    compact = (
        s.replace("\u00a0", "")
        .replace("\u202f", "")
        .replace(" ", "")
    )
    if compact.isdigit():
        return compact
    return ""


def parse_drogeri_order_from_excel_worksheet(ws):
    """Один заказ «Дрогери Ритейл» по подписям колонок в Excel."""
    columns, col_errors = _find_drogeri_excel_column_map(ws)
    if col_errors:
        return None, col_errors
    header_row = max(row for _, row in columns.values())
    max_row = int(ws.max_row or 0)
    items: list[dict] = []
    errors: list[str] = []
    empty_streak = 0
    name_col = columns["name"][0]
    qty_col = columns["qty"][0]
    buyer_col = columns["buyer_order"][0]
    unit_col = columns["unit"][0]

    for row in range(header_row + 1, max_row + 1):
        if _drogeri_excel_data_row_empty(ws, row, columns):
            empty_streak += 1
            if empty_streak >= 3:
                break
            continue
        empty_streak = 0

        product_raw = _excel_cell_at(ws, name_col, row)
        product_s = (
            _strip_trailing_name_suffix(_normalize_str(product_raw))
            if product_raw is not None and str(product_raw).strip() != ""
            else ""
        )
        qty_raw = _excel_cell_at(ws, qty_col, row)
        qty, qty_recognized = _parse_excel_quantity(qty_raw)
        buyer_order = _parse_drogeri_excel_buyer_order(_excel_cell_at(ws, buyer_col, row))
        unit = _parse_excel_unit(_excel_cell_at(ws, unit_col, row))

        if not product_s and not qty_recognized:
            continue
        if not product_s:
            errors.append(f"Строка {row}: не указано наименование.")
            continue
        if not qty_recognized or qty <= 0:
            errors.append(f"Строка {row}: «{product_s}» — не распознано количество.")
            continue
        if not buyer_order:
            errors.append(f"Строка {row}: «{product_s}» — не указан номер заказа.")
            continue
        items.append(
            {
                "name": product_s,
                "quantity": qty,
                "unit": unit,
                "buyer_order": buyer_order,
            }
        )

    if not items:
        msg = "В файле нет позиций для заказа Дрогери Ритейл."
        if errors:
            msg = errors[0]
        return None, errors or [msg]

    return (
        {
            "client": _DROGERI_EXCEL_CLIENT,
            "ship_date": datetime.date.today().isoformat(),
            "items": items,
        },
        errors,
    )


def import_drogeri_order_from_excel_bytes(data: bytes, *, modified_by: str = ""):
    if not HAVE_OPENPYXL or load_workbook is None:
        return {
            "error": "no_openpyxl",
            "message": "На сервере не установлен openpyxl (pip install openpyxl).",
        }
    if not data:
        return {"error": "validation", "message": "Файл пустой."}
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception:
        return {
            "error": "invalid_file",
            "message": "Не удалось прочитать Excel. Используйте файл .xlsx.",
        }
    try:
        ws = wb.active
        order, parse_errors = parse_drogeri_order_from_excel_worksheet(ws)
    finally:
        wb.close()
    if not order:
        msg = "Не удалось разобрать файл для Дрогери Ритейл."
        if parse_errors:
            msg = parse_errors[0]
        return {
            "error": "validation",
            "message": msg,
            "details": parse_errors,
        }

    import_errors: list[str] = list(parse_errors)
    modifier = _normalize_str(modified_by)
    items_payload: list[dict] = []
    skipped_count = 0
    skipped_names: list[str] = []

    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        name_lookup = _build_product_name_import_lookup(cur)
        for it in order["items"]:
            pid, article, pname = _resolve_product_id_for_excel_import(
                cur, it["name"], name_lookup=name_lookup
            )
            excel_name = _normalize_str(it.get("name", ""))
            if pid is None:
                skipped_count += 1
                skipped_names.append(excel_name)
                items_payload.append(
                    {
                        "product_id": None,
                        "article": "",
                        "name": excel_name,
                        "quantity": float(it.get("quantity") or 0),
                        "unit": it.get("unit") or "piece",
                        "buyer_order": _normalize_str(it.get("buyer_order") or ""),
                    }
                )
                continue
            items_payload.append(
                {
                    "product_id": pid,
                    "article": article,
                    "name": pname,
                    "quantity": it["quantity"],
                    "unit": it["unit"],
                    "buyer_order": _normalize_str(it.get("buyer_order") or ""),
                }
            )

        if not items_payload:
            con.close()
            return {
                "error": "validation",
                "message": "В файле нет позиций для заказа.",
                "details": import_errors,
            }

        body = {
            "ship_date": order["ship_date"],
            "client": order["client"],
            "items": items_payload,
        }
        pack = _normalize_order_items_body(body, allow_zero_quantity=True)
        if pack.get("error"):
            con.close()
            return {
                "error": "validation",
                "message": pack.get("message", "ошибка валидации"),
                "details": import_errors,
            }
        normalized = pack["normalized"]
        normalized, res_err = _resolve_normalized_product_ids(
            normalized, cur, allow_unresolved=True
        )
        if res_err:
            con.close()
            return {
                "error": "validation",
                "message": res_err.get("message", "ошибка"),
                "details": import_errors,
            }
        ship = pack["ship"]
        client_n = pack["client"]
        names_summary = _format_order_names_summary(normalized)
        baseline = len(normalized)
        delivery_extra = _delivery_type_for_client(client_n, "Обычный")
        extra_info = _encode_import_skip_extra_info(
            skipped_count, baseline, delivery_extra, skipped_names
        )
        cur.execute(
            """
            INSERT INTO orders (
              ship_date, client, assembled_percent, names, extra_info,
              buyer_order_mode, last_edited_by
            )
            VALUES (?, ?, 0, ?, ?, ?, ?)
            """,
            (ship, client_n, names_summary, extra_info, "multiple", modifier),
        )
        oid = int(cur.lastrowid)
        for pid, article, name, qty, unit, line_buyer, line_total_qty in normalized:
            cur.execute(
                """
                INSERT INTO order_items (
                  order_id, product_id, article, name, quantity, unit, buyer_order,
                  total_order_quantity
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (oid, pid, article, name, qty, unit, line_buyer, line_total_qty),
            )
        con.commit()
        con.close()

    if skipped_count > 0:
        import_errors.append(
            _format_excel_import_skipped_message(client_n, skipped_count, skipped_names)
        )

    return {
        "ok": True,
        "id": oid,
        "created_ids": [oid],
        "count": 1,
        "errors": import_errors,
    }


def parse_orders_from_excel_worksheet(ws):
    """Строки с клиентом в C начинают заказ; следующие строки без C — позиции того же заказа.

    Дата в M: при объединённых ячейках Excel значение часто только в первой строке блока.
    Если в строке M пусто — используется последняя дата из M выше, пока не встретится новая.
    """
    errors: list[str] = []
    orders: list[dict] = []
    current: dict | None = None
    last_ship_iso = ""
    last_qty = 0.0
    last_qty_recognized = False

    def flush_current():
        nonlocal current
        if not current:
            return
        client = current.get("client") or ""
        ship = current.get("ship_date") or ""
        items = current.get("items") or []
        start = current.get("_start_row", "?")
        if not client and not items:
            current = None
            return
        if not client:
            errors.append(f"Строка {start}: в заказе не указан клиент (колонка C).")
        elif not ship:
            errors.append(f"Заказ «{client}» (строка {start}): не указана дата отгрузки (колонка M).")
        else:
            delivery_texts = current.get("_delivery_texts") or []
            orders.append(
                {
                    "client": client,
                    "ship_date": ship,
                    "items": items,
                    "extra_info": _delivery_type_for_client(
                        client,
                        _excel_delivery_extra_info_from_texts(delivery_texts),
                    ),
                    "_delivery_texts": list(delivery_texts),
                }
            )
        current = None

    def _append_delivery_text_to_current(row: int) -> None:
        nonlocal current
        if current is None:
            return
        t = _excel_row_delivery_text(ws, row)
        if t:
            current.setdefault("_delivery_texts", []).append(t)

    max_row = int(ws.max_row or 0)
    for row in range(2, max_row + 1):
        if _excel_row_is_empty(ws, row):
            continue
        client_raw = _excel_cell(ws, _EXCEL_ORDER_COL_CLIENT, row)
        ship_raw = _excel_cell(ws, _EXCEL_ORDER_COL_SHIP, row)
        product_raw = _excel_cell(ws, _EXCEL_ORDER_COL_PRODUCT, row)
        qty_raw = _excel_cell(ws, _EXCEL_ORDER_COL_QTY, row)
        unit_raw = _excel_cell(ws, _EXCEL_ORDER_COL_UNIT, row)

        if ship_raw is not None and str(ship_raw).strip() != "":
            ship_iso_row = _excel_value_to_ship_iso(ship_raw)
            if ship_iso_row:
                last_ship_iso = ship_iso_row

        client_s = (
            _normalize_str(client_raw)
            if client_raw is not None and str(client_raw).strip() != ""
            else ""
        )
        if client_s:
            flush_current()
            last_qty = 0.0
            last_qty_recognized = False
            current = {
                "client": client_s,
                "ship_date": last_ship_iso,
                "items": [],
                "_delivery_texts": [],
                "_start_row": row,
            }
            _append_delivery_text_to_current(row)
        elif current is not None and last_ship_iso and not current.get("ship_date"):
            current["ship_date"] = last_ship_iso
        else:
            _append_delivery_text_to_current(row)

        product_s = (
            _strip_trailing_name_suffix(_normalize_str(product_raw))
            if product_raw is not None and str(product_raw).strip() != ""
            else ""
        )
        if not product_s:
            continue
        if current is None:
            errors.append(
                f"Строка {row}: позиция без заказа — сначала укажите клиента в колонке C."
            )
            continue
        if not current.get("ship_date") and last_ship_iso:
            current["ship_date"] = last_ship_iso
        qty, qty_recognized = _parse_excel_quantity(qty_raw)
        if qty_recognized:
            last_qty = qty
            last_qty_recognized = True
        elif qty_raw is None or str(qty_raw).strip() == "":
            if last_qty_recognized:
                qty = last_qty
                qty_recognized = True
            else:
                q_unit, ok_unit = _parse_excel_quantity(unit_raw)
                if ok_unit:
                    qty = q_unit
                    qty_recognized = True
                    last_qty = qty
                    last_qty_recognized = True
        if not qty_recognized:
            hint = ""
            if qty_raw is not None and str(qty_raw).strip().startswith("="):
                hint = " Сохраните файл в Excel перед загрузкой (формула без значения)."
            errors.append(
                f"Строка {row}: количество в колонке J не распознано "
                f"(«{product_s}») — позиция добавлена с количеством 0.{hint}"
            )
            qty = 0.0
        unit = _parse_excel_unit(unit_raw)
        current["items"].append(
            {"name": product_s, "quantity": qty, "unit": unit}
        )

    flush_current()
    return merge_parsed_excel_orders(orders), errors


def merge_parsed_excel_orders(orders: list[dict]) -> list[dict]:
    """Один заказ на пару (клиент, дата отгрузки); одинаковые товары — сумма количества."""
    if not orders:
        return []
    buckets: dict[tuple[str, str], dict] = {}
    order_keys: list[tuple[str, str]] = []

    for order in orders:
        client = _normalize_str(order.get("client", ""))
        ship = (order.get("ship_date") or "").strip()
        if not client or not ship:
            continue
        key = (client, ship)
        if key not in buckets:
            buckets[key] = {
                "client": client,
                "ship_date": ship,
                "items": [],
                "_delivery_texts": [],
            }
            order_keys.append(key)
        bucket = buckets[key]
        bucket["_delivery_texts"].extend(order.get("_delivery_texts") or [])
        item_map: dict[tuple[str, str], dict] = {
            (
                _strip_trailing_name_suffix(_normalize_str(it.get("name", ""))),
                (it.get("unit") or "piece"),
            ): dict(it)
            for it in bucket["items"]
        }
        for it in order.get("items") or []:
            name = _strip_trailing_name_suffix(_normalize_str(it.get("name", "")))
            unit = (it.get("unit") or "piece").strip().lower()
            if unit not in ("box", "set", "piece"):
                unit = "piece"
            try:
                qty = float(it.get("quantity") or 0)
            except (TypeError, ValueError):
                qty = 0
            if not name:
                continue
            ik = (name, unit)
            if ik in item_map:
                item_map[ik]["quantity"] = float(item_map[ik]["quantity"]) + qty
            else:
                item_map[ik] = {"name": name, "quantity": qty, "unit": unit}
        bucket["items"] = list(item_map.values())

    out: list[dict] = []
    for key in order_keys:
        b = buckets[key]
        delivery_texts = b.pop("_delivery_texts", [])
        out.append(
            {
                "client": b["client"],
                "ship_date": b["ship_date"],
                "items": b["items"],
                "extra_info": _delivery_type_for_client(
                    b["client"],
                    _excel_delivery_extra_info_from_texts(delivery_texts),
                ),
            }
        )
    return out


def _merge_import_items_payload(items: list[dict]) -> list[dict]:
    """Суммирует количество по product_id (или наименованию) и единице."""
    merged: dict[tuple, dict] = {}
    for it in items:
        unit = (it.get("unit") or "piece").strip().lower()
        if unit not in ("box", "set", "piece"):
            unit = "piece"
        pid_raw = it.get("product_id")
        if pid_raw is not None and str(pid_raw).strip() != "":
            key = ("id", int(pid_raw), unit)
        else:
            key = ("name", _normalize_str(it.get("name", "")), unit)
        if key in merged:
            merged[key]["quantity"] = float(merged[key]["quantity"]) + float(
                it.get("quantity") or 0
            )
        else:
            merged[key] = {
                "product_id": int(pid_raw) if pid_raw is not None else None,
                "article": it.get("article", ""),
                "name": it.get("name", ""),
                "quantity": float(it.get("quantity") or 0),
                "unit": unit,
            }
    return list(merged.values())


def _format_excel_import_skipped_message(
    client: str, skipped_count: int, skipped_names: list[str]
) -> str:
    """Текст предупреждения при импорте: сколько строк и какие наименования не нашлись."""
    client_s = _normalize_str(client) or "?"
    if skipped_count == 1:
        word = "строка"
    elif 2 <= skipped_count <= 4:
        word = "строки"
    else:
        word = "строк"
    msg = (
        f"Заказ «{client_s}»: пропущено {skipped_count} {word} "
        f"(нет совпадения в номенклатуре)."
    )
    preview = [_normalize_str(n) for n in (skipped_names or []) if _normalize_str(n)]
    if preview:
        shown = preview[:5]
        tail = f" и ещё {len(preview) - len(shown)}" if len(preview) > len(shown) else ""
        msg += " " + "; ".join(shown) + tail
    return msg


def _resolve_product_id_for_excel_import(
    cur, name: str, name_lookup: dict[str, int] | None = None
):
    """Совпадение наименования с номенклатурой (точное + нормализация как в поиске, без нечёткого подбора)."""
    name_n = _strip_trailing_name_suffix(_normalize_str(name))
    if not name_n:
        return None, "", ""
    rid = _try_resolve_product_id(cur, "", name_n, name_lookup=name_lookup)
    if rid is None:
        return None, "", name_n
    row = cur.execute(
        "SELECT article, name FROM products WHERE id = ?",
        (rid,),
    ).fetchone()
    if not row:
        return None, "", name_n
    return (
        rid,
        _normalize_str(row["article"]),
        _normalize_str(row["name"]),
    )


def import_orders_from_excel_bytes(data: bytes, *, modified_by: str = ""):
    if not HAVE_OPENPYXL or load_workbook is None:
        return {
            "error": "no_openpyxl",
            "message": "На сервере не установлен openpyxl (pip install openpyxl).",
        }
    if not data:
        return {"error": "validation", "message": "Файл пустой."}
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception:
        return {
            "error": "invalid_file",
            "message": "Не удалось прочитать Excel. Используйте файл .xlsx.",
        }
    try:
        ws = wb.active
        parsed_orders, parse_errors = parse_orders_from_excel_worksheet(ws)
    finally:
        wb.close()
    if not parsed_orders:
        msg = "В файле нет заказов для импорта."
        if parse_errors:
            msg = parse_errors[0]
        return {
            "error": "validation",
            "message": msg,
            "details": parse_errors,
        }
    created_ids: list[int] = []
    import_errors: list[str] = list(parse_errors)
    modifier = _normalize_str(modified_by)
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        name_lookup = _build_product_name_import_lookup(cur)
        for order in parsed_orders:
            items_payload = []
            skipped_count = 0
            skipped_names: list[str] = []
            for it in order["items"]:
                pid, article, pname = _resolve_product_id_for_excel_import(
                    cur, it["name"], name_lookup=name_lookup
                )
                excel_name = _normalize_str(it.get("name", ""))
                if pid is None:
                    skipped_count += 1
                    skipped_names.append(excel_name)
                    items_payload.append(
                        {
                            "product_id": None,
                            "article": "",
                            "name": excel_name,
                            "quantity": float(it.get("quantity") or 0),
                            "unit": it.get("unit") or "piece",
                        }
                    )
                    continue
                items_payload.append(
                    {
                        "product_id": pid,
                        "article": article,
                        "name": pname,
                        "quantity": it["quantity"],
                        "unit": it["unit"],
                    }
                )
            items_payload = _merge_import_items_payload(items_payload)
            if not items_payload:
                ship = _format_ship_date_storage(order.get("ship_date", ""))
                client_n = _normalize_str(order.get("client", ""))
                if not ship or not client_n:
                    import_errors.append(
                        f"Заказ «{order.get('client', '')}»: укажите дату отгрузки и клиента."
                    )
                    continue
                names_summary = ""
                baseline = 0
                delivery_extra = _delivery_type_for_client(
                    client_n, _normalize_str(order.get("extra_info", "")) or "Обычный"
                )
                extra_info = _encode_import_skip_extra_info(
                    skipped_count, baseline, delivery_extra, skipped_names
                )
                cur.execute(
                    """
                    INSERT INTO orders (
                      ship_date, client, assembled_percent, names, extra_info,
                      last_edited_by
                    )
                    VALUES (?, ?, 0, ?, ?, ?)
                    """,
                    (ship, client_n, names_summary, extra_info, modifier),
                )
                oid = int(cur.lastrowid)
                created_ids.append(oid)
                if skipped_count > 0:
                    import_errors.append(
                        _format_excel_import_skipped_message(
                            client_n, skipped_count, skipped_names
                        )
                        + " Заказ создан без позиций."
                    )
                continue
            body = {
                "ship_date": order["ship_date"],
                "client": order["client"],
                "items": items_payload,
            }
            pack = _normalize_order_items_body(body, allow_zero_quantity=True)
            if pack.get("error"):
                import_errors.append(
                    f"Заказ «{order['client']}»: {pack.get('message', 'ошибка валидации')}"
                )
                continue
            normalized = pack["normalized"]
            normalized, res_err = _resolve_normalized_product_ids(
                normalized, cur, allow_unresolved=True
            )
            if res_err:
                import_errors.append(
                    f"Заказ «{order['client']}»: {res_err.get('message', 'ошибка')}"
                )
                continue
            ship = pack["ship"]
            client_n = pack["client"]
            names_summary = _format_order_names_summary(normalized)
            baseline = len(normalized)
            delivery_extra = _delivery_type_for_client(
                client_n, _normalize_str(order.get("extra_info", "")) or "Обычный"
            )
            extra_info = _encode_import_skip_extra_info(
                skipped_count, baseline, delivery_extra, skipped_names
            )
            cur.execute(
                """
                INSERT INTO orders (
                  ship_date, client, assembled_percent, names, extra_info,
                  last_edited_by
                )
                VALUES (?, ?, 0, ?, ?, ?)
                """,
                (ship, client_n, names_summary, extra_info, modifier),
            )
            oid = int(cur.lastrowid)
            for pid, article, name, qty, unit, line_buyer, line_total_qty in normalized:
                cur.execute(
                    """
                    INSERT INTO order_items (
                      order_id, product_id, article, name, quantity, unit, buyer_order,
                      total_order_quantity
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (oid, pid, article, name, qty, unit, line_buyer, line_total_qty),
                )
            created_ids.append(oid)
            if skipped_count > 0:
                import_errors.append(
                    _format_excel_import_skipped_message(
                        order["client"], skipped_count, skipped_names
                    )
                )
        if created_ids:
            con.commit()
        con.close()
    if not created_ids:
        return {
            "error": "validation",
            "message": import_errors[0]
            if import_errors
            else "Не удалось создать заказы.",
            "details": import_errors,
        }
    return {
        "ok": True,
        "created_ids": created_ids,
        "count": len(created_ids),
        "errors": import_errors,
    }


def update_order_with_items(order_id: int, body: dict, *, modified_by: str = ""):
    pack = _normalize_order_items_body(body)
    if pack.get("error"):
        return pack
    ship = pack["ship"]
    client_n = pack["client"]
    normalized = pack["normalized"]
    buyer_order_mode = pack.get("buyer_order_mode") or ""
    order_buyer_order = pack.get("buyer_order") or ""
    client_city = pack.get("client_city") or ""
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        normalized, res_err = _resolve_normalized_product_ids(normalized, cur)
        if res_err:
            con.close()
            return res_err
        names_summary = _format_order_names_summary(normalized)
        row = cur.execute(
            """
            SELECT id, ship_date, client, assembled_percent, extra_info, assemble_state,
                   buyer_order_mode, buyer_order, client_city, order_readiness,
                   assemble_revision, assemble_state_updated_at
            FROM orders WHERE id = ?
            """,
            (order_id,),
        ).fetchone()
        if not row:
            con.close()
            return {"error": "not_found", "message": "Заказ не найден."}
        if _order_is_shipment_locked(row):
            con.close()
            return _order_shipment_locked_error()
        apct = max(0, min(100, int(row["assembled_percent"] or 0)))
        skipped, baseline, skipped_names, user_xinfo = _parse_import_skip_extra_info(
            row["extra_info"] or ""
        )
        if skipped > 0 and len(normalized) >= baseline + skipped:
            skipped = 0
            baseline = 0
            skipped_names = []
        delivery_type = _delivery_type_for_client(
            client_n, body.get("extra_info", user_xinfo)
        )
        xinfo = _encode_import_skip_extra_info(
            skipped, baseline, delivery_type, skipped_names
        )
        xasm = row["assemble_state"] or ""
        new_item_sigs = [_order_item_save_signature(t) for t in normalized]
        old_item_sigs = _fetch_order_item_signatures(cur, order_id)
        assemble_remapped = False
        if old_item_sigs != new_item_sigs and xasm:
            index_remap = _build_order_line_index_remap(old_item_sigs, new_item_sigs)
            xasm, assemble_remapped = _remap_assemble_state_line_indices(
                xasm, index_remap
            )
        assemble_rev = max(0, int(row["assemble_revision"] or 0))
        assemble_updated_at = (row["assemble_state_updated_at"] or "").strip()
        if assemble_remapped:
            assemble_rev += 1
            assemble_updated_at = datetime.datetime.now(datetime.timezone.utc).strftime(
                "%Y-%m-%dT%H:%M:%SZ"
            )
        if (
            (row["ship_date"] or "") == ship
            and _normalize_str(row["client"] or "") == client_n
            and (row["buyer_order_mode"] or "") == buyer_order_mode
            and (row["buyer_order"] or "") == order_buyer_order
            and _normalize_str(row["client_city"] or "") == client_city
            and (row["extra_info"] or "") == xinfo
            and old_item_sigs == new_item_sigs
        ):
            con.close()
            return {
                "ok": True,
                "id": int(order_id),
                "unchanged": True,
                "order_readiness": _order_readiness_from_row(row),
            }
        modifier = _normalize_str(modified_by)
        current_readiness = _order_readiness_from_row(row)
        effective_readiness = current_readiness
        if current_readiness == _ORDER_READINESS_ASSEMBLED:
            effective_readiness = _ORDER_READINESS_ASSEMBLING
        cur.execute(
            """
            UPDATE orders SET ship_date = ?, client = ?, names = ?, assembled_percent = ?,
              extra_info = ?, assemble_state = ?, buyer_order_mode = ?, buyer_order = ?,
              client_city = ?, order_readiness = ?,
              assemble_revision = ?, assemble_state_updated_at = ?,
              last_edited_by = CASE WHEN ? != '' THEN ? ELSE last_edited_by END
            WHERE id = ?
            """,
            (
                ship,
                client_n,
                names_summary,
                apct,
                xinfo,
                xasm,
                buyer_order_mode,
                order_buyer_order,
                client_city,
                effective_readiness,
                assemble_rev,
                assemble_updated_at,
                modifier,
                modifier,
                order_id,
            ),
        )
        cur.execute("DELETE FROM order_items WHERE order_id = ?", (order_id,))
        for pid, article, name, qty, unit, line_buyer, line_total_qty in normalized:
            cur.execute(
                """
                INSERT INTO order_items (
                  order_id, product_id, article, name, quantity, unit, buyer_order,
                  total_order_quantity
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (order_id, pid, article, name, qty, unit, line_buyer, line_total_qty),
            )
        con.commit()
        con.close()
    return {"ok": True, "id": int(order_id), "order_readiness": effective_readiness}


def _order_item_row_to_dict(ir):
    """Строка JOIN order_items + products → dict как в fetch_order_detail."""
    pid = ir["product_id"]
    keys = ir.keys() if hasattr(ir, "keys") else ()
    vol = float(ir["volume_ml"] or 0) if "volume_ml" in keys else 0.0
    vol_unit = _normalize_volume_unit(ir["volume_unit"]) if "volume_unit" in keys else "ml"
    return {
        "product_id": int(pid) if pid is not None else None,
        "article": ir["article"] or "",
        "name": ir["name"] or "",
        "quantity": float(ir["quantity"] or 0),
        "unit": (ir["unit"] or "piece").strip().lower(),
        "buyer_order": (ir["buyer_order"] or "").strip()
        if "buyer_order" in keys
        else "",
        "total_order_quantity": float(ir["total_order_quantity"])
        if "total_order_quantity" in keys and ir["total_order_quantity"] is not None
        else None,
        "pieces_in_box": max(0, int(ir["pieces_in_box"] or 0)),
        "sets_in_box": max(1, int(ir["sets_in_box"] or 1)),
        "pieces_per_set": max(1, int(ir["pieces_per_set"] or 1)),
        "row_layout": max(0, int(ir["row_layout"] or 0)),
        "max_rows": max(0, int(ir["max_rows"] or 0)),
        "box_weight": float(ir["box_weight"] or 0),
        "volume_ml": vol,
        "volume_unit": vol_unit,
    }


def _normalize_volume_unit(value) -> str:
    u = (value or "").strip().lower()
    if u in ("g", "gr", "г", "гр", "gram", "grams"):
        return "g"
    return "ml"


def fetch_order_detail(order_id: int):
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        row = cur.execute(
            """
            SELECT id, ship_date, client, assembled_percent, names, extra_info, assemble_state,
                   buyer_order_mode, buyer_order, client_city, total_order_quantity,
                   lab_sscc_seq_start, lab_sscc_shipped, lab_sscc_pallet_count,
                   assemble_revision, assemble_state_updated_at, order_readiness,
                   last_edited_by, last_assembled_by, last_modified_by
            FROM orders WHERE id = ?
            """,
            (order_id,),
        ).fetchone()
        if not row:
            con.close()
            return None
        if _is_lab_industries_client(str(row["client"] or "")) and not int(
            row["lab_sscc_shipped"] or 0
        ):
            _lab_sscc_sync_all_unshipped_seq_starts(cur)
            row = cur.execute(
                """
                SELECT id, ship_date, client, assembled_percent, names, extra_info, assemble_state,
                       buyer_order_mode, buyer_order, client_city, total_order_quantity,
                       lab_sscc_seq_start, lab_sscc_shipped, lab_sscc_pallet_count,
                       assemble_revision, assemble_state_updated_at, order_readiness,
                       last_edited_by, last_assembled_by, last_modified_by
                FROM orders WHERE id = ?
                """,
                (order_id,),
            ).fetchone()
        item_rows = cur.execute(
            """
            SELECT oi.product_id AS product_id, oi.article AS article, oi.name AS name,
                   oi.quantity AS quantity, oi.unit AS unit, oi.buyer_order AS buyer_order,
                   oi.total_order_quantity AS total_order_quantity,
                   p.pieces_in_box AS pieces_in_box,
                   p.sets_in_box AS sets_in_box,
                   p.pieces_per_set AS pieces_per_set,
                   p.row_layout AS row_layout,
                   p.max_rows AS max_rows,
                   p.box_weight AS box_weight,
                   p.volume_ml AS volume_ml,
                   p.volume_unit AS volume_unit
            FROM order_items oi
            LEFT JOIN products p ON p.id = oi.product_id
            WHERE oi.order_id = ?
            ORDER BY oi.id
            """,
            (order_id,),
        ).fetchall()
        con.close()
    items = [_order_item_row_to_dict(ir) for ir in item_rows]
    if not items and (row["names"] or "").strip():
        items = _synthetic_items_from_order_names(row["names"] or "")
    extra_fields = _order_list_fields_from_extra_info(
        row["extra_info"] or "", row["client"] or ""
    )
    return {
        "id": int(row["id"]),
        "ship_date": row["ship_date"] or "",
        "client": row["client"] or "",
        "assembled_percent": max(0, min(100, int(row["assembled_percent"] or 0))),
        "order_readiness": _order_readiness_from_row(row),
        "names": row["names"] or "",
        "assemble_state": _assemble_state_cell_to_api(row["assemble_state"]),
        "assemble_revision": max(0, int(row["assemble_revision"] or 0))
        if "assemble_revision" in row.keys()
        else 0,
        "assemble_state_updated_at": (row["assemble_state_updated_at"] or "").strip()
        if "assemble_state_updated_at" in row.keys()
        else "",
        "buyer_order_mode": (row["buyer_order_mode"] or "").strip()
        if "buyer_order_mode" in row.keys()
        else "",
        "buyer_order": (row["buyer_order"] or "").strip()
        if "buyer_order" in row.keys()
        else "",
        "client_city": (row["client_city"] or "").strip()
        if "client_city" in row.keys()
        else "",
        "total_order_quantity": float(row["total_order_quantity"])
        if row["total_order_quantity"] is not None
        and "total_order_quantity" in row.keys()
        else None,
        "lab_sscc_seq_start": max(1, int(row["lab_sscc_seq_start"]))
        if row["lab_sscc_seq_start"] is not None and "lab_sscc_seq_start" in row.keys()
        else None,
        "lab_sscc_shipped": bool(int(row["lab_sscc_shipped"] or 0))
        if "lab_sscc_shipped" in row.keys()
        else False,
        "lab_sscc_pallet_count": int(row["lab_sscc_pallet_count"])
        if row["lab_sscc_pallet_count"] is not None
        and "lab_sscc_pallet_count" in row.keys()
        else None,
        "lab_sscc_last_shipped": get_lab_sscc_last_shipped(),
        **_order_editor_fields_from_row(row),
        "items": items,
        **extra_fields,
    }


def delete_order(order_id: int) -> dict:
    """Удаляет заказ; позиции order_items удаляются каскадом.
    PDF/XLSX паллетных листов на сервере не кэшируются — генерируются в памяти по запросу, отдельных файлов под заказ нет."""
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        row = cur.execute(
            """
            SELECT client, order_readiness, COALESCE(lab_sscc_shipped, 0) AS lab_sscc_shipped
            FROM orders WHERE id = ?
            """,
            (order_id,),
        ).fetchone()
        if not row:
            con.close()
            return {"error": "not_found", "message": "Заказ не найден."}
        if _order_is_shipment_locked(row):
            con.close()
            return {
                "error": "order_shipped",
                "message": "Заказ отгружен — удаление недоступно.",
            }
        resync_lab = bool(
            _is_lab_industries_client(str(row["client"] or ""))
            and not int(row["lab_sscc_shipped"] or 0)
        )
        cur.execute("DELETE FROM orders WHERE id = ?", (order_id,))
        n = cur.rowcount
        if n > 0 and resync_lab:
            _lab_sscc_sync_all_unshipped_seq_starts(cur)
        con.commit()
        con.close()
    if n > 0:
        return {"ok": True}
    return {"error": "not_found", "message": "Заказ не найден."}


def fetch_orders(session: dict | None = None):
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        _lab_sscc_sync_all_unshipped_seq_starts(cur)
        rows = cur.execute(
            """
            SELECT id, ship_date, client, assembled_percent, names, extra_info, assemble_state,
                   client_city, lab_sscc_seq_start, lab_sscc_shipped, order_readiness,
                   assemble_revision, assemble_state_updated_at,
                   last_edited_by, last_assembled_by, last_modified_by
            FROM orders ORDER BY id DESC
            """
        ).fetchall()
        ids = [int(r["id"]) for r in rows]
        messages_map = _order_chat_has_messages_map(cur, ids) if ids else {}
        unread_map = _order_chat_unread_map(cur, session, ids) if session else {}
        items_by_oid = {oid: [] for oid in ids}
        if ids:
            placeholders = ",".join("?" * len(ids))
            item_rows = con.execute(
                f"""
                SELECT oi.order_id AS order_id, oi.product_id AS product_id,
                       oi.article AS article, oi.name AS name,
                       oi.quantity AS quantity, oi.unit AS unit,
                       p.pieces_in_box AS pieces_in_box,
                       p.sets_in_box AS sets_in_box,
                       p.pieces_per_set AS pieces_per_set,
                       p.row_layout AS row_layout,
                       p.max_rows AS max_rows,
                       p.box_weight AS box_weight,
                       p.volume_ml AS volume_ml,
                       p.volume_unit AS volume_unit
                FROM order_items oi
                LEFT JOIN products p ON p.id = oi.product_id
                WHERE oi.order_id IN ({placeholders})
                ORDER BY oi.order_id, oi.id
                """,
                ids,
            ).fetchall()
            for ir in item_rows:
                oid = int(ir["order_id"])
                if oid in items_by_oid:
                    items_by_oid[oid].append(_order_item_row_to_dict(ir))
        con.close()
    out = []
    for row in rows:
        oid = int(row["id"])
        items = items_by_oid.get(oid, [])
        if not items and (row["names"] or "").strip():
            items = _synthetic_items_from_order_names(row["names"] or "")
        extra_fields = _order_list_fields_from_extra_info(
            row["extra_info"] or "", row["client"] or ""
        )
        out.append(
            {
                "id": oid,
                "ship_date": row["ship_date"] or "",
                "client": row["client"] or "",
                "client_city": (row["client_city"] or "").strip()
                if "client_city" in row.keys()
                else "",
                "assembled_percent": max(0, min(100, int(row["assembled_percent"] or 0))),
                "order_readiness": _order_readiness_from_row(row),
                "names": row["names"] or "",
                "assemble_state": _assemble_state_cell_to_api(row["assemble_state"]),
                "assemble_revision": max(0, int(row["assemble_revision"] or 0))
                if "assemble_revision" in row.keys()
                else 0,
                "assemble_state_updated_at": (row["assemble_state_updated_at"] or "").strip()
                if "assemble_state_updated_at" in row.keys()
                else "",
                "lab_sscc_seq_start": max(1, int(row["lab_sscc_seq_start"]))
                if row["lab_sscc_seq_start"] is not None
                and "lab_sscc_seq_start" in row.keys()
                else None,
                "lab_sscc_shipped": bool(int(row["lab_sscc_shipped"] or 0))
                if "lab_sscc_shipped" in row.keys()
                else False,
                **_order_editor_fields_from_row(row),
                "items": items,
                "chat_has_messages": bool(messages_map.get(oid, False)),
                "chat_has_unread": bool(unread_map.get(oid, False)),
                **extra_fields,
            }
        )
    return out


def fetch_nomenclature():
    """Список номенклатуры; имена без хвоста «, шт» / «, набор», при расхождении — правка в БД."""
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        rows = cur.execute(
            """
            SELECT id, article, name,
                   pieces_in_box, sets_in_box, pieces_per_set, row_layout, max_rows, box_weight, volume_ml, volume_unit
            FROM products ORDER BY id
            """
        ).fetchall()
        updates: list[tuple[str, int]] = []
        out = []
        for row in rows:
            pib = int(row["pieces_in_box"] or 0)
            sib = max(1, int(row["sets_in_box"] or 1))
            pps = _computed_pieces_per_set_for_api(pib, sib)
            raw_name = row["name"] if row["name"] is not None else ""
            clean_name = _strip_trailing_name_suffix(raw_name)
            if clean_name != raw_name:
                updates.append((clean_name, int(row["id"])))
            out.append(
                {
                    "id": row["id"],
                    "article": row["article"] or "",
                    "name": clean_name,
                    "pieces_in_box": pib,
                    "sets_in_box": sib,
                    "pieces_per_set": pps,
                    "row_layout": int(row["row_layout"] or 0),
                    "max_rows": int(row["max_rows"] or 0),
                    "box_weight": float(row["box_weight"] or 0),
                    "volume_ml": float(row["volume_ml"] or 0),
                    "volume_unit": _normalize_volume_unit(
                        row["volume_unit"] if "volume_unit" in row.keys() else "ml"
                    ),
                }
            )
        for clean_name, rid in updates:
            cur.execute(
                "UPDATE products SET name = ? WHERE id = ?",
                (clean_name, rid),
            )
        if updates:
            con.commit()
        con.close()
    return out


def update_nomenclature_row(
    row_id: int,
    article: str,
    name: str,
    pieces_in_box: int = 0,
    sets_in_box: int = 1,
    row_layout: int = 0,
    max_rows: int = 0,
    box_weight: float = 0,
    volume_ml: float = 0,
    volume_unit: str = "ml",
):
    norm = _normalize_packaging(pieces_in_box, sets_in_box)
    if isinstance(norm, dict):
        return norm
    pib, sib, pps = norm
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        cur.execute(
            """
            UPDATE products SET
              article = ?, name = ?,
              pieces_in_box = ?, sets_in_box = ?, pieces_per_set = ?, row_layout = ?, max_rows = ?, box_weight = ?, volume_ml = ?, volume_unit = ?
            WHERE id = ?
            """,
            (
                article.strip(),
                _strip_trailing_name_suffix(name.strip()),
                pib,
                sib,
                pps,
                int(row_layout),
                int(max_rows),
                float(box_weight),
                max(0.0, float(volume_ml)),
                _normalize_volume_unit(volume_unit),
                row_id,
            ),
        )
        affected = cur.rowcount
        con.commit()
        con.close()
    return True if affected > 0 else False


def _normalize_updates_entry(raw, index: int) -> dict:
    if not isinstance(raw, dict):
        return {
            "error": "validation",
            "message": f"Запись #{index + 1}: ожидался объект.",
        }
    date = _normalize_str(str(raw.get("date", "")))
    if date:
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date):
            return {
                "error": "validation",
                "message": f"Запись #{index + 1}: дата должна быть в формате ГГГГ-ММ-ДД.",
            }
    title = _normalize_str(str(raw.get("title", "")))
    if not title:
        return {
            "error": "validation",
            "message": f"Запись #{index + 1}: укажите заголовок.",
        }
    lines_raw = raw.get("lines")
    if lines_raw is None:
        lines_raw = []
    if not isinstance(lines_raw, list):
        return {
            "error": "validation",
            "message": f"Запись #{index + 1}: пункты списка должны быть массивом строк.",
        }
    lines = [_normalize_str(str(line)) for line in lines_raw]
    lines = [line for line in lines if line]
    if not lines:
        return {
            "error": "validation",
            "message": f"Запись #{index + 1}: добавьте хотя бы один пункт.",
        }
    return {"date": date, "title": title, "lines": lines}


def _normalize_updates_payload(body) -> dict:
    if not isinstance(body, dict):
        return {"error": "validation", "message": "Ожидался JSON-объект."}
    entries_raw = body.get("entries")
    if not isinstance(entries_raw, list):
        return {"error": "validation", "message": "Поле entries должно быть массивом."}
    entries = []
    for idx, raw in enumerate(entries_raw):
        item = _normalize_updates_entry(raw, idx)
        if item.get("error"):
            return item
        entries.append(item)
    return {"entries": _sort_updates_entries(entries)}


def _updates_entry_date_sort_key(entry: dict) -> float:
    date = _normalize_str(str(entry.get("date", "")))
    if not date:
        return float("-inf")
    try:
        return datetime.datetime.strptime(date, "%Y-%m-%d").replace(
            tzinfo=datetime.timezone.utc
        ).timestamp()
    except ValueError:
        return float("-inf")


def _sort_updates_entries(entries: list) -> list:
    return sorted(
        entries,
        key=_updates_entry_date_sort_key,
        reverse=True,
    )


def fetch_updates() -> dict:
    if not UPDATES_JSON_PATH.exists():
        return {"entries": []}
    try:
        data = json.loads(UPDATES_JSON_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {"entries": []}
    if not isinstance(data, dict):
        return {"entries": []}
    entries_raw = data.get("entries")
    if not isinstance(entries_raw, list):
        return {"entries": []}
    entries = []
    for idx, raw in enumerate(entries_raw):
        item = _normalize_updates_entry(raw, idx)
        if item.get("error"):
            continue
        entries.append(item)
    return {"entries": _sort_updates_entries(entries)}


def save_updates(body: dict) -> dict:
    pack = _normalize_updates_payload(body)
    if pack.get("error"):
        return pack
    payload = {"entries": pack["entries"]}
    text = json.dumps(payload, ensure_ascii=False, indent=2) + "\n"
    tmp_path = UPDATES_JSON_PATH.with_suffix(".json.tmp")
    try:
        tmp_path.write_text(text, encoding="utf-8")
        tmp_path.replace(UPDATES_JSON_PATH)
    except OSError as exc:
        return {
            "error": "io",
            "message": f"Не удалось сохранить обновления: {exc}",
        }
    return {"ok": True, **payload}


def _utc_now_iso() -> str:
    return datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _feedback_author_name_from_session(session: dict | None) -> str:
    return _auth_display_name(session)


def _feedback_user_label_from_row(row) -> str:
    display_name = _normalize_str(row["display_name"] or "")
    if display_name:
        return display_name
    login = _normalize_str(row["login"] or "")
    if login:
        return login
    return "Пользователь"


def _feedback_row_is_closed(row) -> bool:
    if "is_closed" not in row.keys():
        return False
    return bool(int(row["is_closed"] or 0))


FEEDBACK_TASK_STATUSES = frozenset({"in_progress", "resolved", "rejected"})


def _feedback_row_task_status(row) -> str:
    if row and "task_status" in row.keys():
        status = _normalize_str(row["task_status"] or "")
        if status in FEEDBACK_TASK_STATUSES:
            return status
    if _feedback_row_is_closed(row):
        return "resolved"
    return "in_progress"


def _feedback_has_unread(row, *, is_admin: bool) -> bool:
    last_at = (row["last_message_at"] or "").strip() if "last_message_at" in row.keys() else ""
    if not last_at:
        return False
    last_author = int(row["last_author_user_id"] or 0) if "last_author_user_id" in row.keys() else 0
    if is_admin:
        if last_author == 0:
            return False
        read_at = (row["admin_last_read_at"] or "").strip() if "admin_last_read_at" in row.keys() else ""
    else:
        if last_author != 0:
            return False
        read_at = (row["user_last_read_at"] or "").strip() if "user_last_read_at" in row.keys() else ""
    if not read_at:
        return True
    return read_at < last_at


def _feedback_thread_row_to_list_item(row, *, is_admin: bool) -> dict:
    out = {
        "id": int(row["id"]),
        "subject": row["subject"] or "",
        "created_at": row["created_at"] or "",
        "updated_at": row["updated_at"] or "",
        "message_count": max(0, int(row["message_count"] or 0)),
        "task_status": _feedback_row_task_status(row),
        "has_unread": _feedback_has_unread(row, is_admin=is_admin),
    }
    if is_admin:
        out["user_id"] = int(row["user_id"] or 0)
        out["user_name"] = _feedback_user_label_from_row(row)
    return out


def _feedback_thread_list_select_sql(*, admin: bool) -> str:
    base = """
        SELECT t.id, t.user_id, t.subject, t.created_at, t.updated_at,
               t.is_closed, t.task_status, t.user_last_read_at, t.admin_last_read_at,
               (SELECT COUNT(*) FROM feedback_messages m WHERE m.thread_id = t.id) AS message_count,
               (SELECT author_user_id FROM feedback_messages m2
                WHERE m2.thread_id = t.id ORDER BY m2.id DESC LIMIT 1) AS last_author_user_id,
               (SELECT created_at FROM feedback_messages m3
                WHERE m3.thread_id = t.id ORDER BY m3.id DESC LIMIT 1) AS last_message_at
    """
    if admin:
        return (
            base
            + """,
               u.display_name, u.login
        FROM feedback_threads t
        LEFT JOIN app_users u ON u.id = t.user_id
        ORDER BY t.updated_at DESC, t.id DESC
        """
        )
    return (
        base
        + """
        FROM feedback_threads t
        WHERE t.user_id = ?
        ORDER BY t.updated_at DESC, t.id DESC
        """
    )


def _feedback_mark_thread_read(cur, thread_id: int, session: dict | None) -> None:
    if not session:
        return
    now = _utc_now_iso()
    if session.get("is_admin"):
        cur.execute(
            "UPDATE feedback_threads SET admin_last_read_at = ? WHERE id = ?",
            (now, int(thread_id)),
        )
    else:
        cur.execute(
            "UPDATE feedback_threads SET user_last_read_at = ? WHERE id = ?",
            (now, int(thread_id)),
        )


def fetch_feedback_threads(session: dict | None) -> dict:
    if not session:
        return {"error": "unauthorized", "message": "Требуется авторизация."}
    is_admin = bool(session.get("is_admin"))
    user_id = int(session.get("user_id") or 0)
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        if is_admin:
            rows = cur.execute(_feedback_thread_list_select_sql(admin=True)).fetchall()
        else:
            if user_id <= 0:
                con.close()
                return {"threads": []}
            rows = cur.execute(
                _feedback_thread_list_select_sql(admin=False),
                (user_id,),
            ).fetchall()
        con.close()
    threads = [_feedback_thread_row_to_list_item(r, is_admin=is_admin) for r in rows]
    return {"threads": threads}


def _feedback_can_access_thread(session: dict | None, thread_row) -> bool:
    if not session or not thread_row:
        return False
    if session.get("is_admin"):
        return True
    uid = int(session.get("user_id") or 0)
    return uid > 0 and int(thread_row["user_id"] or 0) == uid


def fetch_feedback_thread_detail(session: dict | None, thread_id: int) -> dict:
    if not session:
        return {"error": "unauthorized", "message": "Требуется авторизация."}
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        row = cur.execute(
            """
            SELECT t.id, t.user_id, t.subject, t.created_at, t.updated_at,
                   t.is_closed, t.task_status, t.user_last_read_at, t.admin_last_read_at,
                   u.display_name, u.login,
                   (SELECT author_user_id FROM feedback_messages m2
                    WHERE m2.thread_id = t.id ORDER BY m2.id DESC LIMIT 1) AS last_author_user_id,
                   (SELECT created_at FROM feedback_messages m3
                    WHERE m3.thread_id = t.id ORDER BY m3.id DESC LIMIT 1) AS last_message_at
            FROM feedback_threads t
            LEFT JOIN app_users u ON u.id = t.user_id
            WHERE t.id = ?
            """,
            (int(thread_id),),
        ).fetchone()
        if not row:
            con.close()
            return {"error": "not_found", "message": "Беседа не найдена."}
        if not _feedback_can_access_thread(session, row):
            con.close()
            return {"error": "forbidden", "message": "Нет доступа к этой беседе."}
        messages = cur.execute(
            """
            SELECT id, author_user_id, author_name, body, created_at
            FROM feedback_messages
            WHERE thread_id = ?
            ORDER BY id ASC
            """,
            (int(thread_id),),
        ).fetchall()
        _feedback_mark_thread_read(cur, int(thread_id), session)
        row = cur.execute(
            """
            SELECT t.id, t.user_id, t.subject, t.created_at, t.updated_at,
                   t.is_closed, t.task_status, t.user_last_read_at, t.admin_last_read_at,
                   u.display_name, u.login,
                   (SELECT author_user_id FROM feedback_messages m2
                    WHERE m2.thread_id = t.id ORDER BY m2.id DESC LIMIT 1) AS last_author_user_id,
                   (SELECT created_at FROM feedback_messages m3
                    WHERE m3.thread_id = t.id ORDER BY m3.id DESC LIMIT 1) AS last_message_at
            FROM feedback_threads t
            LEFT JOIN app_users u ON u.id = t.user_id
            WHERE t.id = ?
            """,
            (int(thread_id),),
        ).fetchone()
        con.commit()
        con.close()
    is_admin = bool(session.get("is_admin"))
    out = {
        "id": int(row["id"]),
        "subject": row["subject"] or "",
        "created_at": row["created_at"] or "",
        "updated_at": row["updated_at"] or "",
        "is_closed": _feedback_row_is_closed(row),
        "task_status": _feedback_row_task_status(row),
        "has_unread": _feedback_has_unread(row, is_admin=is_admin),
        "messages": [
            {
                "id": int(m["id"]),
                "author_user_id": int(m["author_user_id"] or 0),
                "author_name": m["author_name"] or "",
                "body": m["body"] or "",
                "created_at": m["created_at"] or "",
                "is_admin": int(m["author_user_id"] or 0) == 0,
            }
            for m in messages
        ],
    }
    if is_admin:
        out["user_id"] = int(row["user_id"] or 0)
        out["user_name"] = _feedback_user_label_from_row(row)
    return out


def create_feedback_thread(session: dict | None, body: dict) -> dict:
    if not session:
        return {"error": "unauthorized", "message": "Требуется авторизация."}
    if session.get("is_admin"):
        return {
            "error": "forbidden",
            "message": "Администратор не создаёт обращения — отвечайте в существующих беседах.",
        }
    user_id = int(session.get("user_id") or 0)
    if user_id <= 0:
        return {"error": "forbidden", "message": "Создание обращений недоступно."}
    subject = _normalize_str(body.get("subject", ""))
    message = _normalize_str(body.get("message", ""))
    if not subject:
        return {"error": "validation", "message": "Укажите тему беседы."}
    if len(subject) > 200:
        return {"error": "validation", "message": "Тема не длиннее 200 символов."}
    if not message:
        return {"error": "validation", "message": "Напишите сообщение."}
    if len(message) > 8000:
        return {"error": "validation", "message": "Сообщение слишком длинное."}
    author_name = _feedback_author_name_from_session(session)
    now = _utc_now_iso()
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        cur.execute(
            """
            INSERT INTO feedback_threads (user_id, subject, created_at, updated_at)
            VALUES (?, ?, ?, ?)
            """,
            (user_id, subject, now, now),
        )
        thread_id = int(cur.lastrowid)
        cur.execute(
            """
            INSERT INTO feedback_messages (
              thread_id, author_user_id, author_name, body, created_at
            ) VALUES (?, ?, ?, ?, ?)
            """,
            (thread_id, user_id, author_name, message, now),
        )
        cur.execute(
            """
            UPDATE feedback_threads
            SET user_last_read_at = ?, updated_at = ?
            WHERE id = ?
            """,
            (now, now, thread_id),
        )
        con.commit()
        con.close()
    return {"ok": True, "id": thread_id}


def add_feedback_message(session: dict | None, thread_id: int, body: dict) -> dict:
    if not session:
        return {"error": "unauthorized", "message": "Требуется авторизация."}
    message = _normalize_str(body.get("message", ""))
    if not message:
        return {"error": "validation", "message": "Напишите сообщение."}
    if len(message) > 8000:
        return {"error": "validation", "message": "Сообщение слишком длинное."}
    is_admin = bool(session.get("is_admin"))
    user_id = int(session.get("user_id") or 0)
    author_user_id = 0 if is_admin else user_id
    if not is_admin and user_id <= 0:
        return {"error": "forbidden", "message": "Отправка сообщений недоступна."}
    author_name = _feedback_author_name_from_session(session)
    now = _utc_now_iso()
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        row = cur.execute(
            "SELECT id, user_id FROM feedback_threads WHERE id = ?",
            (int(thread_id),),
        ).fetchone()
        if not row:
            con.close()
            return {"error": "not_found", "message": "Беседа не найдена."}
        if not _feedback_can_access_thread(session, row):
            con.close()
            return {"error": "forbidden", "message": "Нет доступа к этой беседе."}
        cur.execute(
            """
            INSERT INTO feedback_messages (
              thread_id, author_user_id, author_name, body, created_at
            ) VALUES (?, ?, ?, ?, ?)
            """,
            (int(thread_id), author_user_id, author_name, message, now),
        )
        msg_id = int(cur.lastrowid)
        if is_admin:
            cur.execute(
                """
                UPDATE feedback_threads
                SET updated_at = ?, admin_last_read_at = ?
                WHERE id = ?
                """,
                (now, now, int(thread_id)),
            )
        else:
            cur.execute(
                """
                UPDATE feedback_threads
                SET updated_at = ?, user_last_read_at = ?, task_status = 'in_progress'
                WHERE id = ?
                """,
                (now, now, int(thread_id)),
            )
        con.commit()
        con.close()
    return {
        "ok": True,
        "id": msg_id,
        "message": {
            "id": msg_id,
            "author_user_id": author_user_id,
            "author_name": author_name,
            "body": message,
            "created_at": now,
            "is_admin": is_admin,
        },
    }


def _order_chat_read_user_id(session: dict | None) -> int:
    if not session:
        return 0
    return int(session.get("user_id") or 0)


def _order_chat_author_name(session: dict | None) -> str:
    return _auth_display_name(session)


def _order_chat_is_own_message(
    author_user_id: int, author_name: str, session: dict | None
) -> bool:
    if not session:
        return False
    uid = _order_chat_read_user_id(session)
    aid = int(author_user_id or 0)
    if uid > 0 and aid > 0:
        return uid == aid
    if uid > 0:
        return uid == aid
    return (author_name or "").strip() == _order_chat_author_name(session)


def _order_chat_mark_read(cur, order_id: int, session: dict | None) -> None:
    if not session:
        return
    uid = _order_chat_read_user_id(session)
    now = _utc_now_iso()
    cur.execute(
        """
        INSERT INTO order_chat_reads (order_id, user_id, last_read_at)
        VALUES (?, ?, ?)
        ON CONFLICT(order_id, user_id) DO UPDATE SET
          last_read_at = excluded.last_read_at
        """,
        (int(order_id), uid, now),
    )


def _order_chat_has_messages_map(cur, order_ids: list[int]) -> dict[int, bool]:
    if not order_ids:
        return {}
    placeholders = ",".join("?" * len(order_ids))
    rows = cur.execute(
        f"""
        SELECT order_id, COUNT(*) AS cnt
        FROM order_messages
        WHERE order_id IN ({placeholders})
        GROUP BY order_id
        """,
        order_ids,
    ).fetchall()
    return {int(r["order_id"]): int(r["cnt"] or 0) > 0 for r in rows}


def _order_chat_unread_map(
    cur, session: dict | None, order_ids: list[int]
) -> dict[int, bool]:
    if not session or not order_ids:
        return {}
    uid = _order_chat_read_user_id(session)
    placeholders = ",".join("?" * len(order_ids))
    rows = cur.execute(
        f"""
        SELECT m.order_id, m.author_user_id, m.author_name, m.created_at,
               COALESCE(r.last_read_at, '') AS last_read_at
        FROM order_messages m
        INNER JOIN (
            SELECT order_id, MAX(id) AS max_id
            FROM order_messages
            WHERE order_id IN ({placeholders})
            GROUP BY order_id
        ) latest ON latest.max_id = m.id
        LEFT JOIN order_chat_reads r
          ON r.order_id = m.order_id AND r.user_id = ?
        """,
        (*order_ids, uid),
    ).fetchall()
    out: dict[int, bool] = {}
    for row in rows:
        oid = int(row["order_id"])
        if _order_chat_is_own_message(
            int(row["author_user_id"] or 0),
            row["author_name"] or "",
            session,
        ):
            out[oid] = False
            continue
        last_at = (row["created_at"] or "").strip()
        read_at = (row["last_read_at"] or "").strip()
        out[oid] = bool(last_at) and (not read_at or read_at < last_at)
    return out


def fetch_order_chat(session: dict | None, order_id: int) -> dict:
    if not session:
        return {"error": "unauthorized", "message": "Требуется авторизация."}
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        exists = cur.execute(
            "SELECT id FROM orders WHERE id = ?",
            (int(order_id),),
        ).fetchone()
        if not exists:
            con.close()
            return {"error": "not_found", "message": "Заказ не найден."}
        rows = cur.execute(
            """
            SELECT id, author_user_id, author_name, body, created_at
            FROM order_messages
            WHERE order_id = ?
            ORDER BY id ASC
            """,
            (int(order_id),),
        ).fetchall()
        _order_chat_mark_read(cur, int(order_id), session)
        con.commit()
        con.close()
    messages = [
        {
            "id": int(m["id"]),
            "author_user_id": int(m["author_user_id"] or 0),
            "author_name": m["author_name"] or "",
            "body": m["body"] or "",
            "created_at": m["created_at"] or "",
            "is_own": _order_chat_is_own_message(
                int(m["author_user_id"] or 0),
                m["author_name"] or "",
                session,
            ),
        }
        for m in rows
    ]
    return {
        "order_id": int(order_id),
        "messages": messages,
        "chat_has_unread": False,
    }


def add_order_message(session: dict | None, order_id: int, body: dict) -> dict:
    if not session:
        return {"error": "unauthorized", "message": "Требуется авторизация."}
    message = _normalize_str(body.get("message", ""))
    if not message:
        return {"error": "validation", "message": "Напишите сообщение."}
    if len(message) > 8000:
        return {"error": "validation", "message": "Сообщение слишком длинное."}
    author_user_id = _order_chat_read_user_id(session)
    author_name = _order_chat_author_name(session)
    now = _utc_now_iso()
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        exists = cur.execute(
            "SELECT id FROM orders WHERE id = ?",
            (int(order_id),),
        ).fetchone()
        if not exists:
            con.close()
            return {"error": "not_found", "message": "Заказ не найден."}
        cur.execute(
            """
            INSERT INTO order_messages (
              order_id, author_user_id, author_name, body, created_at
            ) VALUES (?, ?, ?, ?, ?)
            """,
            (int(order_id), author_user_id, author_name, message, now),
        )
        msg_id = int(cur.lastrowid)
        _order_chat_mark_read(cur, int(order_id), session)
        con.commit()
        con.close()
    return {
        "ok": True,
        "id": msg_id,
        "message": {
            "id": msg_id,
            "author_user_id": author_user_id,
            "author_name": author_name,
            "body": message,
            "created_at": now,
            "is_own": True,
        },
    }


def delete_order_message(
    session: dict | None, order_id: int, message_id: int
) -> dict:
    if not session:
        return {"error": "unauthorized", "message": "Требуется авторизация."}
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        row = cur.execute(
            """
            SELECT id, order_id, author_user_id, author_name
            FROM order_messages
            WHERE id = ? AND order_id = ?
            """,
            (int(message_id), int(order_id)),
        ).fetchone()
        if not row:
            con.close()
            return {"error": "not_found", "message": "Сообщение не найдено."}
        if not _order_chat_is_own_message(
            int(row["author_user_id"] or 0),
            row["author_name"] or "",
            session,
        ):
            con.close()
            return {
                "error": "forbidden",
                "message": "Можно удалить только своё сообщение.",
            }
        cur.execute(
            "DELETE FROM order_messages WHERE id = ? AND order_id = ?",
            (int(message_id), int(order_id)),
        )
        con.commit()
        con.close()
    return {"ok": True}


def patch_feedback_thread(session: dict | None, thread_id: int, body: dict) -> dict:
    if not session:
        return {"error": "unauthorized", "message": "Требуется авторизация."}
    if not session.get("is_admin"):
        return {
            "error": "forbidden",
            "message": "Менять статус задачи может только администратор.",
        }
    if not isinstance(body, dict) or "task_status" not in body:
        return {
            "error": "validation",
            "message": "Передайте task_status: in_progress, resolved или rejected.",
        }
    task_status = _normalize_str(body.get("task_status", ""))
    if task_status not in FEEDBACK_TASK_STATUSES:
        return {
            "error": "validation",
            "message": "Недопустимый статус задачи.",
        }
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        row = cur.execute(
            "SELECT id, user_id FROM feedback_threads WHERE id = ?",
            (int(thread_id),),
        ).fetchone()
        if not row:
            con.close()
            return {"error": "not_found", "message": "Беседа не найдена."}
        if not _feedback_can_access_thread(session, row):
            con.close()
            return {"error": "forbidden", "message": "Нет доступа к этой беседе."}
        cur.execute(
            "UPDATE feedback_threads SET task_status = ? WHERE id = ?",
            (task_status, int(thread_id)),
        )
        con.commit()
        con.close()
    return {"ok": True, "id": int(thread_id), "task_status": task_status}


def delete_nomenclature_row(row_id: int) -> bool:
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        cur.execute("DELETE FROM products WHERE id = ?", (int(row_id),))
        affected = cur.rowcount
        con.commit()
        con.close()
    return affected > 0


def _normalize_str(value: str) -> str:
    return (value or "").strip()


def find_nomenclature_create_conflicts(article: str, name: str):
    article_n = _normalize_str(article)
    name_n = _normalize_str(name)
    conflicts = []
    if not article_n or not name_n:
        return {
            "error": "validation",
            "message": "Укажите непустой артикул и наименование.",
        }
    art_cf = article_n.casefold()
    name_cf = name_n.casefold()
    for item in fetch_nomenclature():
        ea = _normalize_str(item["article"])
        en = _normalize_str(item["name"])
        if not ea and not en:
            continue
        row_id = item["id"]
        if art_cf == ea.casefold():
            conflicts.append(
                {
                    "kind": "article",
                    "you": article_n,
                    "existing_id": row_id,
                    "existing_article": ea,
                    "existing_name": en or "—",
                }
            )
        if name_cf == en.casefold():
            conflicts.append(
                {
                    "kind": "name",
                    "you": name_n,
                    "existing_id": row_id,
                    "existing_article": ea or "—",
                    "existing_name": en,
                }
            )
    if conflicts:
        return {"error": "duplicate", "conflicts": conflicts}
    return None


def insert_nomenclature_row(
    article: str,
    name: str,
    pieces_in_box: int,
    sets_in_box: int,
    row_layout: int,
    max_rows: int,
    box_weight: float,
    volume_ml: float = 0,
    volume_unit: str = "ml",
):
    article_n = _normalize_str(article)
    name_n = _strip_trailing_name_suffix(_normalize_str(name))
    conflict = find_nomenclature_create_conflicts(article_n, name_n)
    if conflict:
        return conflict
    norm = _normalize_packaging(pieces_in_box, sets_in_box)
    if isinstance(norm, dict):
        return norm
    pib, sib, pps = norm
    with DB_LOCK:
        con = get_connection()
        cur = con.cursor()
        next_id = cur.execute("SELECT COALESCE(MAX(id), 0) + 1 FROM products").fetchone()[0]
        cur.execute(
            """
            INSERT INTO products (
              id, article, name,
              pieces_in_box, sets_in_box, pieces_per_set, row_layout, max_rows, box_weight, volume_ml, volume_unit
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                int(next_id),
                article_n,
                name_n,
                pib,
                sib,
                pps,
                int(row_layout),
                int(max_rows),
                float(box_weight),
                max(0.0, float(volume_ml)),
                _normalize_volume_unit(volume_unit),
            ),
        )
        con.commit()
        con.close()
    return {"ok": True, "id": int(next_id)}


class ApiHandler(BaseHTTPRequestHandler):
    _auth_session: dict | None = None

    def _bearer_token(self) -> str | None:
        auth = self.headers.get("Authorization", "") or ""
        if auth.startswith("Bearer "):
            token = auth[7:].strip()
            return token or None
        return None

    def _client_ip(self) -> str:
        xff = (self.headers.get("X-Forwarded-For") or "").strip()
        if xff:
            return xff.split(",")[0].strip()
        xri = (self.headers.get("X-Real-IP") or "").strip()
        if xri:
            return xri
        return str(self.client_address[0] if self.client_address else "")

    def _require_shared_secret(self, expected: str) -> bool:
        if not expected:
            self._send_json(
                503,
                {
                    "error": "config",
                    "message": "Секрет не настроен на сервере.",
                },
            )
            return False
        provided = self._bearer_token()
        if not _secrets_match(provided, expected):
            self._send_json(
                403,
                {"error": "forbidden", "message": "Неверный секрет."},
            )
            return False
        return True

    def _ensure_authenticated(self) -> bool:
        path = urlparse(self.path).path
        if _is_public_api_path(path):
            return True
        session = resolve_auth_session(self._bearer_token())
        if not session:
            self._send_json(
                401,
                {"error": "unauthorized", "message": "Требуется вход в систему."},
            )
            return False
        self._auth_session = session
        return True

    def _handle_auth_login(self) -> None:
        try:
            body = self._read_json_body()
        except json.JSONDecodeError:
            self._send_json(400, {"error": "Invalid JSON"})
            return
        login_raw = body.get("login", "")
        ip = self._client_ip()
        banned = login_guard_status(ip, str(login_raw or ""))
        if banned:
            self._send_json(429, banned)
            return
        user = authenticate_app_login(login_raw, body.get("password", ""))
        if not user:
            ban_now = login_guard_register_failure(ip, str(login_raw or ""))
            if ban_now:
                self._send_json(429, ban_now)
                return
            self._send_json(
                401,
                {
                    "error": "invalid_credentials",
                    "message": "Неверный логин или пароль.",
                },
            )
            return
        login_guard_clear_success(ip, user.get("login") or str(login_raw or ""))
        token = create_auth_session(user)
        self._send_json(
            200,
            {
                "token": token,
                "user": {
                    "user_id": int(user.get("user_id", 0)),
                    "login": user["login"],
                    "display_name": user.get("display_name") or "",
                    "department": user.get("department") or "",
                    "is_admin": bool(user.get("is_admin")),
                    "permissions": user.get("permissions") or _admin_permissions(),
                },
            },
        )

    def _require_admin(self) -> bool:
        if (self._auth_session or {}).get("is_admin"):
            return True
        self._send_json(
            403,
            {
                "error": "forbidden",
                "message": "Доступ только для администратора.",
            },
        )
        return False

    def _session_permissions(self) -> dict:
        session = self._auth_session or {}
        if session.get("is_admin"):
            return _admin_permissions()
        uid = int(session.get("user_id") or 0)
        if uid > 0:
            db_perms = fetch_app_user_permissions(uid)
            if db_perms is not None:
                token = self._bearer_token()
                if token:
                    with _AUTH_SESSIONS_LOCK:
                        stored = _AUTH_SESSIONS.get(token)
                        if stored and int(stored.get("user_id") or 0) == uid:
                            stored["permissions"] = dict(db_perms)
                return db_perms
        perms = session.get("permissions")
        if isinstance(perms, dict):
            return {
                "orders": bool(perms.get("orders")),
                "nomenclature": bool(perms.get("nomenclature")),
                "manage_users": bool(perms.get("manage_users")),
                "feedback": bool(perms.get("feedback")),
                "order_monitoring": bool(perms.get("order_monitoring")),
            }
        return {
            "orders": True,
            "nomenclature": True,
            "manage_users": False,
            "feedback": True,
            "order_monitoring": False,
        }

    def _ensure_feedback_permission(self) -> bool:
        session = self._auth_session or {}
        if session.get("is_admin"):
            return True
        return self._require_permission(
            "feedback", "Нет доступа к жалобам и предложениям."
        )

    def _require_permission(self, key: str, message: str) -> bool:
        if self._session_permissions().get(key):
            return True
        self._send_json(403, {"error": "forbidden", "message": message})
        return False

    def _require_users_manager(self) -> bool:
        if (self._auth_session or {}).get("is_admin"):
            return True
        if self._session_permissions().get("manage_users"):
            return True
        self._send_json(
            403,
            {
                "error": "forbidden",
                "message": "Нет доступа к управлению пользователями.",
            },
        )
        return False

    def _handle_auth_me(self) -> None:
        session = self._auth_session or {}
        is_admin = bool(session.get("is_admin"))
        department = ""
        if not is_admin:
            uid = int(session.get("user_id") or 0)
            if uid > 0:
                department = fetch_app_user_department(uid)
        self._send_json(
            200,
            {
                "user_id": int(session.get("user_id") or 0),
                "login": session.get("login") or "",
                "display_name": session.get("display_name") or "",
                "department": department,
                "department_options": fetch_app_user_department_options(),
                "is_admin": is_admin,
                "permissions": self._session_permissions(),
            },
        )

    def _handle_auth_logout(self) -> None:
        revoke_auth_session(self._bearer_token())
        self._send_json(200, {"ok": True})

    def _ensure_path_permissions(self, path: str, method: str = "GET") -> bool:
        if _path_needs_orders_permission(path):
            session = self._auth_session or {}
            if not session.get("is_admin"):
                perms = self._session_permissions()
                if not perms.get("orders"):
                    if not (
                        perms.get("order_monitoring")
                        and _path_allowed_for_order_monitoring(path, method)
                    ):
                        if not self._require_permission(
                            "orders", "Нет доступа к заказам."
                        ):
                            return False
        if _path_needs_nomenclature_permission(path):
            if not self._require_permission(
                "nomenclature", "Нет доступа к номенклатуре."
            ):
                return False
        return True

    def _send_json(self, status: int, data):
        payload = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)

    def _read_json_body(self):
        length = int(self.headers.get("Content-Length", "0"))
        if length <= 0:
            return {}
        raw = self.rfile.read(length).decode("utf-8")
        return json.loads(raw) if raw else {}

    def _read_multipart_file_field(self, field_name: str = "file"):
        content_type = self.headers.get("Content-Type", "")
        if "multipart/form-data" not in content_type:
            return None, {
                "error": "validation",
                "message": "Ожидается загрузка файла (multipart/form-data).",
            }
        boundary = None
        for piece in content_type.split(";"):
            piece = piece.strip()
            if piece.startswith("boundary="):
                boundary = piece.split("=", 1)[1].strip().strip('"')
                break
        if not boundary:
            return None, {
                "error": "validation",
                "message": "Некорректная загрузка файла.",
            }
        length = int(self.headers.get("Content-Length", "0"))
        if length <= 0:
            return None, {"error": "validation", "message": "Файл не передан."}
        body = self.rfile.read(length)
        delimiter = ("--" + boundary).encode("ascii", "ignore")
        name_token = f'name="{field_name}"'.encode("ascii", "ignore")
        for part in body.split(delimiter):
            if not part or part in (b"--", b"--\r\n"):
                continue
            chunk = part
            if chunk.startswith(b"\r\n"):
                chunk = chunk[2:]
            if chunk.endswith(b"--"):
                chunk = chunk[:-2]
            if chunk.endswith(b"\r\n"):
                chunk = chunk[:-2]
            header_end = chunk.find(b"\r\n\r\n")
            if header_end < 0:
                continue
            headers = chunk[:header_end]
            if name_token not in headers:
                continue
            data = chunk[header_end + 4 :]
            if data.endswith(b"\r\n"):
                data = data[:-2]
            if not data:
                return None, {"error": "validation", "message": "Файл пустой."}
            return data, None
        return None, {"error": "validation", "message": "Файл не передан."}

    def _respond_lab_industries_pallet_sheets_pdf(self, body: dict):
        """POST JSON: { \"pallets\": [{ \"article\", \"pallet_number\" }, ...] } — ЛАБ Индастриз."""
        try:
            from packing_sheets_lab_industries import (  # noqa: PLC0415
                build_lab_industries_pallet_sheets_pdf_bytes,
                lab_pallet_pdf_error_message,
            )
        except ImportError:
            self._send_json(
                503,
                {"error": "module", "message": "Модуль паллетных листов ЛАБ недоступен."},
            )
            return
        pallets_raw = body.get("pallets")
        if not isinstance(pallets_raw, list):
            self._send_json(
                400,
                {
                    "error": "validation_pallets",
                    "message": lab_pallet_pdf_error_message("validation_pallets"),
                },
            )
            return
        lab_sscc_seq_start = 1
        order_id_raw = body.get("order_id")
        order_oid: int | None = None
        if order_id_raw is not None and str(order_id_raw).strip() != "":
            try:
                order_oid = int(order_id_raw)
            except (TypeError, ValueError):
                order_oid = None
        if order_oid is not None:
            with DB_LOCK:
                con = get_connection()
                cur = con.cursor()
                override: dict[int, int] = {}
                n_body = _lab_sscc_count_physical_pallets_from_pdf_rows(
                    pallets_raw if isinstance(pallets_raw, list) else []
                )
                if n_body > 0:
                    override[order_oid] = n_body
                _lab_sscc_sync_all_unshipped_seq_starts(cur, override)
                row = cur.execute(
                    "SELECT lab_sscc_seq_start FROM orders WHERE id = ?",
                    (order_oid,),
                ).fetchone()
                con.commit()
                con.close()
                if row and row[0] is not None:
                    lab_sscc_seq_start = max(1, int(row[0]))
        pallets_for_pdf = []
        for p in pallets_raw:
            if not isinstance(p, dict):
                continue
            row = dict(p)
            row["lab_sscc_seq_start"] = lab_sscc_seq_start
            if "lab_sscc_pallet_index" not in row:
                row["lab_sscc_pallet_index"] = len(pallets_for_pdf) + 1
            pallets_for_pdf.append(row)
        blob, err, err_detail = build_lab_industries_pallet_sheets_pdf_bytes(
            pallets_for_pdf
        )
        if err:
            status_map = {
                "validation_pallets": 400,
                "no_fpdf": 503,
                "no_barcode": 503,
                "no_font": 503,
                "pdf_build": 500,
                "barcode_fetch": 500,
            }
            status = status_map.get(err, 500)
            msg = err_detail or lab_pallet_pdf_error_message(err)
            self._send_json(status, {"error": err, "message": msg})
            return
        self.send_response(200)
        self.send_header("Content-Type", "application/pdf")
        self.send_header("Content-Length", str(len(blob)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(blob)

    def _respond_arnest_unirus_pallet_sheets_pdf(self, body: dict):
        """POST JSON: { \"pallets\": [...] } с полями article, batch_number, manufacturing_date_raw, expiry_date_raw
        (даты — 6 цифр ДДММГГ), опционально pallet_number для штрих-кода строки 14 (1500000+суффикс),
        либо только { \"page_count\": N } — пустые страницы A4."""
        blob, err, err_detail = build_arnest_unirus_pallet_sheets_pdf_bytes(body)
        if err:
            status_map = {
                "validation_pallets": 400,
                "validation_pallet": 400,
                "validation": 400,
                "no_fpdf": 503,
                "no_barcode": 503,
                "no_line2_font": 503,
                "pdf_build": 500,
                "barcode_fetch": 500,
            }
            status = status_map.get(err, 500)
            msg = err_detail or _arnest_pallet_pdf_error_message(err)
            self._send_json(status, {"error": err, "message": msg})
            return
        self.send_response(200)
        self.send_header("Content-Type", "application/pdf")
        self.send_header("Content-Length", str(len(blob)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(blob)

    def _respond_orders_batches_export_pdf(self, body: dict):
        """POST JSON: { \"orders\": [{ ship_date, client, lines: [{ name, batches }] }] }."""
        blob, err, err_detail = build_orders_batches_export_pdf_bytes(body)
        if err:
            status_map = {
                "validation": 400,
                "no_fpdf": 503,
                "no_line2_font": 503,
                "pdf_build": 500,
            }
            status = status_map.get(err, 500)
            msg = err_detail or _arnest_pallet_pdf_error_message(err)
            self._send_json(status, {"error": err, "message": msg})
            return
        self.send_response(200)
        self.send_header("Content-Type", "application/pdf")
        self.send_header("Content-Length", str(len(blob)))
        self.send_header(
            "Content-Disposition",
            'attachment; filename="partii-zakazov.pdf"',
        )
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(blob)

    def _respond_arnest_unirus_pallet_sheets_xlsx(self, body: dict):
        """POST JSON: { \"pallets\": [...] }; опционально ship_date, order_id — для шапки. Файл только в ответе, на диск не пишется."""
        blob, err, err_detail = build_arnest_unirus_pallet_sheets_xlsx_bytes(body)
        if err:
            status_map = {
                "validation_pallets": 400,
                "validation_pallet": 400,
                "no_openpyxl": 503,
                "xlsx_build": 500,
            }
            status = status_map.get(err, 500)
            msg = err_detail or _arnest_pallet_pdf_error_message(err)
            self._send_json(status, {"error": err, "message": msg})
            return
        self.send_response(200)
        self.send_header(
            "Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.send_header("Content-Length", str(len(blob)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(blob)

    def _respond_print_arnest_unirus_pallet_sheets_raw(self, body: dict):
        """Тот же JSON, что для PDF; сервер отправляет сформированный PDF на принтер (TCP JetDirect)."""
        blob, err, err_detail = build_arnest_unirus_pallet_sheets_pdf_bytes(body)
        if err:
            status_map = {
                "validation_pallets": 400,
                "validation_pallet": 400,
                "validation": 400,
                "no_fpdf": 503,
                "no_barcode": 503,
                "no_line2_font": 503,
                "pdf_build": 500,
                "barcode_fetch": 500,
            }
            status = status_map.get(err, 500)
            msg = err_detail or _arnest_pallet_pdf_error_message(err)
            self._send_json(status, {"error": err, "message": msg})
            return
        ok, info, nbytes = send_pdf_to_raw_lan_printer(blob)
        if not ok:
            self._send_json(
                502,
                {"error": "print_transport", "message": info or "Ошибка отправки на принтер"},
            )
            return
        self._send_json(
            200,
            {"ok": True, "printer": info, "sent_bytes": nbytes},
        )

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        if _is_auth_me_path(path):
            if not self._ensure_authenticated():
                return
            self._handle_auth_me()
            return
        if not self._ensure_authenticated():
            return
        if _is_updates_path(path):
            self._send_json(200, fetch_updates())
            return
        if _is_database_backup_path(path):
            if not self._require_admin():
                return
            blob, err = export_database_bytes()
            if err:
                status = 404 if err.get("error") == "not_found" else 500
                self._send_json(status, err)
                return
            stamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
            filename = f"paletlist-warehouse-{stamp}.db"
            self.send_response(200)
            self.send_header("Content-Type", "application/octet-stream")
            self.send_header("Content-Length", str(len(blob)))
            self.send_header(
                "Content-Disposition",
                f'attachment; filename="{filename}"',
            )
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            self.wfile.write(blob)
            return
        if _is_server_status_path(path):
            if not self._require_admin():
                return
            try:
                self._send_json(200, collect_server_status())
            except Exception as exc:  # noqa: BLE001
                self._send_json(
                    500,
                    {
                        "error": "server_status",
                        "message": f"Не удалось собрать показатели: {exc}",
                    },
                )
            return
        if _is_unique_clients_list_path(path):
            self._send_json(200, fetch_unique_clients())
            return
        if _is_feedback_threads_list_path(path):
            if not self._ensure_feedback_permission():
                return
            result = fetch_feedback_threads(self._auth_session)
            err = result.get("error")
            if err == "unauthorized":
                self._send_json(401, result)
                return
            self._send_json(200, result)
            return
        feedback_tid = _parse_feedback_thread_detail_id(path)
        if feedback_tid is not None:
            if not self._ensure_feedback_permission():
                return
            result = fetch_feedback_thread_detail(self._auth_session, feedback_tid)
            err = result.get("error")
            if err == "not_found":
                self._send_json(404, result)
                return
            if err == "forbidden":
                self._send_json(403, result)
                return
            if err == "unauthorized":
                self._send_json(401, result)
                return
            self._send_json(200, result)
            return
        if not self._ensure_path_permissions(path, "GET"):
            return
        packing_html_oid = _parse_orders_packing_sheets_html_id(path)
        if packing_html_oid is not None:
            detail = fetch_order_detail(packing_html_oid)
            if detail is None:
                self._send_json(404, {"error": "Not found"})
                return
            if build_generic_packing_sheets_html is None:
                self._send_json(
                    503,
                    {"error": "module", "message": "Модуль упаковочных листов недоступен."},
                )
                return
            html, err = build_generic_packing_sheets_html(detail)
            if err:
                known = {"arnest_client", "no_assembly", "no_pallets"}
                if err in known:
                    status_map = {"arnest_client": 400, "no_assembly": 400, "no_pallets": 400}
                    status = status_map.get(err, 400)
                    msg = (
                        generic_packing_error_message(err)
                        if generic_packing_error_message
                        else str(err)
                    )
                    self._send_json(status, {"error": err, "message": msg})
                    return
                self._send_json(500, {"error": "packing_sheet", "message": str(err)})
                return
            b = html.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(b)))
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            self.wfile.write(b)
            return
        packing_oid = _parse_orders_packing_sheets_pdf_id(path)
        if packing_oid is not None:
            detail = fetch_order_detail(packing_oid)
            if detail is None:
                self._send_json(404, {"error": "Not found"})
                return
            try:
                from packing_sheets_lab_industries import (  # noqa: PLC0415
                    build_lab_industries_pallet_sheets_pdf_from_order,
                    is_lab_industries_client,
                    lab_pallet_pdf_error_message,
                )
            except ImportError:
                build_lab_industries_pallet_sheets_pdf_from_order = None  # type: ignore
                is_lab_industries_client = lambda _c: False  # type: ignore
                lab_pallet_pdf_error_message = lambda c: str(c)  # type: ignore
            if (
                build_lab_industries_pallet_sheets_pdf_from_order is not None
                and is_lab_industries_client(str(detail.get("client") or ""))
            ):
                lab_sscc_seq_start = get_or_assign_lab_sscc_seq_start(packing_oid)
                detail["lab_sscc_seq_start"] = lab_sscc_seq_start
                pdf_blob, err, err_detail = (
                    build_lab_industries_pallet_sheets_pdf_from_order(
                        detail, lab_sscc_seq_start=lab_sscc_seq_start
                    )
                )
                if err:
                    status_map = {
                        "lab_client": 400,
                        "no_assembly": 400,
                        "no_pallets": 400,
                        "validation_pallets": 400,
                        "no_fpdf": 503,
                        "no_barcode": 503,
                        "no_font": 503,
                        "pdf_build": 500,
                        "barcode_fetch": 500,
                    }
                    status = status_map.get(err, 500)
                    msg = err_detail or lab_pallet_pdf_error_message(err)
                    self._send_json(status, {"error": err, "message": msg})
                    return
                self.send_response(200)
                self.send_header("Content-Type", "application/pdf")
                self.send_header("Content-Length", str(len(pdf_blob)))
                self.send_header("Cache-Control", "no-store")
                self.end_headers()
                self.wfile.write(pdf_blob)
                return
            if build_generic_packing_sheets_pdf_bytes is None:
                self._send_json(
                    503,
                    {"error": "module", "message": "Модуль упаковочных листов недоступен."},
                )
                return
            pdf_blob, err = build_generic_packing_sheets_pdf_bytes(detail)
            if err:
                known = {
                    "arnest_client",
                    "lab_client",
                    "no_assembly",
                    "no_pallets",
                    "no_pdf_engine",
                    "pdf_empty",
                }
                if err in known:
                    status_map = {
                        "arnest_client": 400,
                        "lab_client": 400,
                        "no_assembly": 400,
                        "no_pallets": 400,
                        "no_pdf_engine": 503,
                        "pdf_empty": 500,
                    }
                    status = status_map.get(err, 400)
                    msg = (
                        generic_packing_error_message(err)
                        if generic_packing_error_message
                        else str(err)
                    )
                    self._send_json(status, {"error": err, "message": msg})
                    return
                if err.startswith("pdf_render:"):
                    msg = (
                        generic_packing_error_message(err)
                        if generic_packing_error_message
                        else err
                    )
                    self._send_json(500, {"error": "pdf_render", "message": msg})
                    return
                self._send_json(500, {"error": "packing_sheet", "message": str(err)})
                return
            self.send_response(200)
            self.send_header("Content-Type", "application/pdf")
            self.send_header("Content-Length", str(len(pdf_blob)))
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            self.wfile.write(pdf_blob)
            return
        oid = _parse_orders_detail_id(path)
        if oid is not None:
            detail = fetch_order_detail(oid)
            if detail is None:
                self._send_json(404, {"error": "Not found"})
                return
            self._send_json(200, detail)
            return
        assemble_sync_oid = _parse_orders_assemble_sync_id(path)
        if assemble_sync_oid is not None:
            qs = parse_qs(parsed.query)
            since_rev = None
            if "since_rev" in qs:
                raw_rev = (qs.get("since_rev") or [""])[0]
                if raw_rev not in ("", None):
                    try:
                        since_rev = int(raw_rev)
                    except (TypeError, ValueError):
                        self._send_json(
                            400,
                            {
                                "error": "validation",
                                "message": "since_rev: нужно целое число.",
                            },
                        )
                        return
            sync = fetch_order_assemble_sync_fields(assemble_sync_oid, since_rev)
            if sync is None:
                self._send_json(404, {"error": "not_found", "message": "Заказ не найден."})
                return
            self._send_json(200, {"ok": True, "id": int(assemble_sync_oid), **sync})
            return
        order_chat_oid = _parse_order_messages_id(path)
        if order_chat_oid is not None:
            if not self._ensure_path_permissions(path, "GET"):
                return
            result = fetch_order_chat(self._auth_session, order_chat_oid)
            err = result.get("error")
            if err == "not_found":
                self._send_json(404, result)
                return
            if err == "unauthorized":
                self._send_json(401, result)
                return
            self._send_json(200, result)
            return
        if _is_orders_list_path(path):
            if not self._ensure_path_permissions(path, "GET"):
                return
            self._send_json(200, fetch_orders(session=self._auth_session))
            return
        if _is_users_list_path(path):
            if not self._require_users_manager():
                return
            self._send_json(200, {"users": fetch_app_users()})
            return
        if _is_pallet_printer_raw_ping_path(path):
            ok, addr, msg = check_raw_pallet_printer_reachable()
            self._send_json(
                200,
                {"reachable": ok, "printer": addr, "message": msg},
            )
            return
        if not _is_nomenclature_list_path(path):
            self._send_json(404, {"error": "Not found"})
            return
        items = fetch_nomenclature()
        qs = parse_qs(parsed.query)
        if "q" in qs:
            query = (qs.get("q", [""])[0] or "").strip()
            try:
                limit = max(1, min(50, int((qs.get("limit", ["5"]) or ["5"])[0])))
            except (TypeError, ValueError):
                limit = 5
            if query:
                items = search_nomenclature(items, query, limit)
            else:
                items = []
        self._send_json(200, items)

    def do_POST(self):
        parsed = urlparse(self.path)
        path = parsed.path
        if _is_auth_login_path(path):
            self._handle_auth_login()
            return
        if _is_deploy_path(path):
            if not self._require_shared_secret(_deploy_secret()):
                return
            # Body optional; drain if present
            try:
                self._read_json_body()
            except json.JSONDecodeError:
                pass
            result = start_background_deploy()
            err = result.get("error")
            if err:
                self._send_json(500, result)
                return
            self._send_json(202, result)
            return
        if _is_ssh_allow_path(path):
            if not self._require_shared_secret(_ssh_allow_secret()):
                return
            try:
                body = self._read_json_body()
            except json.JSONDecodeError:
                body = {}
            ip = _normalize_str((body or {}).get("ip") or "") or self._client_ip()
            result = allow_ssh_ip(ip)
            err = result.get("error")
            if err == "validation":
                self._send_json(400, result)
                return
            if err:
                self._send_json(500, result)
                return
            self._send_json(200, result)
            return
        if not self._ensure_authenticated():
            return
        if _is_database_restore_path(path):
            if not self._require_admin():
                return
            length = int(self.headers.get("Content-Length", "0") or 0)
            if length > _DB_RESTORE_MAX_BYTES:
                self._send_json(
                    400,
                    {
                        "error": "validation",
                        "message": "Файл слишком большой (лимит 200 МБ).",
                    },
                )
                return
            data, read_err = self._read_multipart_file_field("file")
            if read_err:
                self._send_json(400, read_err)
                return
            result = restore_database_from_bytes(data)
            err = result.get("error")
            if err == "validation":
                self._send_json(400, result)
                return
            if err == "io":
                self._send_json(500, result)
                return
            if err == "database":
                self._send_json(500, result)
                return
            self._send_json(200, result)
            return
        if _is_feedback_threads_list_path(path):
            if not self._ensure_feedback_permission():
                return
            try:
                body = self._read_json_body()
            except json.JSONDecodeError:
                self._send_json(400, {"error": "Invalid JSON"})
                return
            try:
                result = create_feedback_thread(self._auth_session, body)
            except sqlite3.Error as exc:
                self._send_json(500, {"error": "database", "message": str(exc)})
                return
            err = result.get("error")
            if err == "validation":
                self._send_json(400, result)
                return
            if err == "forbidden":
                self._send_json(403, result)
                return
            self._send_json(201, result)
            return
        feedback_msg_tid = _parse_feedback_thread_messages_id(path)
        if feedback_msg_tid is not None:
            if not self._ensure_feedback_permission():
                return
            try:
                body = self._read_json_body()
            except json.JSONDecodeError:
                self._send_json(400, {"error": "Invalid JSON"})
                return
            try:
                result = add_feedback_message(
                    self._auth_session, feedback_msg_tid, body
                )
            except sqlite3.Error as exc:
                self._send_json(500, {"error": "database", "message": str(exc)})
                return
            err = result.get("error")
            if err == "validation":
                self._send_json(400, result)
                return
            if err == "not_found":
                self._send_json(404, result)
                return
            if err == "forbidden":
                self._send_json(403, result)
                return
            if err == "closed":
                self._send_json(400, result)
                return
            self._send_json(201, result)
            return
        order_chat_oid = _parse_order_messages_id(path)
        if order_chat_oid is not None:
            if not self._ensure_path_permissions(path, "POST"):
                return
            try:
                body = self._read_json_body()
            except json.JSONDecodeError:
                self._send_json(400, {"error": "Invalid JSON"})
                return
            try:
                result = add_order_message(
                    self._auth_session, order_chat_oid, body
                )
            except sqlite3.Error as exc:
                self._send_json(500, {"error": "database", "message": str(exc)})
                return
            err = result.get("error")
            if err == "validation":
                self._send_json(400, result)
                return
            if err == "not_found":
                self._send_json(404, result)
                return
            if err == "unauthorized":
                self._send_json(401, result)
                return
            self._send_json(201, result)
            return
        if not self._ensure_path_permissions(path, "POST"):
            return
        if _is_print_arnest_unirus_pallet_sheets_raw_path(path):
            try:
                body = self._read_json_body()
            except json.JSONDecodeError:
                self._send_json(400, {"error": "Invalid JSON"})
                return
            self._respond_print_arnest_unirus_pallet_sheets_raw(body)
            return
        if _is_arnest_unirus_pallet_sheets_xlsx_path(path):
            try:
                body = self._read_json_body()
            except json.JSONDecodeError:
                self._send_json(400, {"error": "Invalid JSON"})
                return
            self._respond_arnest_unirus_pallet_sheets_xlsx(body)
            return
        if _is_arnest_unirus_pallet_sheets_pdf_path(path):
            try:
                body = self._read_json_body()
            except json.JSONDecodeError:
                self._send_json(400, {"error": "Invalid JSON"})
                return
            self._respond_arnest_unirus_pallet_sheets_pdf(body)
            return
        if _is_lab_industries_pallet_sheets_pdf_path(path):
            try:
                body = self._read_json_body()
            except json.JSONDecodeError:
                self._send_json(400, {"error": "Invalid JSON"})
                return
            self._respond_lab_industries_pallet_sheets_pdf(body)
            return
        if _is_lab_sscc_reset_path(path):
            try:
                body = self._read_json_body() or {}
            except json.JSONDecodeError:
                self._send_json(400, {"error": "Invalid JSON"})
                return
            if not isinstance(body, dict):
                body = {}
            try:
                last_raw = body.get("last_shipped")
                next_raw = body.get("next_pallet")
                last_arg = int(last_raw) if last_raw is not None else None
                next_arg = int(next_raw) if next_raw is not None else None
            except (TypeError, ValueError):
                self._send_json(
                    400,
                    {
                        "error": "validation",
                        "message": "last_shipped и next_pallet — целые числа.",
                    },
                )
                return
            if last_arg is None and next_arg is None:
                next_arg = 12
            try:
                result = reset_lab_sscc_pallet_counter(
                    last_shipped=last_arg,
                    next_pallet=next_arg,
                )
            except sqlite3.Error as exc:
                self._send_json(500, {"error": "database", "message": str(exc)})
                return
            self._send_json(200, result)
            return
        if _is_orders_batches_export_pdf_path(path):
            try:
                body = self._read_json_body()
            except json.JSONDecodeError:
                self._send_json(400, {"error": "Invalid JSON"})
                return
            self._respond_orders_batches_export_pdf(body)
            return
        if _is_orders_import_excel_path(path):
            data, read_err = self._read_multipart_file_field("file")
            if read_err:
                self._send_json(400, read_err)
                return
            try:
                result = import_orders_from_excel_bytes(
                    data, modified_by=_auth_display_name(self._auth_session)
                )
            except sqlite3.Error as exc:
                self._send_json(500, {"error": "database", "message": str(exc)})
                return
            err = result.get("error")
            if err == "no_openpyxl":
                self._send_json(503, result)
                return
            if err in ("validation", "invalid_file"):
                self._send_json(400, result)
                return
            self._send_json(201, result)
            return
        if _is_orders_import_drogeri_excel_path(path):
            data, read_err = self._read_multipart_file_field("file")
            if read_err:
                self._send_json(400, read_err)
                return
            try:
                result = import_drogeri_order_from_excel_bytes(
                    data, modified_by=_auth_display_name(self._auth_session)
                )
            except sqlite3.Error as exc:
                self._send_json(500, {"error": "database", "message": str(exc)})
                return
            err = result.get("error")
            if err == "no_openpyxl":
                self._send_json(503, result)
                return
            if err in ("validation", "invalid_file"):
                self._send_json(400, result)
                return
            self._send_json(201, result)
            return
        presence_oid = _parse_orders_assemble_presence_id(path)
        if presence_oid is not None:
            try:
                body = self._read_json_body()
            except json.JSONDecodeError:
                self._send_json(400, {"error": "Invalid JSON"})
                return
            ua = self.headers.get("User-Agent", "")
            body = body if isinstance(body, dict) else {}
            since_rev = None
            include_sync = "since_rev" in body
            if include_sync:
                raw_rev = body.get("since_rev")
                if raw_rev is None or raw_rev == "":
                    since_rev = None
                else:
                    try:
                        since_rev = int(raw_rev)
                    except (TypeError, ValueError):
                        self._send_json(
                            400,
                            {
                                "error": "validation",
                                "message": "since_rev: нужно целое число или null.",
                            },
                        )
                        return
            result = touch_assemble_presence(
                presence_oid,
                body.get("client_id", ""),
                ua,
                since_rev=since_rev,
                include_assemble_sync=include_sync,
            )
            err = result.get("error")
            if err == "validation":
                self._send_json(400, result)
                return
            if err == "not_found":
                self._send_json(404, result)
                return
            self._send_json(200, result)
            return
        lab_ship_oid = _parse_orders_lab_ship_id(path)
        if lab_ship_oid is not None:
            try:
                result = confirm_lab_order_shipment(lab_ship_oid)
            except sqlite3.Error as exc:
                self._send_json(500, {"error": "database", "message": str(exc)})
                return
            err = result.get("error")
            if err == "not_found":
                self._send_json(404, result)
                return
            if err in ("lab_client", "no_pallets", "already_shipped"):
                self._send_json(400, result)
                return
            self._send_json(200, result)
            return
        if _is_users_list_path(path):
            if not self._require_users_manager():
                return
            try:
                body = self._read_json_body()
            except json.JSONDecodeError:
                self._send_json(400, {"error": "Invalid JSON"})
                return
            try:
                result = create_app_user(body)
            except sqlite3.Error as exc:
                self._send_json(500, {"error": "database", "message": str(exc)})
                return
            err = result.get("error")
            if err == "validation":
                self._send_json(400, result)
                return
            if err == "duplicate":
                self._send_json(409, result)
                return
            self._send_json(201, result)
            return
        if _is_orders_list_path(path):
            try:
                body = self._read_json_body()
            except json.JSONDecodeError:
                self._send_json(400, {"error": "Invalid JSON"})
                return
            try:
                result = insert_order_with_items(
                    body, modified_by=_auth_display_name(self._auth_session)
                )
            except sqlite3.Error as exc:
                self._send_json(500, {"error": "database", "message": str(exc)})
                return
            except Exception as exc:
                self._send_json(
                    500,
                    {
                        "error": "server",
                        "message": f"Ошибка сохранения заказа: {exc}",
                    },
                )
                return
            err = result.get("error")
            if err == "validation":
                self._send_json(400, result)
                return
            self._send_json(201, result)
            return
        if not _is_nomenclature_list_path(parsed.path):
            self._send_json(
                404,
                {
                    "error": "Not found",
                    "message": "Неизвестный POST-маршрут. Обновите код api_server и выполните sudo systemctl restart paletlist-api.",
                },
            )
            return
        try:
            body = self._read_json_body()
        except json.JSONDecodeError:
            self._send_json(400, {"error": "Invalid JSON"})
            return

        def _int_body(key: str) -> int:
            try:
                return int(body.get(key, 0))
            except (TypeError, ValueError):
                return 0

        def _float_body(key: str) -> float:
            try:
                return float(body.get(key, 0))
            except (TypeError, ValueError):
                return 0.0

        def _sets_in_box_body() -> int:
            try:
                v = int(body.get("sets_in_box", 1))
            except (TypeError, ValueError):
                return 1
            return max(1, v)

        try:
            result = insert_nomenclature_row(
                body.get("article", ""),
                body.get("name", ""),
                _int_body("pieces_in_box"),
                _sets_in_box_body(),
                _int_body("row_layout"),
                _int_body("max_rows"),
                _float_body("box_weight"),
                max(0.0, _float_body("volume_ml")),
                body.get("volume_unit", "ml"),
            )
        except sqlite3.Error as exc:
            self._send_json(
                500,
                {"error": "database", "message": str(exc)},
            )
            return
        err = result.get("error")
        if err == "validation":
            self._send_json(400, result)
            return
        if err == "duplicate":
            self._send_json(409, result)
            return
        self._send_json(201, result)

    def do_PUT(self):
        if not self._ensure_authenticated():
            return
        parsed = urlparse(self.path)
        path = parsed.path
        if _is_updates_path(path):
            if not self._require_admin():
                return
            try:
                body = self._read_json_body()
            except json.JSONDecodeError:
                self._send_json(400, {"error": "Invalid JSON"})
                return
            result = save_updates(body)
            err = result.get("error")
            if err == "validation":
                self._send_json(400, result)
                return
            if err == "io":
                self._send_json(500, result)
                return
            self._send_json(200, result)
            return
        if not _is_auth_account_path(path) and not _parse_user_id_path(path):
            if not self._ensure_path_permissions(path, "PUT"):
                return
        if _is_auth_account_path(path):
            try:
                body = self._read_json_body()
            except json.JSONDecodeError:
                self._send_json(400, {"error": "Invalid JSON"})
                return
            try:
                result = update_auth_account(
                    self._auth_session or {},
                    body,
                    self._bearer_token(),
                )
            except sqlite3.Error as exc:
                self._send_json(500, {"error": "database", "message": str(exc)})
                return
            err = result.get("error")
            if err == "validation":
                self._send_json(400, result)
                return
            if err == "duplicate":
                self._send_json(409, result)
                return
            if err == "forbidden":
                self._send_json(403, result)
                return
            if err == "not_found":
                self._send_json(404, result)
                return
            self._send_json(200, result)
            return
        user_id = _parse_user_id_path(path)
        if user_id is not None:
            if not self._require_users_manager():
                return
            try:
                body = self._read_json_body()
            except json.JSONDecodeError:
                self._send_json(400, {"error": "Invalid JSON"})
                return
            try:
                result = update_app_user(user_id, body)
            except sqlite3.Error as exc:
                self._send_json(500, {"error": "database", "message": str(exc)})
                return
            err = result.get("error")
            if err == "validation":
                self._send_json(400, result)
                return
            if err == "duplicate":
                self._send_json(409, result)
                return
            if err == "forbidden":
                self._send_json(403, result)
                return
            if err == "not_found":
                self._send_json(404, result)
                return
            self._send_json(200, result)
            return
        oid = _parse_orders_detail_id(path)
        if oid is not None:
            try:
                body = self._read_json_body()
            except json.JSONDecodeError:
                self._send_json(400, {"error": "Invalid JSON"})
                return
            try:
                result = update_order_with_items(
                    oid, body, modified_by=_auth_display_name(self._auth_session)
                )
            except sqlite3.Error as exc:
                self._send_json(500, {"error": "database", "message": str(exc)})
                return
            except Exception as exc:
                self._send_json(
                    500,
                    {
                        "error": "server",
                        "message": f"Ошибка сохранения заказа: {exc}",
                    },
                )
                return
            err = result.get("error")
            if err == "validation":
                self._send_json(400, result)
                return
            if err == "not_found":
                self._send_json(404, result)
                return
            if err == "order_shipped":
                self._send_json(403, result)
                return
            self._send_json(200, result)
            return
        if not parsed.path.startswith("/api/nomenclature/"):
            self._send_json(404, {"error": "Not found"})
            return
        try:
            row_id = int(parsed.path.rsplit("/", 1)[-1])
        except ValueError:
            self._send_json(400, {"error": "Invalid id"})
            return
        try:
            body = self._read_json_body()
        except json.JSONDecodeError:
            self._send_json(400, {"error": "Invalid JSON"})
            return
        def _int_body(key: str) -> int:
            try:
                return int(body.get(key, 0))
            except (TypeError, ValueError):
                return 0

        def _float_body(key: str) -> float:
            try:
                return float(body.get(key, 0))
            except (TypeError, ValueError):
                return 0.0

        def _sets_in_box_body() -> int:
            try:
                v = int(body.get("sets_in_box", 1))
            except (TypeError, ValueError):
                return 1
            return max(1, v)

        ok = update_nomenclature_row(
            row_id,
            body.get("article", ""),
            body.get("name", ""),
            _int_body("pieces_in_box"),
            _sets_in_box_body(),
            _int_body("row_layout"),
            _int_body("max_rows"),
            _float_body("box_weight"),
            max(0.0, _float_body("volume_ml")),
            body.get("volume_unit", "ml"),
        )
        if isinstance(ok, dict) and ok.get("error") == "validation":
            self._send_json(400, ok)
            return
        if not ok:
            self._send_json(404, {"error": "Item not found"})
            return
        self._send_json(200, {"ok": True})

    def do_PATCH(self):
        if not self._ensure_authenticated():
            return
        parsed = urlparse(self.path)
        path = parsed.path
        feedback_tid = _parse_feedback_thread_detail_id(path)
        if feedback_tid is not None:
            if not self._ensure_feedback_permission():
                return
            try:
                body = self._read_json_body()
            except json.JSONDecodeError:
                self._send_json(400, {"error": "Invalid JSON"})
                return
            try:
                result = patch_feedback_thread(
                    self._auth_session, feedback_tid, body
                )
            except sqlite3.Error as exc:
                self._send_json(500, {"error": "database", "message": str(exc)})
                return
            err = result.get("error")
            if err == "validation":
                self._send_json(400, result)
                return
            if err == "not_found":
                self._send_json(404, result)
                return
            if err == "forbidden":
                self._send_json(403, result)
                return
            self._send_json(200, result)
            return
        if not self._ensure_path_permissions(path, "PUT"):
            return
        oid = _parse_orders_detail_id(path)
        if oid is None:
            self._send_json(404, {"error": "Not found"})
            return
        try:
            body = self._read_json_body()
        except json.JSONDecodeError:
            self._send_json(400, {"error": "Invalid JSON"})
            return
        try:
            result = patch_order_assembly(
                oid, body, modified_by=_auth_display_name(self._auth_session)
            )
        except sqlite3.Error as exc:
            self._send_json(500, {"error": "database", "message": str(exc)})
            return
        err = result.get("error")
        if err == "validation":
            self._send_json(400, result)
            return
        if err == "not_found":
            self._send_json(404, result)
            return
        if err == "conflict":
            self._send_json(409, result)
            return
        if err == "order_shipped":
            self._send_json(403, result)
            return
        self._send_json(200, result)

    def do_DELETE(self):
        if not self._ensure_authenticated():
            return
        parsed = urlparse(self.path)
        path = parsed.path
        user_id = _parse_user_id_path(path)
        if user_id is None and not self._ensure_path_permissions(path, "DELETE"):
            return
        if user_id is not None:
            if not self._require_users_manager():
                return
            try:
                result = delete_app_user(user_id)
            except sqlite3.Error as exc:
                self._send_json(500, {"error": "database", "message": str(exc)})
                return
            err = result.get("error")
            if err == "forbidden":
                self._send_json(403, result)
                return
            if err == "not_found":
                self._send_json(404, result)
                return
            self._send_json(200, result)
            return
        presence_oid = _parse_orders_assemble_presence_id(path)
        if presence_oid is not None:
            try:
                body = self._read_json_body()
            except json.JSONDecodeError:
                body = {}
            result = leave_assemble_presence(
                presence_oid,
                (body or {}).get("client_id", ""),
            )
            err = result.get("error")
            if err == "validation":
                self._send_json(400, result)
                return
            self._send_json(200, result)
            return
        order_msg_ids = _parse_order_message_detail_id(path)
        if order_msg_ids is not None:
            order_id, message_id = order_msg_ids
            try:
                result = delete_order_message(
                    self._auth_session, order_id, message_id
                )
            except sqlite3.Error as exc:
                self._send_json(500, {"error": "database", "message": str(exc)})
                return
            err = result.get("error")
            if err == "not_found":
                self._send_json(404, result)
                return
            if err == "forbidden":
                self._send_json(403, result)
                return
            if err == "unauthorized":
                self._send_json(401, result)
                return
            self._send_json(200, result)
            return
        oid = _parse_orders_detail_id(path)
        if oid is not None:
            try:
                result = delete_order(oid)
            except sqlite3.Error as exc:
                self._send_json(500, {"error": "database", "message": str(exc)})
                return
            err = result.get("error")
            if err == "not_found":
                self._send_json(404, result)
                return
            if err == "order_shipped":
                self._send_json(403, result)
                return
            self._send_json(200, result)
            return
        if not parsed.path.startswith("/api/nomenclature/"):
            self._send_json(404, {"error": "Not found"})
            return
        try:
            row_id = int(parsed.path.rsplit("/", 1)[-1])
        except ValueError:
            self._send_json(400, {"error": "Invalid id"})
            return
        try:
            ok = delete_nomenclature_row(row_id)
        except sqlite3.Error as exc:
            self._send_json(500, {"error": "database", "message": str(exc)})
            return
        if not ok:
            self._send_json(404, {"error": "Item not found"})
            return
        self._send_json(200, {"ok": True})

    def log_message(self, format, *args):
        return


def main():
    server = ThreadingHTTPServer((HOST, PORT), ApiHandler)
    print(f"API server listening on http://{HOST}:{PORT}")
    server.serve_forever()


init_db()

if __name__ == "__main__":
    main()
